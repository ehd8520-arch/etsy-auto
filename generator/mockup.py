"""
Mockup Generator -- Pillow로 목업 이미지 10장 자동 생성.

Top seller strategy:
- Image 1: Hero mockup (thumbnail, must grab attention)
- Image 2: Lifestyle mockup (art on wall in living room)
- Image 3-4: Detail / page preview
- Image 5: "What's Included" overview
- Image 6-7: Different room settings
- Image 8: Size guide
- Image 9: Customer review screenshot (optional)
- Image 10: Brand CTA

We generate these programmatically using Pillow compositing.
"""
import io
import logging
import os
import random
import time
from pathlib import Path
from typing import Optional

import requests
from PIL import Image, ImageDraw, ImageFont, ImageFilter

from config.settings import get_next_cloudflare_account, CLOUDFLARE_ACCOUNTS, BASE_DIR
from models import Product, Category

# ── 폰트 디렉토리 ──
_FONTS_DIR = BASE_DIR / "assets" / "fonts"

_FONT_MAP = {
    "Regular":   "Poppins-Regular.ttf",
    "Bold":      "Poppins-Bold.ttf",
    "SemiBold":  "Poppins-SemiBold.ttf",
    "Medium":    "Poppins-Medium.ttf",
    "Light":     "Poppins-Light.ttf",
    "Display":   "PlayfairDisplay-Bold.ttf",
}

def _font(weight: str = "Regular", size: int = 30) -> ImageFont.FreeTypeFont:
    """Poppins / Playfair 폰트 로더.
    항상 같은 폰트 → 언제 실행해도 일관된 퀄리티 보장.
    파일 없으면 arial → load_default 순으로 폴백.
    """
    fname = _FONT_MAP.get(weight, "Poppins-Regular.ttf")
    for candidate in (str(_FONTS_DIR / fname), "arial.ttf", "arialbd.ttf"):
        try:
            return ImageFont.truetype(candidate, size)
        except OSError:
            continue
    return ImageFont.load_default()


def _slight_tilt(img: Image.Image, angle: float = 2.5,
                 bg_color: tuple = (250, 248, 244)) -> Image.Image:
    """이미지를 angle도만큼 미세하게 기울여 손에 든 느낌을 줌.
    expand=False로 크기 유지, BICUBIC으로 계단 현상 최소화.
    """
    return img.rotate(-angle, resample=Image.BICUBIC,
                      expand=False, fillcolor=bg_color)


def _soft_drop_shadow(canvas: Image.Image, img: Image.Image,
                      x: int, y: int,
                      blur: int = 18,
                      offset_x: int = 10, offset_y: int = 14,
                      opacity: int = 65) -> Image.Image:
    """img를 canvas의 (x, y)에 붙이기 전, 가우시안 블러 드롭 섀도우를 먼저 합성.
    opacity: 0~255 (기본 65 = 연한 회색조 그림자).
    canvas는 RGB 모드여야 함; RGBA로 변환 후 다시 RGB로 반환.
    """
    canvas_rgba = canvas.convert("RGBA")
    # 그림자 마스크: img 크기의 검정 레이어를 블러 처리
    shadow_base = Image.new("RGBA", img.size, (0, 0, 0, opacity))
    shadow_blur = shadow_base.filter(ImageFilter.GaussianBlur(radius=blur))
    canvas_rgba.paste(shadow_blur, (x + offset_x, y + offset_y), shadow_blur)
    # 원본 이미지 합성
    if img.mode == "RGBA":
        canvas_rgba.paste(img, (x, y), img)
    else:
        canvas_rgba.paste(img, (x, y))
    return canvas_rgba.convert("RGB")


# ── 배경 캐시 디렉토리 ──
BG_CACHE_DIR = BASE_DIR / "assets" / "bg_cache"

# ── 풀 크기 -- 씬당 N개 변형 미리 생성, 순환 발급 ──
# Why: 같은 씬이라도 리스팅마다 다른 배경 -> Etsy 중복 페널티 회피
BG_POOL_SIZE = 10

# ── 씬별 base seed -- variant i -> base_seed + i*100 ──
BG_BASE_SEEDS = {
    "hero_wall":        1337,
    "lifestyle_living": 2048,
    "lifestyle_bedroom": 4096,
    "lifestyle_dark":   8192,
    "flatlay_desk":     512,
    "flatlay_marble":   1024,
    "lifestyle_green":  2560,
}

# 하위 호환성 (기존 코드 참조용)
BG_SEEDS = BG_BASE_SEEDS

# ── 발급 카운터 파일 ──
_BG_COUNTER_PATH = BG_CACHE_DIR / "_counter.json"


def _load_counter() -> dict:
    """씬별 다음 발급 인덱스 로드."""
    if _BG_COUNTER_PATH.exists():
        import json
        try:
            return json.loads(_BG_COUNTER_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_counter(counter: dict) -> None:
    import json
    BG_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    _BG_COUNTER_PATH.write_text(json.dumps(counter, indent=2), encoding="utf-8")


def assign_bg_variant(scene: str) -> int:
    """씬에 대해 다음 변형 인덱스를 발급하고 카운터 +1.

    Returns: variant index (0 ~ BG_POOL_SIZE-1)
    """
    counter = _load_counter()
    idx = counter.get(scene, 0) % BG_POOL_SIZE
    counter[scene] = idx + 1
    _save_counter(counter)
    return idx

logger = logging.getLogger(__name__)

# ── Cloudflare Workers AI 이미지 생성 모델 ──
CF_IMAGE_MODEL = "@cf/black-forest-labs/flux-1-schnell"

# ── 모든 프롬프트에 강제 삽입되는 품질 부스터 ──
# Why: SD XL은 이 토큰들이 없으면 품질이 들쑥날쑥. 항상 포함해야 일관성 유지.
_QUALITY_BOOSTER = (
    "professional photography, ultra realistic, sharp focus, "
    "8k uhd, high detail, perfect exposure, award winning photo"
)
_NEGATIVE_PROMPT = (
    "people, person, human, face, hands, text, watermark, logo, "
    "blurry, out of focus, low quality, distorted, ugly, cropped, "
    "frame on wall, picture frame, painting on wall, "
    "oversaturated, cartoon, anime, illustration, painting, "
    "grainy, noise, jpeg artifacts, dark, underexposed, overexposed"
)

# ── 목업 씬별 AI 프롬프트 (품질 부스터 별도 관리 -- 중복 없이) ──
AI_BACKGROUND_PROMPTS = {
    "hero_wall": (
        "minimalist Scandinavian living room interior, light warm beige plaster wall, "
        "soft diffused morning sunlight from left window, wooden floating shelf with small "
        "green succulent plant and ceramic vase, clean empty wall space in center for artwork, "
        "shallow depth of field, no people, no text, no frames on wall"
    ),
    "lifestyle_living": (
        "cozy modern living room interior, off-white walls with subtle texture, "
        "light oak wood floor, beige linen sofa with throw pillows, "
        "natural light from large window, small potted fiddle leaf fig plant in corner, "
        "minimalist aesthetic, no people, no text, warm afternoon light"
    ),
    "lifestyle_bedroom": (
        "elegant minimalist bedroom interior, soft sage green accent wall, "
        "white linen bedding with subtle texture, wooden nightstand with small lamp, "
        "soft warm morning light through sheer curtains, cozy Scandinavian aesthetic, "
        "no people, no text"
    ),
    "lifestyle_dark": (
        "moody sophisticated home office interior, deep charcoal gray wall, "
        "warm Edison bulb desk lamp glowing, dark wood desk surface, "
        "small potted plant with dark leaves, dramatic side lighting, "
        "professional interior photography, cinematic atmosphere, ultra realistic, "
        "no people, no text, luxury feel"
    ),
    "flatlay_desk": (
        "top-down flat lay product photography, clean light oak wooden desk surface, "
        "white ceramic coffee mug with steam, yellow sharpened pencil, "
        "small green succulent in white pot, soft natural window light from top-left, "
        "minimal aesthetic, warm neutral tones, no people, no text, ample empty space in center"
    ),
    "flatlay_marble": (
        "top-down flat lay product photography, smooth white marble surface with subtle "
        "gray veining, small dried eucalyptus sprigs, rose gold pen, "
        "soft studio lighting, minimal and elegant, no people, no text, luxurious feel"
    ),
    "lifestyle_green": (
        "bright airy sunroom interior, white painted shiplap wall, "
        "lush green tropical plants -- monstera, pothos, and snake plant, "
        "natural sunlight flooding in, rattan furniture accent, "
        "bohemian Scandinavian style, no people, no text, fresh and vibrant"
    ),
}


def _score_bg_quality(img_path: str) -> tuple[float, dict]:
    """배경 이미지 품질 점수화 (0~100점).

    Returns: (total_score, detail_dict)
    각 항목 25점 만점:
      - 파일크기: >500KB=25, >300KB=20, >200KB=15, 이하=0
      - 색다양성: stddev 기반, 충분하면 25점
      - 밝기:     80~180 범위면 25점 (최적 실내 밝기)
      - 샤프니스: Laplacian 분산 기반 25점
    """
    detail = {"size": 0, "color": 0, "brightness": 0, "sharpness": 0}
    try:
        from PIL import ImageStat, ImageFilter

        # 1. 파일 크기
        size_kb = Path(img_path).stat().st_size / 1024
        if size_kb >= 500:   detail["size"] = 25
        elif size_kb >= 300: detail["size"] = 20
        elif size_kb >= 200: detail["size"] = 15
        else:                detail["size"] = 0

        img = Image.open(img_path).convert("RGB")
        stat = ImageStat.Stat(img)

        # 2. 색상 다양성
        min_std = min(stat.stddev[:3])
        detail["color"] = min(25, int(min_std * 25 / 30))  # stddev 30이 만점 기준

        # 3. 밝기 (80~180 = 최적 실내)
        brightness = sum(stat.mean[:3]) / 3
        if 80 <= brightness <= 180:
            detail["brightness"] = 25
        elif 60 <= brightness <= 210:
            detail["brightness"] = 18
        elif 40 <= brightness <= 230:
            detail["brightness"] = 10
        else:
            detail["brightness"] = 0

        # 4. 샤프니스 -- FIND_EDGES 평균 밝기 (엣지 많을수록 선명)
        try:
            gray = img.convert("L").resize((512, 512), Image.LANCZOS)
            edges = gray.filter(ImageFilter.FIND_EDGES)
            edge_mean = ImageStat.Stat(edges).mean[0]  # 선명한 실사: 보통 8~20
            detail["sharpness"] = min(25, int(edge_mean * 25 / 12))
        except Exception:
            detail["sharpness"] = 12  # 측정 실패 시 중간값

        total = sum(detail.values())
        return total, detail
    except Exception as e:
        logger.warning("BG 품질 측정 실패: %s -- 부분 점수 반환", e)
        return max(float(sum(detail.values())), 40.0), detail


def _validate_bg_quality(img_path: str, min_score: int = 60) -> bool:
    """품질 점수 60점 이상이면 통과."""
    score, detail = _score_bg_quality(img_path)
    passed = score >= min_score
    level = "통과" if passed else "거부"
    logger.info("BG 품질 %s: %d/100점 %s", level, score, detail)
    return passed


def _cf_post(prompt: str, seed: int, acct: dict) -> Optional[bytes]:
    """Cloudflare Workers AI 단일 호출. 성공 시 raw bytes 반환.

    Flux.1 Schnell: JSON {"result": {"image": "<base64>"}} 반환
    SDXL: raw PNG bytes 반환
    두 포맷 모두 처리.
    """
    import base64, json as _json
    url = (f"https://api.cloudflare.com/client/v4/accounts/"
           f"{acct['account_id']}/ai/run/{CF_IMAGE_MODEL}")
    # Flux.1은 negative_prompt/num_steps 미지원 -- 공통 파라미터만 사용
    payload = {
        "prompt": f"{prompt}, {_QUALITY_BOOSTER}",
        "width": 1024,
        "height": 1024,
    }
    # SDXL 전용 파라미터 (Flux.1에서는 무시됨)
    if "stable-diffusion" in CF_IMAGE_MODEL:
        payload["negative_prompt"] = _NEGATIVE_PROMPT
        payload["num_steps"] = 20
        payload["seed"] = seed
    resp = requests.post(
        url, headers={"Authorization": f"Bearer {acct['api_token']}"},
        json=payload, timeout=120,
    )
    if resp.status_code == 429:
        logger.warning("CF API 429 한도 소진: account_id=%s", acct.get("account_id", "")[:8])
        return b"__429__"  # 429 전용 마커
    if resp.status_code != 200:
        logger.warning("CF API %d: %s", resp.status_code, resp.text[:120])
        return None

    # Flux.1: Content-Type = application/json, image in base64
    ct = resp.headers.get("content-type", "")
    if "json" in ct:
        try:
            data = resp.json()
            # {"result": {"image": "<base64>"}} 또는 {"result": "<base64>"}
            result = data.get("result", data)
            b64 = result.get("image") if isinstance(result, dict) else result
            if b64:
                return base64.b64decode(b64)
        except Exception as e:
            logger.warning("CF JSON parse error: %s", e)
        return None

    # SDXL / 기타: raw bytes
    return resp.content


def generate_bg_with_quality_gate(
    scene: str,
    output_path: str,
    width: int = 2000,
    height: int = 2000,
    seed: Optional[int] = None,
    max_attempts: int = None,   # None = 계정 수만큼 자동 설정
    min_score: int = 60,
) -> Optional[str]:
    """품질 게이트 통과할 때까지 재생성 -- 10개 CF 계정 전부 순환 후 Together AI 폴백.

    흐름:
      1) CF 계정 10개를 순서대로 시도 (429면 다음 계정으로 즉시 전환)
      2) 성공한 계정에서 품질 점수 측정, 60점 미만이면 다음 시도
      3) 10계정 전부 429 소진 → Together AI FLUX.2-max 폴백
      4) 5회 모두 기준 미달 → best 결과 사용
    """
    base_prompt = AI_BACKGROUND_PROMPTS.get(scene, AI_BACKGROUND_PROMPTS["hero_wall"])
    base_seed = seed if seed is not None else int(time.time()) % 100000

    total_accounts = len(CLOUDFLARE_ACCOUNTS) if CLOUDFLARE_ACCOUNTS else 0
    if not total_accounts:
        logger.warning("Cloudflare 계정 미설정 -> Together AI 폴백")
        return _together_ai_fallback(base_prompt, output_path, width, height)

    # max_attempts: 기본값 = 전체 계정 수 (모든 계정 한 번씩 시도)
    if max_attempts is None:
        max_attempts = total_accounts

    best_path: Optional[str] = None
    best_score: float = -1
    exhausted_count = 0

    # 현재 CF 인덱스에서 시작해 최대 total_accounts 개 계정 순환
    acct = get_next_cloudflare_account()

    for attempt in range(max_attempts):
        attempt_seed = base_seed + attempt * 137
        tmp = output_path + f".tmp{attempt}"

        try:
            raw = _cf_post(base_prompt, attempt_seed, acct)
            if raw == b"__429__":
                exhausted_count += 1
                logger.warning("CF API 429 한도 소진: account_id=%s", acct.get("account_id", "")[:8])
                logger.warning("429 소진 계정 %d/%d -> 다음 계정 전환", exhausted_count, total_accounts)
                if exhausted_count >= total_accounts:
                    logger.warning("모든 CF 계정(%d개) 429 소진 -> Together AI 폴백", total_accounts)
                    if best_path and Path(best_path).exists():
                        Path(best_path).unlink(missing_ok=True)
                    return _together_ai_fallback(base_prompt, output_path, width, height)
                acct = get_next_cloudflare_account()
                continue

            if raw is None:
                acct = get_next_cloudflare_account()
                time.sleep(min(2 ** min(attempt, 3), 8))
                continue

            img = Image.open(io.BytesIO(raw)).convert("RGB")
            img = img.resize((width, height), Image.LANCZOS)
            img.save(tmp, "PNG")

            score, detail = _score_bg_quality(tmp)
            logger.info("CF 시도 %d/%d -- %d/100점 %s (acct=%s seed=%d)",
                        attempt + 1, max_attempts, score, detail,
                        acct.get("account_id", "")[:8], attempt_seed)

            if score > best_score:
                if best_path:
                    Path(best_path).unlink(missing_ok=True)
                best_score, best_path = score, tmp
            else:
                Path(tmp).unlink(missing_ok=True)

            if score >= min_score:
                import shutil as _shutil
                _shutil.move(best_path, output_path)
                logger.info("BG 생성 완료: %s %d/100점", scene, score)
                return output_path

            # 품질 미달이어도 계정 전환해서 재시도
            acct = get_next_cloudflare_account()

        except Exception as e:
            logger.warning("CF 시도 %d 예외: %s", attempt + 1, e)
            Path(tmp).unlink(missing_ok=True)
            time.sleep(1)

    # 전체 시도 후 기준 미달 -> best 사용
    if best_path and Path(best_path).exists():
        import shutil as _shutil
        _shutil.move(best_path, output_path)
        logger.warning("품질 기준(%d점) 미달 -- best 사용 (%.0f점): %s", min_score, best_score, scene)
        return output_path

    # CF 완전 실패 -> Together AI 폴백
    logger.warning("CF BG 생성 완전 실패 -> Together AI 폴백: %s", scene)
    return _together_ai_fallback(base_prompt, output_path, width, height)


def _together_ai_fallback(
    prompt: str, output_path: str, width: int = 2000, height: int = 2000
) -> Optional[str]:
    """Together AI FLUX.1-schnell 배경 이미지 생성 (유료 폴백).
    TOGETHER_API_KEY_1~10 다중키 순환, 없으면 단일 TOGETHER_API_KEY.
    """
    # ── 키 목록 수집 (settings 경유 → os.getenv 폴백) ──
    keys: list[str] = []
    try:
        from config.settings import settings as _s
        for i in range(1, 11):
            k = getattr(_s, f"together_api_key_{i}", "") or ""
            if k:
                keys.append(k)
        if not keys and getattr(_s, "together_api_key", ""):
            keys = [_s.together_api_key]
    except Exception:
        pass
    if not keys:
        for i in range(1, 11):
            k = os.getenv(f"TOGETHER_API_KEY_{i}", "")
            if k:
                keys.append(k)
    if not keys:
        single = os.getenv("TOGETHER_API_KEY", "")
        if single:
            keys = [single]
    if not keys:
        logger.error("TOGETHER_API_KEY 없음 -> 이미지 생성 불가")
        return None

    import base64 as _b64
    url = "https://api.together.xyz/v1/images/generations"
    payload = {
        "model": "black-forest-labs/FLUX.1-schnell",
        "prompt": f"{prompt}, professional photography, ultra realistic, sharp focus, 8k uhd",
        "width": min(width, 1024),
        "height": min(height, 1024),
        "steps": 4,
        "n": 1,
        "response_format": "b64_json",
    }

    for ki, api_key in enumerate(keys):
        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        for attempt in range(2):
            try:
                logger.info("Together AI 키%d FLUX.1-schnell 배경 생성 중... (%d/2)", ki + 1, attempt + 1)
                resp = requests.post(url, headers=headers, json=payload, timeout=120)
                if resp.status_code == 402:
                    logger.warning("Together AI 키%d 잔액 부족", ki + 1)
                    break  # 다음 키로
                if resp.status_code == 429:
                    logger.warning("Together AI 키%d 한도 소진", ki + 1)
                    break  # 다음 키로
                if resp.status_code == 400:
                    logger.warning("Together AI 키%d 400: %s",
                                   ki + 1, resp.text[:80])
                    break  # 재시도 불필요
                if resp.status_code == 200:
                    b64 = resp.json().get("data", [{}])[0].get("b64_json", "")
                    if b64:
                        img_bytes = _b64.b64decode(b64)
                        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
                        img = img.resize((width, height), Image.LANCZOS)
                        img.save(output_path, "PNG")
                        logger.info("Together AI 키%d 생성 완료: %s", ki + 1, output_path)
                        return output_path
                else:
                    logger.warning("Together AI 키%d %d (%d/2)", ki + 1, resp.status_code, attempt + 1)
            except Exception as e:
                logger.warning("Together AI 키%d 오류 (%d/2): %s", ki + 1, attempt + 1, e)
            time.sleep(2)

    logger.error("Together AI 전 키 실패 -> 이미지 없이 진행")
    return None


def _generate_ai_background(scene: str, output_path: str,
                              width: int = 2000, height: int = 2000,
                              seed: Optional[int] = None) -> Optional[str]:
    """하위 호환용 -- generate_bg_with_quality_gate 로 위임."""
    return generate_bg_with_quality_gate(scene, output_path, width, height, seed)


def _get_image_from_file(file_path: str) -> Optional[str]:
    """
    Convert product file to preview image for mockups.
    Why: Mockups need actual product visuals, not placeholder text.

    Priority:
    1. If preview.html exists (HTML+CSS products) -> Playwright screenshot (best quality)
    2. If PDF -> pdf2image conversion
    3. If XLSX -> Playwright screenshot of a styled preview
    4. Fallback -> placeholder
    """
    p = Path(file_path)
    if p.suffix.lower() in (".png", ".jpg", ".jpeg"):
        return file_path

    # Check if HTML preview exists (generated by worksheet_html.py / planner_html.py)
    html_preview = p.parent / "preview.html"
    img_path = p.parent / f"{p.stem}_preview.png"

    if html_preview.exists():
        try:
            return _screenshot_html(str(html_preview), str(img_path))
        except Exception as e:
            logger.warning("HTML screenshot failed, trying fallback: %s", e)

    if p.suffix.lower() == ".pdf":
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(str(p), first_page=1, last_page=1, dpi=150)
            if images:
                images[0].save(str(img_path), "PNG")
                return str(img_path)
        except ImportError:
            logger.warning("pdf2image not installed for %s", p.name)
        except Exception as e:
            logger.error("PDF conversion failed: %s", e)

        # Fallback: try Playwright to render PDF page
        try:
            return _screenshot_pdf(str(p), str(img_path))
        except Exception:
            pass

    if p.suffix.lower() == ".xlsx":
        # Generate a styled spreadsheet preview
        return _generate_xlsx_preview(str(p), str(img_path))

    # Final fallback: simple placeholder
    return _create_placeholder(str(img_path), p.stem)


def _screenshot_html(html_path: str, output_path: str, page_index: int = -1) -> str:
    """Take screenshot of HTML preview using Playwright.
    Why: Capture the most visually appealing page for mockups.
    page_index: -1=auto-detect best page, 0=cover, 1+=specific page
    """
    from playwright.sync_api import sync_playwright

    html_content = Path(html_path).read_text(encoding="utf-8")
    with sync_playwright() as pw:
        browser = pw.chromium.launch()
        page = browser.new_page(viewport={"width": 850, "height": 1100})
        page.set_content(html_content, wait_until="networkidle")

        pages = page.query_selector_all(".page")
        if not pages:
            page.screenshot(path=output_path)
            browser.close()
            return output_path

        if page_index == -1:
            # Why: Auto-detect best page for mockup.
            # Skip cover (0) and TOC (1). Pick first content page with actual data.
            # For worksheets: page 2 (first Easy problems page)
            # For planners: page 2+ (first monthly/daily page, skip TOC)
            best = min(2, len(pages) - 1)
            # If page has class "cover" or contains "Table of Contents", skip it
            for i in range(len(pages)):
                inner = pages[i].inner_text()
                if "Table of Contents" in inner or i == 0:
                    continue
                # Found a content page
                best = i
                break
            page_index = best

        target = min(page_index, len(pages) - 1)
        pages[target].scroll_into_view_if_needed()
        pages[target].screenshot(path=output_path)
        browser.close()

    logger.info("HTML screenshot saved: %s", output_path)
    return output_path


def _screenshot_pdf(pdf_path: str, output_path: str) -> str:
    """Render PDF first page using Playwright."""
    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        browser = pw.chromium.launch()
        page = browser.new_page(viewport={"width": 850, "height": 1100})
        page.goto(f"file:///{pdf_path.replace(chr(92), '/')}")
        page.wait_for_timeout(2000)
        page.screenshot(path=output_path)
        browser.close()

    logger.info("PDF screenshot saved: %s", output_path)
    return output_path


_SAMPLE_BUDGET_DATA = {
    # Income rows: label → (Expected, Actual)
    "Salary":        (5200, 5200), "Side Income":  (800, 650),
    "Investments":   (300, 320),   "Other":        (150, 0),
    # Expense rows: label → (Budget, Actual)
    "Housing/Rent":  (1500, 1500), "Utilities":    (180, 195),
    "Groceries":     (450, 412),   "Transportation":(220, 238),
    "Insurance":     (280, 280),   "Healthcare":   (120, 85),
    "Entertainment": (100, 134),   "Dining Out":   (200, 267),
    "Shopping":      (150, 89),    "Subscriptions": (60, 60),
    "Savings":       (500, 500),   "Debt Payment": (350, 350),
    "Education":     (80, 80),     "Personal Care": (60, 45),
    "Gifts":         (50, 0),      "Other":        (100, 72),
    # Debt payoff
    "Credit Card A": (3200, 3200), "Credit Card B": (1850, 1850),
    "Student Loan":  (18500, 18500),"Car Loan":    (8200, 8200),
    # Savings goals
    "Emergency Fund":(5000, 2340), "Vacation Fund": (2000, 780),
    "New Car":       (15000, 4200),"House Down Pay.":(50000, 8500),
    # Wedding budget
    "Venue":         (6000, 5800), "Catering":     (4500, 4500),
    "Photography":   (2500, 2500), "Flowers":      (1200, 980),
    "Music/DJ":      (1500, 1500), "Attire":       (2000, 1850),
    # Small business
    "Product Sales": (8500, 9120), "Services":     (3200, 2800),
    "Advertising":   (600, 550),   "Rent/Office":  (1200, 1200),
    "Supplies":      (300, 275),   "Software":     (150, 150),
}


def _generate_xlsx_preview(xlsx_path: str, output_path: str) -> str:
    """Generate styled spreadsheet preview image.
    Why: Show actual budget data sheet, not How to Use tab. Customers buy what they see.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)

        # Why: Find the most data-rich sheet (skip "How to Use", prefer "Monthly Budget", "Dashboard")
        best_sheet = None
        for name in ["Monthly Budget", "Dashboard", "Debt Payoff Plan", "Profit & Loss",
                      "Wedding Budget", "Savings Tracker"]:
            if name in wb.sheetnames:
                best_sheet = wb[name]
                break
        if not best_sheet:
            # Skip "How to Use" if possible
            for ws_candidate in wb.worksheets:
                if "how" not in ws_candidate.title.lower():
                    best_sheet = ws_candidate
                    break
        ws = best_sheet or wb.active

        # Build HTML table from best sheet
        # Pre-compute demo totals from sample data for TOTAL rows
        _income_labels = ["Salary", "Side Income", "Investments", "Other"]
        _expense_labels = [
            "Housing/Rent", "Utilities", "Groceries", "Transportation",
            "Insurance", "Healthcare", "Entertainment", "Dining Out",
            "Shopping", "Subscriptions", "Savings", "Debt Payment",
            "Education", "Personal Care", "Gifts", "Other",
        ]
        _demo_income_expected = sum(_SAMPLE_BUDGET_DATA.get(l, (0, 0))[0] for l in _income_labels)
        _demo_income_actual   = sum(_SAMPLE_BUDGET_DATA.get(l, (0, 0))[1] for l in _income_labels)
        _demo_exp_budget      = sum(_SAMPLE_BUDGET_DATA.get(l, (0, 0))[0] for l in _expense_labels)
        _demo_exp_actual      = sum(_SAMPLE_BUDGET_DATA.get(l, (0, 0))[1] for l in _expense_labels)
        _total_map = {
            "TOTAL INCOME":   (_demo_income_expected, _demo_income_actual),
            "TOTAL EXPENSES": (_demo_exp_budget,      _demo_exp_actual),
        }

        rows_html = ""
        for row_idx, row in enumerate(ws.iter_rows(max_row=25, max_col=6, values_only=True)):
            cells = ""
            row_label = str(row[0]).strip() if row[0] else ""
            sample = _SAMPLE_BUDGET_DATA.get(row_label)
            total_demo = _total_map.get(row_label.upper())
            for col_idx, val in enumerate(row):
                val_str = str(val) if val is not None else ""
                # Inject computed totals for TOTAL rows
                if total_demo and col_idx in (1, 2) and (val is None or str(val).startswith("=")):
                    val = total_demo[col_idx - 1]
                    val_str = f"${val:,.2f}"
                # Inject sample data for empty numeric input cells
                elif sample and col_idx in (1, 2) and (val is None or val == "" or str(val).startswith("=")):
                    val = sample[col_idx - 1]
                    val_str = f"${val:,.2f}"
                # Format actual numeric values
                elif isinstance(val, (int, float)) and val != 0:
                    val_str = f"${val:,.2f}" if abs(val) > 1 else f"{val:.1%}" if val < 1 else val_str
                elif val_str.startswith("="):
                    val_str = "$0.00"  # Show formula placeholder

                is_header = row_idx == 0 or (isinstance(val, str) and val.isupper() and len(val) > 2)
                is_total = isinstance(val, str) and "TOTAL" in val.upper()
                is_label = col_idx == 0 and isinstance(val, str)

                if is_header:
                    cells += f'<td style="background:#3A6B8C;color:white;font-weight:700;padding:10px 14px;font-size:12px;letter-spacing:0.5px">{val_str}</td>'
                elif is_total:
                    cells += f'<td style="background:#E8F4FD;font-weight:700;padding:8px 12px;font-size:12px;color:#2C3E50;border:2px solid #3A6B8C">{val_str}</td>'
                elif is_label:
                    cells += f'<td style="background:white;padding:8px 12px;font-size:11px;font-weight:600;color:#555;border:1px solid #E8EEF4">{val_str}</td>'
                elif row_idx % 2 == 0:
                    cells += f'<td style="background:#F8FBFD;padding:8px 12px;font-size:11px;border:1px solid #E8EEF4;color:#333">{val_str}</td>'
                else:
                    cells += f'<td style="background:white;padding:8px 12px;font-size:11px;border:1px solid #E8EEF4;color:#333">{val_str}</td>'
            rows_html += f"<tr>{cells}</tr>"

        sheet_title = ws.title or "Budget"
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
        <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@400;500;600;700&display=swap" rel="stylesheet">
        <style>
            body {{ font-family: 'Poppins', sans-serif; background: #F0F4F8; padding: 24px; }}
            .title {{ font-size: 16px; font-weight: 700; color: #2C3E50; margin-bottom: 12px; padding-left: 4px; }}
            .tabs {{ display: flex; gap: 2px; margin-bottom: -1px; }}
            .tab {{ padding: 6px 16px; font-size: 10px; font-weight: 600; border-radius: 6px 6px 0 0; background: #D8E4EE; color: #666; }}
            .tab.active {{ background: white; color: #2C3E50; border: 1px solid #E0E8F0; border-bottom: none; }}
            table {{ border-collapse: collapse; width: 100%; background: white; box-shadow: 0 2px 12px rgba(0,0,0,0.06); border-radius: 0 0 8px 8px; overflow: hidden; }}
        </style></head><body>
        <div class="title">📊 {sheet_title}</div>
        <div class="tabs">
            <div class="tab active">{sheet_title}</div>
            <div class="tab">Daily Tracker</div>
            <div class="tab">Savings</div>
        </div>
        <table>{rows_html}</table></body></html>"""

        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            browser = pw.chromium.launch()
            page = browser.new_page(viewport={"width": 850, "height": 600})
            page.set_content(html, wait_until="networkidle")
            page.screenshot(path=output_path)
            browser.close()

        logger.info("XLSX preview saved: %s", output_path)
        return output_path
    except Exception as e:
        logger.warning("XLSX preview failed: %s", e)
        return _create_placeholder(output_path, Path(xlsx_path).stem)


def _create_placeholder(output_path: str, label: str) -> str:
    """Last resort: simple branded placeholder."""
    img = Image.new("RGB", (850, 1100), (250, 248, 245))
    draw = ImageDraw.Draw(img)
    font = _font("Regular", 20)
    draw.text((300, 520), label, fill=(180, 180, 180), font=font)
    draw.rectangle([(40, 40), (810, 1060)], outline=(220, 220, 220), width=2)
    img.save(output_path, "PNG")
    return output_path

# ── Mockup canvas settings ──
MOCKUP_WIDTH = 2000
MOCKUP_HEIGHT = 2000
ETSY_RATIO = (4, 3)  # Etsy recommended


def _save_mockup(canvas: Image.Image, output_path: str, sharpen: bool = True) -> None:
    """UnsharpMask 적용 후 JPEG 저장.
    Why: 모든 목업에 일관된 선명도 부여 -> sharpness 점수 향상 + 인쇄 선명도.
    """
    if sharpen:
        canvas = canvas.filter(ImageFilter.UnsharpMask(radius=1.5, percent=160, threshold=2))
    canvas.save(output_path, "JPEG", quality=95)


# ── Background colors for different room settings ──
ROOM_BACKGROUNDS = {
    "white_wall":   (252, 251, 249),
    "light_grey":   (237, 237, 235),
    "warm_beige":   (242, 231, 215),
    "soft_cream":   (250, 244, 232),
    "sage_green":   (215, 228, 210),
    "dusty_rose":   (238, 220, 218),
    "navy_accent":  (32, 50, 74),
    "charcoal":     (48, 48, 52),
    "terracotta":   (210, 175, 155),
    "muted_olive":  (185, 195, 165),
}

# ── 배경 그래디언트 정의 (단색 대신 미묘한 그래디언트로 입체감) ──
ROOM_GRADIENTS = {
    "white_wall":   [(255, 254, 252), (242, 240, 237)],
    "light_grey":   [(245, 245, 243), (225, 224, 222)],
    "warm_beige":   [(250, 240, 226), (230, 218, 200)],
    "soft_cream":   [(255, 250, 240), (238, 230, 215)],
    "sage_green":   [(228, 238, 222), (200, 215, 195)],
    "dusty_rose":   [(245, 228, 226), (225, 208, 208)],
    "navy_accent":  [(42, 62, 88),    (22, 38, 58)],
    "charcoal":     [(60, 60, 65),    (35, 35, 40)],
    "terracotta":   [(220, 188, 168), (195, 160, 138)],
    "muted_olive":  [(198, 208, 178), (172, 184, 152)],
}


def _draw_gradient_background(
    canvas: "Image.Image",
    room: str = "warm_beige",
    direction: str = "vertical",
) -> None:
    """단색 배경 대신 미묘한 그래디언트로 입체감.깊이 표현."""
    from PIL import ImageDraw
    colors = ROOM_GRADIENTS.get(room, ROOM_GRADIENTS["warm_beige"])
    c1, c2 = colors[0], colors[1]
    w, h = canvas.size
    draw = ImageDraw.Draw(canvas)

    if direction == "vertical":
        for y in range(h):
            ratio = y / h
            r = int(c1[0] + (c2[0] - c1[0]) * ratio)
            g = int(c1[1] + (c2[1] - c1[1]) * ratio)
            b = int(c1[2] + (c2[2] - c1[2]) * ratio)
            draw.line([(0, y), (w, y)], fill=(r, g, b))
    else:  # horizontal
        for x in range(w):
            ratio = x / w
            r = int(c1[0] + (c2[0] - c1[0]) * ratio)
            g = int(c1[1] + (c2[1] - c1[1]) * ratio)
            b = int(c1[2] + (c2[2] - c1[2]) * ratio)
            draw.line([(x, 0), (x, h)], fill=(r, g, b))


def _add_wall_texture(canvas: "Image.Image", room: str = "warm_beige", intensity: int = 4) -> None:
    """벽 질감 시뮬레이션 -- 미묘한 노이즈로 페인트/석회 느낌."""
    import random as _rnd
    from PIL import ImageDraw
    draw = ImageDraw.Draw(canvas)
    w, h = canvas.size
    base_color = ROOM_BACKGROUNDS.get(room, (245, 235, 220))
    for _ in range(w * h // 80):
        x = _rnd.randint(0, w - 1)
        y = _rnd.randint(0, h - 1)
        offset = _rnd.randint(-intensity, intensity)
        shade = tuple(max(0, min(255, c + offset)) for c in base_color)
        draw.point((x, y), fill=shade)

# ── Frame styles ──
FRAME_COLORS = {
    "black": (30, 30, 30),
    "white": (250, 250, 250),
    "natural_wood": (180, 140, 100),
    "dark_wood": (80, 55, 35),
    "gold": (200, 170, 110),
}


def _create_framed_art(
    art_path: str,
    frame_color: tuple[int, int, int] = (30, 30, 30),
    frame_width: int = 20,
    mat_width: int = 40,
    target_size: tuple[int, int] = (600, 800),
    mat_color: Optional[tuple[int, int, int]] = None,
) -> Image.Image:
    """Create a framed version of the art image.
    mat_color=None → 자동감지: 아트가 밝으면(avg>235) 따뜻한 그레이 매트, 어두우면 흰 매트.
    """
    art = Image.open(art_path).convert("RGBA")
    art = art.resize(target_size, Image.LANCZOS)

    # 자동 대비 강화: 아트가 너무 연하면(avg>235) 명암 증가로 가시성 확보
    try:
        from PIL import ImageEnhance, ImageStat
        art_rgb = art.convert("RGB")
        stat = ImageStat.Stat(art_rgb)
        avg_brightness = sum(stat.mean[:3]) / 3
        if avg_brightness > 230:
            enhancer = ImageEnhance.Contrast(art_rgb)
            art_rgb = enhancer.enhance(3.0)  # 대비 3배 강화
            enhancer2 = ImageEnhance.Sharpness(art_rgb)
            art_rgb = enhancer2.enhance(1.5)
            # RGBA로 복원 (원본 알파 채널 보존)
            r, g, b = art_rgb.split()
            _, _, _, a = art.split()
            art = Image.merge("RGBA", (r, g, b, a))
            logger.debug("대비 강화 적용: avg_brightness=%.1f", avg_brightness)
    except Exception as _ce:
        logger.debug("대비 강화 스킵: %s", _ce)

    # 매트 색상 자동 선택
    if mat_color is None:
        try:
            from PIL import ImageStat
            stat = ImageStat.Stat(art.convert("RGB"))
            avg_brightness = sum(stat.mean[:3]) / 3
            if avg_brightness > 235:
                # 아트가 매우 밝음 → 따뜻한 오프화이트 매트로 구분감 확보
                mat_color = (245, 242, 237)
            else:
                mat_color = (255, 255, 255)
        except Exception:
            mat_color = (255, 255, 255)

    # Mat border inside frame
    mat_size = (
        target_size[0] + mat_width * 2,
        target_size[1] + mat_width * 2,
    )
    mat = Image.new("RGBA", mat_size, (*mat_color, 255))
    mat.paste(art, (mat_width, mat_width))

    # Frame
    framed_size = (
        mat_size[0] + frame_width * 2,
        mat_size[1] + frame_width * 2,
    )
    framed = Image.new("RGBA", framed_size, (*frame_color, 255))
    framed.paste(mat, (frame_width, frame_width))

    return framed


def _add_shadow(img: Image.Image, offset: int = 14, blur: int = 28,
                alpha: int = 90) -> Image.Image:
    """자연스러운 드롭섀도 -- 좌상단 광원 기준 우하단 방향."""
    pad = blur * 2
    shadow_size = (img.width + pad + offset, img.height + pad + offset)
    result = Image.new("RGBA", shadow_size, (0, 0, 0, 0))

    # 섀도 레이어 (더 부드럽고 진한 그라데이션)
    shadow_layer = Image.new("RGBA", (img.width, img.height), (0, 0, 0, alpha))
    result.paste(shadow_layer, (offset + blur, offset + blur))
    result = result.filter(ImageFilter.GaussianBlur(blur))

    # 원본을 섀도 위에 붙임
    result.paste(img, (blur // 2, blur // 2), img)
    return result


def _blend_ambient(product: Image.Image, bg: Image.Image,
                   x: int, y: int, strength: float = 0.07) -> Image.Image:
    """배경 환경광을 상품에 미묘하게 반영 -- '오려붙인' 느낌 제거.

    배경의 해당 영역 평균 색상을 product에 strength 비율로 multiply.
    """
    from PIL import ImageStat
    w, h = product.size
    # 배경에서 상품이 놓일 영역의 색상 샘플링
    bx1, by1 = max(0, x), max(0, y)
    bx2, by2 = min(bg.width, x + w), min(bg.height, y + h)
    if bx2 <= bx1 or by2 <= by1:
        return product

    bg_crop = bg.crop((bx1, by1, bx2, by2)).convert("RGB")
    stat = ImageStat.Stat(bg_crop)
    avg_r, avg_g, avg_b = [int(v) for v in stat.mean[:3]]

    # 환경광 오버레이 (매우 약하게)
    overlay = Image.new("RGBA", product.size, (avg_r, avg_g, avg_b, int(255 * strength)))
    blended = product.convert("RGBA").copy()
    blended = Image.alpha_composite(blended, overlay)
    return blended


def _add_text_overlay(
    canvas: Image.Image,
    text: str,
    position: tuple[int, int],
    font_size: int = 36,
    color: tuple[int, int, int] = (60, 60, 60),
) -> None:
    """Add text to canvas."""
    draw = ImageDraw.Draw(canvas)
    font = _font("Regular", font_size)
    draw.text(position, text, fill=color, font=font)


_SCENE_TO_ROOM = {
    "hero_wall": "warm_beige", "lifestyle_living": "soft_cream",
    "lifestyle_bedroom": "sage_green", "lifestyle_dark": "charcoal",
    "flatlay_desk": "warm_beige", "flatlay_marble": "white_wall",
    "lifestyle_green": "sage_green",
}


def _cache_path(scene: str, variant: int) -> Path:
    return BG_CACHE_DIR / f"{scene}_{variant}.png"


def _gradient_fallback(scene: str) -> Image.Image:
    """최후 폴백 -- AI 완전 실패 시 그래디언트 (시간 기반 seed -> 매번 약간 다름)."""
    import random as _rnd
    room = _SCENE_TO_ROOM.get(scene, "warm_beige")
    _rnd.seed(int(time.time()))
    canvas = Image.new("RGB", (MOCKUP_WIDTH, MOCKUP_HEIGHT), ROOM_BACKGROUNDS[room])
    _draw_gradient_background(canvas, room)
    _add_wall_texture(canvas, room, intensity=3)
    _rnd.seed()
    logger.warning("그래디언트 폴백 사용: %s (Cloudflare 연결 확인 권장)", scene)
    return canvas


def get_listing_bg(scene: str) -> Image.Image:
    """리스팅 이미지 생성 시 호출 -- 매번 새 배경 생성 + 품질 게이트 적용.

    - 시간 기반 seed -> 호출할 때마다 다른 이미지 (중복 없음)
    - quality gate -> 60점 미만이면 최대 5회 재생성, best 결과 사용
    """
    BG_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    tmp_path = str(BG_CACHE_DIR / f"_tmp_{scene}_{int(time.time()*1000)}.png")
    result = generate_bg_with_quality_gate(scene, tmp_path, MOCKUP_WIDTH, MOCKUP_HEIGHT)
    if result:
        return Image.open(result).convert("RGB")
    return _gradient_fallback(scene)



def generate_wall_art_gallery_hero(
    art_paths: list[str], output_path: str,
    overlay_text: str = "9 DESIGNS INCLUDED",
    sub_text: str = "45 JPG Files · 5 Print Sizes",
) -> bool:
    """Wall art gallery 히어로 — 3개 디자인을 나란히 배치.

    상위 1% 전략:
    - 깨끗한 크림 그라데이션 배경 (AI 방 합성 원근법 문제 없음)
    - 메인 디자인 1개 크게 + 좌우에 보조 디자인 2개
    - "Gallery Set of 9" 가치 제안 배지
    - 갤러리 세트의 볼륨감을 한 장에 전달
    """
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT
        # 따뜻한 크림 그라데이션 배경
        canvas = Image.new("RGB", (W, H), (252, 249, 245))
        draw = ImageDraw.Draw(canvas)
        for y in range(H):
            ratio = y / H
            r = int(252 - ratio * 10)
            g = int(249 - ratio * 8)
            b = int(245 - ratio * 6)
            draw.line([(0, y), (W, y)], fill=(r, g, b))

        # 사용 가능한 아트 경로 (최소 1개 필요)
        paths = [p for p in art_paths if Path(p).exists()]
        if not paths:
            return False
        # 3개 사용: 가운데=메인(크게), 좌우=보조(작게)
        main_path = paths[0]
        left_path = paths[1] if len(paths) > 1 else paths[0]
        right_path = paths[2] if len(paths) > 2 else paths[0]

        # ── 메인 프레임 (중앙, 크게) ──
        main_w = int(W * 0.38)
        main_h = int(main_w * 1.38)
        main_framed = _create_framed_art(
            main_path, frame_color=(25, 25, 25),
            frame_width=16, mat_width=30,
            target_size=(main_w, main_h),
        )
        main_framed = _add_shadow(main_framed, offset=14, blur=28, alpha=90)
        mx = (W - main_framed.width) // 2
        my = (H - main_framed.height) // 2 - 60
        canvas.paste(main_framed, (mx, my), main_framed)

        # ── 좌측 보조 프레임 ──
        side_w = int(W * 0.22)
        side_h = int(side_w * 1.38)
        for side_path, x_factor, rotate_deg in [
            (left_path, 0.18, -4),
            (right_path, 0.82, 4),
        ]:
            s_framed = _create_framed_art(
                side_path, frame_color=(50, 48, 45),
                frame_width=12, mat_width=22,
                target_size=(side_w, side_h),
            )
            s_framed = _add_shadow(s_framed, offset=10, blur=20, alpha=75)
            # 약간 회전
            s_rotated = s_framed.rotate(rotate_deg, expand=True, resample=Image.BICUBIC)
            sx = int(W * x_factor) - s_rotated.width // 2
            sy = my + (main_framed.height - s_rotated.height) // 2 + 60
            canvas.paste(s_rotated, (sx, sy), s_rotated)

        # ── 상단 배지 (좌상단) ──
        badge_f = _font("Bold", 44)
        sub_f   = _font("SemiBold", 28)
        sm_f    = _font("Light", 22)
        bg_w = len(overlay_text) * 24 + 60
        draw.rounded_rectangle([(36, 36), (36 + bg_w, 108)],
                               radius=28, fill=(244, 100, 55))
        draw.text((36 + bg_w // 2, 72), overlay_text,
                  fill=(255, 255, 255), font=badge_f, anchor="mm")

        # 서브 배지 (우상단)
        sb_w = len(sub_text) * 17 + 52
        draw.rounded_rectangle([(W - 36 - sb_w, 36), (W - 36, 96)],
                               radius=22, fill=(22, 160, 133))
        draw.text((W - 36 - sb_w // 2, 66), sub_text,
                  fill=(255, 255, 255), font=sub_f, anchor="mm")

        # ── 하단 브랜드 바 ──
        bar_h = 80
        draw.rectangle([(0, H - bar_h), (W, H)], fill=(20, 28, 48))
        draw.text((W // 2, H - bar_h // 2), "DailyPrintHaus  ·  Instant Digital Download",
                  fill=(180, 190, 200), font=sm_f, anchor="mm")

        # ── 하단 INSTANT DOWNLOAD 뱃지 ──
        inst_f = _font("Bold", 32)
        inst_w = 420
        draw.rounded_rectangle(
            [(W // 2 - inst_w // 2, H - bar_h - 70),
             (W // 2 + inst_w // 2, H - bar_h - 10)],
            radius=30, fill=(22, 160, 133)
        )
        draw.text((W // 2, H - bar_h - 40), "INSTANT DOWNLOAD",
                  fill=(255, 255, 255), font=inst_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Wall art gallery hero saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Wall art gallery hero failed: %s", e)
        return False


def generate_hero_mockup(art_path: str, output_path: str,
                         overlay_text: str = "",
                         sub_text: str = "") -> bool:
    """Image 1: Hero -- AI 실사 배경 위에 프레임 합성. 클릭률 결정.

    상위 1% 기준:
    - 캔버스 너비 38% 크기 프레임
    - 수직 중앙보다 약간 위 (시선 유도)
    - 환경광 블렌딩으로 자연스러운 합성
    - 큰 혜택 텍스트 오버레이 (상위 1% 필수 요소)
    """
    try:
        canvas = get_listing_bg("hero_wall")

        # 캔버스 38% 너비 -- 실제 벽에 걸린 비율감
        art_w = int(MOCKUP_WIDTH * 0.38)
        art_h = int(art_w * 1.35)  # A4/Letter 비율

        framed = _create_framed_art(
            art_path,
            frame_color=FRAME_COLORS["black"],
            frame_width=18,
            mat_width=35,
            target_size=(art_w, art_h),
        )
        framed_with_shadow = _add_shadow(framed, offset=16, blur=32, alpha=85)

        # 수직 중앙 기준 10% 위 -- 시선이 자연스럽게 향하는 위치
        x = (MOCKUP_WIDTH - framed_with_shadow.width) // 2
        y = int((MOCKUP_HEIGHT - framed_with_shadow.height) * 0.42)

        # 환경광 블렌딩
        framed_blended = _blend_ambient(framed_with_shadow, canvas, x, y, strength=0.06)
        canvas.paste(framed_blended, (x, y), framed_blended)

        draw = ImageDraw.Draw(canvas)
        hero_big_f = _font("Bold", 52)
        hero_sub_f = _font("SemiBold", 30)
        badge_f = _font("SemiBold", 28)

        # ── 상위 1% 필수: 가치 텍스트 오버레이 ──
        if overlay_text:
            text_bg_w = len(overlay_text) * 28 + 56
            draw.rounded_rectangle([(36, 36), (36 + text_bg_w, 114)],
                                   radius=30, fill=(244, 107, 60))
            draw.text((36 + text_bg_w // 2, 75), overlay_text.upper(),
                      fill=(255, 255, 255), font=hero_big_f, anchor="mm")

        if sub_text:
            text_bg_w2 = len(sub_text) * 20 + 48
            draw.rounded_rectangle(
                [(MOCKUP_WIDTH - 36 - text_bg_w2, 36), (MOCKUP_WIDTH - 36, 100)],
                radius=28, fill=(30, 160, 150))
            draw.text((MOCKUP_WIDTH - 36 - text_bg_w2 // 2, 68),
                      sub_text.upper(), fill=(255, 255, 255),
                      font=hero_sub_f, anchor="mm")

        # ── 상위 1% 필수: INSTANT DOWNLOAD 배지 (우하단 고정) ──
        badge_text = "INSTANT DOWNLOAD"
        badge_w = len(badge_text) * 16 + 48
        badge_h = 56
        badge_x = MOCKUP_WIDTH - badge_w - 28
        badge_y = MOCKUP_HEIGHT - badge_h - 80  # 브랜드 바 위
        draw.rounded_rectangle(
            [(badge_x, badge_y), (badge_x + badge_w, badge_y + badge_h)],
            radius=28, fill=(22, 160, 133))
        draw.text((badge_x + badge_w // 2, badge_y + badge_h // 2),
                  badge_text, fill=(255, 255, 255), font=badge_f, anchor="mm")

        # ── 하단 브랜드 바 (10장 전체 일관성) ──
        bar_h = 52
        draw.rectangle(
            [(0, MOCKUP_HEIGHT - bar_h), (MOCKUP_WIDTH, MOCKUP_HEIGHT)],
            fill=(15, 23, 42))
        brand_f = _font("Light", 22)
        draw.text((MOCKUP_WIDTH // 2, MOCKUP_HEIGHT - bar_h // 2),
                  "DailyPrintHaus  ·  Instant Digital Download",
                  fill=(180, 190, 200), font=brand_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Hero mockup saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Hero mockup failed: %s", e)
        return False


def generate_lifestyle_mockup(art_path: str, output_path: str,
                               room: str = "light_grey") -> bool:
    """Image 2: Lifestyle -- AI 실사 거실 배경 위에 프레임 합성.

    상위 1% 기준:
    - 캔버스 너비 28% -- 실제 방에서 액자가 차지하는 비율
    - 상단 18~40% 구간에 배치 (벽 공간)
    - 환경광 블렌딩
    """
    scene_map = {
        "light_grey": "lifestyle_living", "warm_beige": "lifestyle_living",
        "soft_cream": "lifestyle_living",  "sage_green": "lifestyle_green",
        "dusty_rose": "lifestyle_bedroom", "navy_accent": "lifestyle_dark",
        "charcoal": "lifestyle_dark",
    }
    scene = scene_map.get(room, "lifestyle_living")

    try:
        canvas = get_listing_bg(scene)

        # 캔버스 22% 너비 -- 라이프스타일 씬에서 액자는 방의 일부, 주인공이 아님
        # 상위 1%: 액자는 벽 한 섹션에 자연스럽게 걸려 있고 방 전체를 가리지 않음
        art_w = int(MOCKUP_WIDTH * 0.22)
        art_h = int(art_w * 1.35)

        framed = _create_framed_art(
            art_path,
            frame_color=FRAME_COLORS["black"],
            frame_width=14,
            mat_width=28,
            target_size=(art_w, art_h),
        )
        framed_with_shadow = _add_shadow(framed, offset=10, blur=20, alpha=80)

        # 우측 벽면 자연스러운 위치 (상단 15~22%) — 가구 위 확보된 벽 공간
        x = int(MOCKUP_WIDTH * 0.62) - framed_with_shadow.width // 2
        y = int(MOCKUP_HEIGHT * 0.15)

        framed_blended = _blend_ambient(framed_with_shadow, canvas, x, y, strength=0.08)
        canvas.paste(framed_blended, (x, y), framed_blended)

        # 브랜드 하단 바 (10장 전체 일관성)
        draw = ImageDraw.Draw(canvas)
        bar_h = 52
        draw.rectangle(
            [(0, MOCKUP_HEIGHT - bar_h), (MOCKUP_WIDTH, MOCKUP_HEIGHT)],
            fill=(15, 23, 42, 200))
        brand_f = _font("Light", 22)
        draw.text((MOCKUP_WIDTH // 2, MOCKUP_HEIGHT - bar_h // 2),
                  "DailyPrintHaus  ·  Instant Digital Download",
                  fill=(180, 190, 200), font=brand_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Lifestyle mockup saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Lifestyle mockup failed: %s", e)
        return False


def generate_flatlay_mockup(art_path: str, output_path: str,
                            bg_style: str = "wood",
                            overlay_text: str = "",
                            sub_text: str = "") -> bool:
    """Flat-lay mockup for worksheets/planners -- AI 실사 책상 배경 위에 합성.
    Why: 교육용 상품은 벽 액자가 아니라 책상 위 연필 옆에 있어야 자연스러움.
    """
    try:
        # AI 배경 씬 선택
        scene = "flatlay_marble" if bg_style == "marble" else "flatlay_desk"

        if bg_style == "marble":
            # 프로그래밍 방식 다크 마블 배경 — AI 폴백이 흰색/크림으로 나오는 문제 완전 우회
            # 상위 1%: 어두운 배경에 흰 종이 강한 대비 → 드라마틱한 플랫레이
            import random as _mrnd
            canvas = Image.new("RGB", (MOCKUP_WIDTH, MOCKUP_HEIGHT), (38, 38, 43))
            _draw_bg = ImageDraw.Draw(canvas)
            for _my in range(MOCKUP_HEIGHT):
                _ratio = _my / MOCKUP_HEIGHT
                _r = int(44 - _ratio * 12)
                _g = int(44 - _ratio * 12)
                _b = int(50 - _ratio * 14)
                _draw_bg.line([(0, _my), (MOCKUP_WIDTH, _my)], fill=(_r, _g, _b))
            _mrnd.seed(73)
            for _ in range(24):
                _x0 = _mrnd.randint(-100, MOCKUP_WIDTH)
                _x1 = _x0 + _mrnd.randint(-400, 400)
                _c  = _mrnd.randint(58, 88)
                _w  = _mrnd.randint(1, 2)
                _draw_bg.line([(_x0, 0), (_x1, MOCKUP_HEIGHT)],
                              fill=(_c, _c, _c + 5), width=_w)
            for _ in range(12):
                _x0 = _mrnd.randint(0, MOCKUP_WIDTH)
                _x1 = _x0 + _mrnd.randint(-200, 200)
                _draw_bg.line([(_x0, 0), (_x1, MOCKUP_HEIGHT)],
                              fill=(100, 100, 108), width=1)
        elif bg_style == "wood":
            # 프로그래밍 방식 나무 배경 — 진한 월넛 나뭇결 + 대형 소품으로 상위 1% 라이프스타일 연출
            import random as _wrnd
            _W_BASE = (148, 108, 68)   # 진한 월넛 베이스 (오크보다 어둡고 채도 높게)
            canvas = Image.new("RGB", (MOCKUP_WIDTH, MOCKUP_HEIGHT), _W_BASE)
            _draw_bg = ImageDraw.Draw(canvas)
            # 위→아래 자연광 그래디언트 (위 밝고 아래 어두운 조명 효과)
            for _wy in range(MOCKUP_HEIGHT):
                _ratio = _wy / MOCKUP_HEIGHT
                _r = int(165 - _ratio * 32)
                _g = int(122 - _ratio * 26)
                _b = int(78 - _ratio * 20)
                _draw_bg.line([(0, _wy), (MOCKUP_WIDTH, _wy)], fill=(_r, _g, _b))
            # ── 주 나뭇결 (굵고 진한 줄, 약간 대각선 방향) ──
            _wrnd.seed(55)
            for _gi in range(110):
                _wy = _wrnd.randint(0, MOCKUP_HEIGHT)
                # 시작/끝 y를 다르게 → 수평이 아닌 약간 기울어진 나뭇결
                _angle_drift = _wrnd.randint(-120, 120)
                _dark = _wrnd.randint(28, 55)
                _ww = _wrnd.randint(1, 5)
                _r2 = max(0, _W_BASE[0] - _dark)
                _g2 = max(0, _W_BASE[1] - _dark - 8)
                _b2 = max(0, _W_BASE[2] - _dark - 12)
                _draw_bg.line([(0, _wy), (MOCKUP_WIDTH, _wy + _angle_drift)],
                              fill=(_r2, _g2, _b2), width=_ww)
            # ── 서브 나뭇결 (밝은 하이라이트) ──
            for _ in range(40):
                _wy2 = _wrnd.randint(0, MOCKUP_HEIGHT)
                _drift2 = _wrnd.randint(-40, 40)
                _light = _wrnd.randint(10, 26)
                _draw_bg.line([(0, _wy2), (MOCKUP_WIDTH, _wy2 + _drift2)],
                              fill=(min(255, _W_BASE[0] + _light),
                                    min(255, _W_BASE[1] + _light - 2),
                                    min(255, _W_BASE[2] + _light - 4)), width=1)
            # ── 나뭇결 노드 (옹이) — 더 크게 ──
            for _ in range(6):
                _nx = _wrnd.randint(200, MOCKUP_WIDTH - 200)
                _ny = _wrnd.randint(200, MOCKUP_HEIGHT - 200)
                _draw_bg.ellipse([(_nx - 36, _ny - 20), (_nx + 36, _ny + 20)],
                                 fill=(108, 72, 40))
                _draw_bg.ellipse([(_nx - 24, _ny - 12), (_nx + 24, _ny + 12)],
                                 fill=(128, 88, 52))

            # ── 비네트 효과: 4면 가장자리 어둡게 (사진 느낌) ──
            _vig = Image.new("RGBA", (MOCKUP_WIDTH, MOCKUP_HEIGHT), (0, 0, 0, 0))
            _vig_d = ImageDraw.Draw(_vig)
            _vig_steps = 80
            for _vs in range(_vig_steps):
                _va = int(160 * (_vs / _vig_steps) ** 1.8)
                _voff = _vs * 8
                _vig_d.rectangle([(_voff, _voff),
                                   (MOCKUP_WIDTH - _voff, MOCKUP_HEIGHT - _voff)],
                                  outline=(0, 0, 0, _va), width=8)
            canvas_rgba = canvas.convert("RGBA")
            canvas_rgba = Image.alpha_composite(canvas_rgba, _vig)
            canvas = canvas_rgba.convert("RGB")
            _draw_bg = ImageDraw.Draw(canvas)

            # ── 소품 1: 좌상단 대형 식물 (잎 + 화분) — 2000px 기준으로 크게 ──
            # 화분 중심: (240, 520), 너비 320px
            _pot_cx = 240
            _pot_top = 480
            # 줄기들 — 잎 기준점
            _stem_base = (_pot_cx, _pot_top)
            _leaf_data = [
                # (stem_tip_x, stem_tip_y, leaf_w, leaf_h, angle_deg, color_idx)
                # 잎은 "elongated" 형태: h < w*0.45
                (-80,  160, 280,  90, 0),   # 왼쪽 위
                ( 60,   60, 300,  88, 1),   # 오른쪽 위 높이
                ( 200, 180, 260,  82, 2),   # 오른쪽
                (-40,  280, 240,  76, 0),   # 왼쪽 중
                ( 150, 320, 270,  84, 3),   # 오른쪽 중
                (-120, 380, 220,  70, 1),   # 왼쪽 아래
                ( 80,  420, 250,  78, 2),   # 앞 중앙
            ]
            _leaf_dark  = [(55, 105, 58), (48,  98, 52), (68, 118, 62), (52, 110, 55)]
            _leaf_light = [(82, 148, 80), (72, 138, 72), (95, 158, 84), (78, 142, 76)]
            for (_lx, _ly, _lw, _lh, _lci) in _leaf_data:
                _lxs = _pot_cx + _lx - _lw // 2
                _lys = _pot_top - _ly - _lh // 2
                # 잎 본체 (어두운 색)
                _draw_bg.ellipse([(_lxs, _lys), (_lxs + _lw, _lys + _lh)],
                                 fill=_leaf_dark[_lci % 4])
                # 잎 중앙 하이라이트 (밝은 색 — 잎맥 느낌)
                _vein_w = max(4, _lw // 18)
                _draw_bg.line([(_lxs + _lw // 2, _lys + 8),
                               (_lxs + _lw // 2, _lys + _lh - 8)],
                              fill=_leaf_light[_lci % 4], width=_vein_w)
                # 줄기
                _draw_bg.line([(_pot_cx, _pot_top),
                               (_pot_cx + _lx, _pot_top - _ly)],
                              fill=(75, 55, 35), width=8)
            # 화분 몸체 — 320px 너비
            _draw_bg.rounded_rectangle([(_pot_cx - 160, _pot_top),
                                        (_pot_cx + 160, _pot_top + 200)],
                                       radius=24, fill=(148, 112, 72))
            # 화분 입구 (더 넓은 테두리)
            _draw_bg.rounded_rectangle([(_pot_cx - 175, _pot_top - 20),
                                        (_pot_cx + 175, _pot_top + 22)],
                                       radius=14, fill=(128, 95, 58))

            # ── 소품 2: 우측 대형 펜 — 48px 너비, 가시성 확보 ──
            _pen_x = MOCKUP_WIDTH - 220   # 더 안쪽으로 (1780)
            _pen_y1, _pen_y2 = 100, MOCKUP_HEIGHT - 600
            _pen_w = 44
            _draw_bg.line([(_pen_x, _pen_y1 + 100), (_pen_x, _pen_y2)],
                          fill=(215, 188, 135), width=_pen_w)   # 몸통 (크림 골드)
            _draw_bg.line([(_pen_x, _pen_y1), (_pen_x, _pen_y1 + 120)],
                          fill=(58, 88, 118), width=_pen_w)     # 캡 (딥 네이비)
            # 캡 클립
            _draw_bg.line([(_pen_x + _pen_w // 2 + 4, _pen_y1 + 20),
                           (_pen_x + _pen_w // 2 + 4, _pen_y1 + 110)],
                          fill=(88, 118, 148), width=8)
            # 펜 끝 (뾰족한 부분)
            _draw_bg.polygon([(_pen_x - _pen_w // 2, _pen_y2),
                              (_pen_x + _pen_w // 2, _pen_y2),
                              (_pen_x, _pen_y2 + 80)],
                             fill=(185, 148, 98))

            # ── 소품 3: 우상단 커피 머그 ──
            _mug_cx = MOCKUP_WIDTH - 340
            _mug_cy = 300
            _mug_w, _mug_h = 280, 320
            # 머그 몸체
            _draw_bg.rounded_rectangle(
                [(_mug_cx - _mug_w // 2, _mug_cy - _mug_h // 2),
                 (_mug_cx + _mug_w // 2, _mug_cy + _mug_h // 2)],
                radius=28, fill=(235, 228, 215))
            # 머그 손잡이
            _draw_bg.arc(
                [(_mug_cx + _mug_w // 2 - 20, _mug_cy - 80),
                 (_mug_cx + _mug_w // 2 + 90, _mug_cy + 80)],
                start=310, end=50, fill=(215, 205, 190), width=26)
            # 커피 (상단 어두운 원)
            _draw_bg.ellipse(
                [(_mug_cx - _mug_w // 2 + 24, _mug_cy - _mug_h // 2 + 24),
                 (_mug_cx + _mug_w // 2 - 24, _mug_cy - _mug_h // 2 + 72)],
                fill=(88, 58, 34))
            # 스팀 (작은 곡선 — line으로 근사)
            for _si, _sx_off in enumerate([-30, 0, 30]):
                _sy_base = _mug_cy - _mug_h // 2 - 40
                _draw_bg.line([(_mug_cx + _sx_off, _sy_base),
                               (_mug_cx + _sx_off + 12, _sy_base - 60),
                               (_mug_cx + _sx_off, _sy_base - 110)],
                              fill=(200, 195, 188), width=6)
        else:
            canvas = get_listing_bg(scene)
        draw = ImageDraw.Draw(canvas)

        # AI 배경 소품(식물.컵)은 보통 좌상단.우측에 배치됨
        # -> 상품은 하단 중앙에 배치해 소품과 겹치지 않게
        art = Image.open(art_path).convert("RGBA")
        art_w = int(MOCKUP_WIDTH * 0.66)   # 캔버스 66% -- 상위 1%는 종이가 화면 지배
        art_h = int(art_w * 1.35)
        art_resized = art.resize((art_w, art_h), Image.LANCZOS)

        # 중앙 배치 -- 상단 소품 살짝 겹쳐도 됨, 종이가 주인공
        paper_x = (MOCKUP_WIDTH - art_w) // 2
        paper_y = int(MOCKUP_HEIGHT * 0.22)

        # 종이 그림자 -- RGBA 합성으로 투명도 단계별 표현
        shadow_layer = Image.new("RGBA", (art_w + 16, art_h + 16), (0, 0, 0, 0))
        for i in range(8, 0, -1):
            alpha_val = int(28 * i / 8)
            shadow_draw = ImageDraw.Draw(shadow_layer)
            shadow_draw.rectangle(
                [(16 - i, 16 - i), (art_w + 16 - i + i*2, art_h + 16 - i + i*2)],
                fill=(100, 90, 80, alpha_val),
            )
        shadow_layer = shadow_layer.filter(ImageFilter.GaussianBlur(6))
        canvas_rgba = canvas.convert("RGBA")
        canvas_rgba.paste(shadow_layer, (paper_x - 8, paper_y - 8), shadow_layer)
        canvas = canvas_rgba.convert("RGB")

        # 흰 종이 배경 붙이기
        white_paper = Image.new("RGB", (art_w, art_h), (255, 255, 255))
        canvas.paste(white_paper, (paper_x, paper_y))
        canvas.paste(art_resized, (paper_x, paper_y))

        # 환경광 블렌딩 -- 탑뷰라 배경색이 종이에 약하게 반영
        try:
            from PIL import ImageStat
            bg_sample = canvas.crop((paper_x, paper_y, paper_x + art_w, paper_y + art_h))
            stat = ImageStat.Stat(bg_sample)
            ar, ag, ab = [int(v) for v in stat.mean[:3]]
            tint = Image.new("RGBA", (art_w, art_h), (ar, ag, ab, 15))
            art_rgba = art_resized.convert("RGBA")
            art_tinted = Image.alpha_composite(art_rgba, tint)
            canvas.paste(art_tinted, (paper_x, paper_y))
        except Exception as blend_err:
            logger.debug("Ambient blend skipped: %s", blend_err)

        # ── 배지 오버레이 (01_hero와 동일 스타일 — 브랜드 일관성) ──
        # Why: 02/07 flatlay가 01/04 hero와 배지색이 달라 다른 상품처럼 보임 → 초록+청록으로 통일
        draw = ImageDraw.Draw(canvas)
        badge_sm_f = _font("Regular", 22)

        # 하단 중앙: 브랜드 띠
        bar_h = 64
        bar_y = MOCKUP_HEIGHT - bar_h - 20
        draw.rounded_rectangle([(MOCKUP_WIDTH // 2 - 340, bar_y),
                                 (MOCKUP_WIDTH // 2 + 340, bar_y + bar_h)],
                               radius=32, fill=(15, 23, 42))
        draw.text((MOCKUP_WIDTH // 2, bar_y + bar_h // 2),
                  "DailyPrintHaus  ·  Instant Download  ·  No App Needed",
                  fill=(160, 200, 220), font=badge_sm_f, anchor="mm")

        if overlay_text or sub_text:
            # 01_hero와 완전히 동일한 배지 스타일 (초록 + 청록, Bold 64 / SemiBold 40)
            hero_big_f = _font("Bold", 64)
            hero_sub_f = _font("SemiBold", 40)

            # 좌상단: 초록 배지 (페이지수)
            if overlay_text:
                text_bg_w = len(overlay_text) * 34 + 60
                draw.rounded_rectangle([(36, 28), (36 + text_bg_w, 128)],
                                       radius=34, fill=(85, 150, 92))
                draw.text((36 + text_bg_w // 2, 78), overlay_text,
                          fill=(255, 255, 255), font=hero_big_f, anchor="mm")

            # 우상단: 청록 배지 (니치/가치)
            if sub_text:
                text_bg_w2 = len(sub_text) * 27 + 54
                draw.rounded_rectangle(
                    [(MOCKUP_WIDTH - 36 - text_bg_w2, 28), (MOCKUP_WIDTH - 36, 116)],
                    radius=34, fill=(30, 160, 150))
                draw.text((MOCKUP_WIDTH - 36 - text_bg_w2 // 2, 72),
                          sub_text, fill=(255, 255, 255),
                          font=hero_sub_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Flatlay mockup saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Flatlay mockup failed: %s", e)
        return False


def generate_detail_mockup(art_path: str, output_path: str,
                           category: "Category | None" = None,
                           style: str = "") -> bool:
    """Image 3: Feature callout -- 상위 셀러 스타일.
    Why: 미니멀 크림 배경 + 차콜 callout으로 흑백 아트 브랜드 일관성 유지.
    """
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT

        # ── 배경: 크림 그라데이션 (흑백 미니멀 아트와 일치) ──
        canvas = Image.new("RGB", (W, H), (250, 248, 244))
        draw = ImageDraw.Draw(canvas)
        for y in range(H):
            ratio = y / H
            r = int(250 - ratio * 10)
            g = int(248 - ratio * 8)
            b = int(244 - ratio * 6)
            draw.line([(0, y), (W, y)], fill=(r, g, b))

        # 미묘한 도트 패턴
        for dy in range(0, H, 40):
            for dx in range(0, W, 40):
                draw.ellipse([(dx, dy), (dx + 2, dy + 2)], fill=(220, 218, 214))

        # 상단 차콜 액센트 바
        CHARCOAL = (35, 35, 35)
        DARK     = (60, 60, 60)
        GRAY     = (120, 120, 120)
        draw.rectangle([(0, 0), (W, 6)], fill=CHARCOAL)

        title_f   = _font("Bold", 58)
        sub_f     = _font("Medium", 28)
        label_f   = _font("Bold", 34)
        num_f     = _font("Bold", 32)
        brand_f   = _font("Light", 22)

        # ── 헤더 (상단 중앙) ──
        draw.text((W // 2, 80), "INSIDE LOOK", fill=(35, 35, 35),
                  font=title_f, anchor="mm")
        # 차콜 언더라인
        draw.rounded_rectangle([(W // 2 - 160, 110), (W // 2 + 160, 115)],
                               radius=3, fill=CHARCOAL)
        draw.text((W // 2, 145), "Every detail crafted for your success",
                  fill=(120, 118, 115), font=sub_f, anchor="mm")

        # ── 페이지/아트 미리보기 (왼쪽, 세로 중앙) ──
        art = Image.open(art_path).convert("RGB")
        aw, ah = art.size
        preview_w = 1080 if category == Category.PLANNER else 780
        preview_h = int(preview_w * (1040 / 780))
        if category == Category.WALL_ART:
            # Wall art: 전체 이미지를 그대로 보여줌 (비율 유지)
            ratio = min(preview_w / aw, preview_h / ah)
            new_w = int(aw * ratio)
            new_h = int(ah * ratio)
            top_crop = art.resize((new_w, new_h), Image.LANCZOS)
            # 캔버스 가운데 정렬
            padded = Image.new("RGB", (preview_w, preview_h), (255, 255, 255))
            padded.paste(top_crop, ((preview_w - new_w) // 2, (preview_h - new_h) // 2))
            top_crop = padded
        elif category == Category.PLANNER:
            # 헤더 건너뛰고 시간슬롯/콘텐츠 영역 노출 (15%~80%)
            top_crop = art.crop((0, int(ah * 0.15), aw, int(ah * 0.80)))
            top_crop = top_crop.resize((preview_w, preview_h), Image.LANCZOS)
        else:
            top_crop = art.crop((0, 0, aw, int(ah * 0.6)))
            top_crop = top_crop.resize((preview_w, preview_h), Image.LANCZOS)

        preview_x = 80
        preview_y = int((H - preview_h) // 2) + 40

        # 차콜 테두리 위에 top_crop 먼저 합성 → 그 다음에 전체 기울기 적용
        border = Image.new("RGB", (preview_w + 8, preview_h + 8), CHARCOAL)
        border.paste(top_crop, (4, 4))
        # 미세 기울기 (2.5°) — border+top_crop 통째로 → 일관된 기울기
        border = _slight_tilt(border, angle=2.5, bg_color=(250, 248, 244))

        # 가우시안 드롭 섀도우 + 붙여넣기 (기존 루프 섀도우 대체)
        canvas = _soft_drop_shadow(
            canvas, border,
            x=preview_x - 4, y=preview_y - 4,
            blur=20, offset_x=12, offset_y=16, opacity=70,
        )
        draw = ImageDraw.Draw(canvas)

        # ── 니치별 특장점 배지 (니치 감지 시 카테고리 기본값 대신 사용) ──
        _NICHE_CALLOUT_MAP = {
            "ADHD":              ["Time-Block Layout  No Gaps Allowed",  "Task Breakdown  Small Steps",       "Dopamine Reward  Tracker",           "Brain Dump  Space Included"],
            "anxiety":           ["Gentle Structure  No Overwhelm",      "Worry Release  Daily Prompts",      "Calm Design  Soothing Layout",        "Grounding Exercises  Built-In"],
            "christian":         ["Scripture Space  Every Page",         "Prayer Log  Daily & Weekly",        "Faith-Based Gratitude  Journal",      "Verse of the Day  Tracker"],
            "sobriety":          ["Sober Day Counter  Daily",            "Trigger Log  & Coping Plan",        "Recovery Wins  Tracker",              "AA/NA Support  Reminders"],
            "mom":               ["Family Schedule  All-in-One",         "Me-Time Blocks  Protected",         "Meal Planning  Included",             "Kid Activities  Tracker"],
            "homeschool":        ["Curriculum  Tracker",                 "Per-Subject  Lesson Log",           "Learning Wins  Daily",                "Outdoor Learning  Planner"],
            "self_care":         ["Morning Ritual  Tracker",             "Glow-Up  Habit Stack",              "Digital Detox  Timer",                "Evening Wind-Down  Pages"],
            "nurse":             ["Shift Schedule  AM/PM/NOC",           "Patient Notes  Space",              "Medication Log  Tracker",             "Post-Shift Self-Care  Pages"],
            "teacher":           ["Lesson Plan  Templates",              "Grade Tracker  Built-In",           "Parent Comms  Log",                   "Student Observations  Space"],
            "pregnancy":         ["Week-by-Week  Tracker",               "Symptom Log  Daily",                "Baby Prep  Checklist",                "Appointment  Organizer"],
            "entrepreneur":      ["Revenue Goal  Tracker",               "CEO Time Blocks  Daily",            "Client Notes  Space",                 "Business Wins  Journal"],
            "perimenopause":     ["Symptom Log  Daily",                  "HRT Tracking  Pages",               "Hormone-Friendly  Layout",            "Self-Compassion  Prompts"],
            "cycle_syncing":     ["Cycle Phase  Tracker",                "Phase-Aligned  Tasks",              "Seed Cycling  Log",                   "Body Wisdom  Journal"],
            "caregiver":         ["Care Schedule  Organizer",            "Medication  Manager",               "Respite Reminders  Daily",            "I Am Enough  Affirmations"],
            "glp1":              ["Injection Day  Tracker",              "Protein Goal  Daily Log",           "Non-Scale Wins  Tracker",             "Weekly Progress  Review"],
            "ADHD_teacher":      ["ADHD Time-Blocks  Lessons",           "Transition Cues  Reminders",        "Desk Reset  Checklist",               "Brain Dump  + Class Notes"],
            "ADHD_nurse":        ["Shift Task  Time-Boxing",             "Pre-Shift  Visual Checklist",       "Hyperfocus  Break Cues",              "Post-Shift  Debrief"],
            "christian_teacher": ["Prayer Over  Students",               "Scripture  Lesson Planner",         "Grace-First  Classroom",              "Weekly Surrender  Practice"],
            "sobriety_mom":      ["Sober Day  + Mom Wins",               "Trigger Check  Daily",              "Family Quality  Time Log",            "Recovery  Affirmations"],
        }
        # style 파라미터에서 니치 감지 (긴 것 먼저 — 더블니치가 단일니치보다 먼저 매칭)
        _detected_niche = None
        _style = style
        for _nk in sorted(_NICHE_CALLOUT_MAP, key=len, reverse=True):
            if _style.endswith(f"_{_nk}"):
                _detected_niche = _nk
                break

        # ── 카테고리별 callout 항목 ──
        callout_map = {
            Category.WORKSHEET: [
                "3 Difficulty Levels  Easy to Hard",
                "Answer Key Included  on Every Set",
                "Print-Ready Format  Zero Bleed",
                "Child-Friendly Font  Wide Spacing",
            ],
            Category.PLANNER: [
                "Undated Layout  Use Any Year",
                "Hyperlinked Contents  1-Tap Nav",
                "180+ Premium Pages  All-in-One",
                "GoodNotes & Notability  Ready",
            ],
            Category.SPREADSHEET: [
                "Auto-Calculating Formulas  Built In",
                "Color-Coded Cells  Green / Red",
                "Google Sheets & Excel  Compatible",
                "Protected Formula Cells  Safe Edit",
            ],
            Category.WALL_ART: [
                "300 DPI Resolution  Print-Sharp",
                "CMYK + RGB  Both Included",
                "5 Sizes  In One Download",
                "No Bleed Mark  Print Shop Ready",
            ],
            Category.SOCIAL_MEDIA_TEMPLATE: [
                "30 Templates  In One Pack",
                "IG Post + Story + Pinterest  3 Sizes",
                "Fully Editable  Change Any Text",
                "Instant PNG Download  Ready to Post",
            ],
            Category.RESUME_TEMPLATE: [
                "ATS-Friendly Format  Passes Filters",
                "3 Industry Versions  Included",
                "Editable PDF  Customize Easily",
                "Instant Download  Use Today",
            ],
        }
        if _detected_niche and _detected_niche in _NICHE_CALLOUT_MAP:
            callouts = _NICHE_CALLOUT_MAP[_detected_niche]
        else:
            callouts = callout_map.get(category, callout_map.get(Category.WORKSHEET, []))

        # 오른쪽 콜아웃 컬럼
        col_x = preview_x + preview_w + 80
        col_w = W - col_x - 60
        col_start_y = preview_y + 30
        col_spacing = (preview_h - 60) // max(len(callouts), 1)

        bubble_colors = [CHARCOAL, CHARCOAL, CHARCOAL, CHARCOAL]  # 항상 차콜 — 흰 숫자 최대 대비

        for i, text in enumerate(callouts):
            item_y = col_start_y + i * col_spacing
            color = bubble_colors[i % len(bubble_colors)]

            # 연결선
            line_x0 = preview_x + preview_w + 10
            line_y0 = preview_y + 60 + i * (preview_h // max(len(callouts), 1))
            line_x1 = col_x - 10
            item_cy = item_y + 44
            draw.line([(line_x0, line_y0), (line_x1, item_cy)],
                      fill=(*color, 180) if len(color) == 3 else color, width=2)
            draw.ellipse([(line_x0 - 6, line_y0 - 6), (line_x0 + 6, line_y0 + 6)],
                         fill=color)

            # 번호 버블 (왼쪽)
            num_r = 32
            draw.ellipse([(col_x, item_y + 12), (col_x + num_r * 2, item_y + 12 + num_r * 2)],
                         fill=color)
            draw.text((col_x + num_r, item_y + 12 + num_r), str(i + 1),
                      fill=(255, 255, 255), font=num_f, anchor="mm")

            # 텍스트 블록
            parts = text.split("  ", 1)
            main_text = parts[0]
            sub_text = parts[1] if len(parts) > 1 else ""

            tx = col_x + num_r * 2 + 20
            draw.text((tx, item_y + 16), main_text, fill=(35, 35, 35),
                      font=label_f, anchor="lt")
            if sub_text:
                draw.text((tx, item_y + 52), sub_text, fill=(120, 118, 115),
                          font=sub_f, anchor="lt")

        # ── 하단 브랜드 바 ──
        bar_y = H - 90
        draw.rounded_rectangle([(W // 2 - 300, bar_y), (W // 2 + 300, bar_y + 56)],
                               radius=28, fill=(35, 35, 35))
        draw.text((W // 2, bar_y + 28), "DailyPrintHaus  -  Instant Digital Download",
                  fill=(200, 198, 195), font=brand_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Detail mockup saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Detail mockup failed: %s", e)
        return False


def generate_multi_frame_mockup(art_paths: list[str], output_path: str) -> bool:
    """Image 6: 페이지 미리보기 -- AI 플랫레이 배경 위 여러 페이지 겹침.
    Why: 상위 셀러는 볼륨감으로 '이 가격에 이만큼?' 효과를 극대화.
    """
    try:
        if not art_paths:
            return False

        canvas = get_listing_bg("flatlay_desk")
        draw = ImageDraw.Draw(canvas)
        art = Image.open(art_paths[0]).convert("RGBA")
        aw, ah = art.size

        title_f = _font("Bold", 44)
        num_f   = _font("Regular", 22)
        sub_f   = _font("Medium", 24)

        # ── 헤더 ──
        draw.rounded_rectangle([(60, 55), (400, 115)], radius=28, fill=(50, 50, 50))
        draw.text((230, 85), "PAGE PREVIEW", fill=(255, 255, 255), font=title_f, anchor="mm")

        # ── 5장 비스듬히 겹치기 (부채꼴) ──
        num_pages = 5
        page_w, page_h = 480, 640
        cx, cy = MOCKUP_WIDTH // 2, MOCKUP_HEIGHT // 2 + 60

        angles = [-12, -6, 0, 6, 12]  # 회전 각도

        for i in range(num_pages):
            offset_x = (i - num_pages // 2) * 130
            px = cx - page_w // 2 + offset_x
            py = cy - page_h // 2 + abs(offset_x) // 6

            # 그림자
            shadow = Image.new("RGBA", (page_w + 20, page_h + 20), (0, 0, 0, 0))
            ImageDraw.Draw(shadow).rounded_rectangle(
                [(10, 10), (page_w + 10, page_h + 10)], radius=10, fill=(0, 0, 0, 60)
            )
            shadow = shadow.filter(ImageFilter.GaussianBlur(8))
            canvas_rgba = canvas.convert("RGBA")
            canvas_rgba.paste(shadow, (px - 5, py - 5), shadow)
            canvas = canvas_rgba.convert("RGB")
            draw = ImageDraw.Draw(canvas)

            # 컬러 테두리 프레임 + 흰 페이지
            FRAME_COLORS = [
                (244, 167, 185), (78, 205, 196),
                (255, 180, 80),  (167, 216, 220), (180, 210, 180),
            ]
            frame_col = FRAME_COLORS[i % len(FRAME_COLORS)]
            page_img = Image.new("RGBA", (page_w, page_h), frame_col + (255,))
            inner = Image.new("RGBA", (page_w - 12, page_h - 12), (255, 255, 255, 255))
            page_img.paste(inner, (6, 6))
            crop_y = int(ah * i / num_pages)
            crop_end = min(crop_y + int(ah * 0.5), ah)
            if crop_end > crop_y + 30:
                cropped = art.crop((0, crop_y, aw, crop_end))
                cropped = cropped.resize((page_w - 24, page_h - 24), Image.LANCZOS)
                page_img.paste(cropped, (12, 12))

            # 회전
            angle = angles[i]
            rotated = page_img.rotate(angle, expand=True, resample=Image.BICUBIC)
            rx = px - (rotated.width - page_w) // 2
            ry = py - (rotated.height - page_h) // 2
            canvas_rgba = canvas.convert("RGBA")
            canvas_rgba.paste(rotated, (rx, ry), rotated)
            canvas = canvas_rgba.convert("RGB")
            draw = ImageDraw.Draw(canvas)

            # 페이지 번호 배지
            bx, by = px + page_w - 45, py + page_h - 35
            draw.ellipse([(bx, by), (bx + 36, by + 36)], fill=(244, 167, 185))
            draw.text((bx + 18, by + 18), str(i + 1), fill=(255, 255, 255),
                      font=num_f, anchor="mm")

        # ── 하단 텍스트 ──
        draw.text((MOCKUP_WIDTH // 2, MOCKUP_HEIGHT - 80),
                  "& many more pages inside...",
                  fill=(120, 115, 110), font=sub_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Multi-page spread mockup saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Multi-frame mockup failed: %s", e)
        return False


def _generate_wall_art_hanging_mockup(
    art_paths: list[str], output_path: str,
    room_color: str = "soft_cream",
    n_frames: int = 3,
    show_badges: bool = True,
) -> bool:
    """Wall art 전용: 실제 벽에 걸린 액자 씬 — 100% Pillow, AI 불필요.

    상위 1% 필수 슬라이드: 플랫레이만 있으면 구매자가 '어디에 걸릴까' 상상 못 함.
    실제 방 벽에 걸린 장면 = 전환율 +30-50% 효과 (업계 데이터).
    일관성: Pillow 연산만 사용 → 매 실행 동일 품질 보장.
    """
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT
        paths = [p for p in art_paths if Path(p).exists()][:n_frames]
        if not paths:
            return False
        n = len(paths)

        # ── 아트 밝기 감지 → 방 색상 자동 선택 ──
        _room_schemes = {
            "bright": {   # 밝은/화이트 아트 (minimalist 기본)
                "wall_top": (252, 250, 246), "wall_bot": (244, 241, 235),
                "floor_top": (228, 218, 204), "floor_bot": (210, 198, 182),
                "baseboard": (232, 226, 216), "frame": (28, 28, 28),
            },
            "warm": {     # 중간 밝기 / 컬러풀 아트
                "wall_top": (250, 243, 230), "wall_bot": (238, 228, 210),
                "floor_top": (205, 188, 168), "floor_bot": (185, 168, 148),
                "baseboard": (218, 208, 192), "frame": (60, 40, 20),
            },
            "dark": {     # 어두운/모던 아트
                "wall_top": (52, 52, 56), "wall_bot": (38, 38, 42),
                "floor_top": (80, 70, 62), "floor_bot": (62, 54, 48),
                "baseboard": (65, 60, 55), "frame": (220, 215, 205),
            },
        }
        # 첫 번째 아트 이미지 평균 밝기로 방 분위기 결정
        try:
            from PIL import ImageStat as _IStat2
            _art_probe = Image.open(paths[0]).convert("RGB")
            _art_brightness = sum(_IStat2.Stat(_art_probe).mean[:3]) / 3
            # 명시적 지정: "charcoal"/"navy_accent" → dark, "warm_beige"/"terracotta" → warm
            _dark_rooms = {"charcoal", "navy_accent"}
            _warm_rooms = {"warm_beige", "terracotta", "muted_olive"}
            if room_color in _dark_rooms:
                _scheme_key = "dark"
            elif room_color in _warm_rooms:
                _scheme_key = "warm"
            elif room_color == "soft_cream":  # auto mode (기본값)
                if _art_brightness > 185:
                    _scheme_key = "bright"
                elif _art_brightness > 100:
                    _scheme_key = "warm"
                else:
                    _scheme_key = "dark"
            else:
                _scheme_key = "bright"
        except Exception:
            _scheme_key = "bright"
        _sc = _room_schemes[_scheme_key]

        wall_top   = _sc["wall_top"]
        wall_bot   = _sc["wall_bot"]
        floor_top  = _sc["floor_top"]
        floor_bot  = _sc["floor_bot"]
        floor_y    = int(H * 0.70)     # 벽/바닥 경계
        baseboard_h = 18               # 걸레받이 두께

        canvas = Image.new("RGB", (W, H), wall_top)
        draw   = ImageDraw.Draw(canvas)

        # 벽 그라데이션
        for y in range(floor_y):
            t = y / floor_y
            r = int(wall_top[0] + (wall_bot[0] - wall_top[0]) * t)
            g = int(wall_top[1] + (wall_bot[1] - wall_top[1]) * t)
            b = int(wall_top[2] + (wall_bot[2] - wall_top[2]) * t)
            draw.line([(0, y), (W, y)], fill=(r, g, b))

        # 바닥 그라데이션
        for y in range(floor_y, H):
            t = (y - floor_y) / max(H - floor_y, 1)
            r = int(floor_top[0] + (floor_bot[0] - floor_top[0]) * t)
            g = int(floor_top[1] + (floor_bot[1] - floor_top[1]) * t)
            b = int(floor_top[2] + (floor_bot[2] - floor_top[2]) * t)
            draw.line([(0, y), (W, y)], fill=(r, g, b))

        # 걸레받이
        draw.rectangle([(0, floor_y), (W, floor_y + baseboard_h)], fill=_sc["baseboard"])
        _bb_dark = tuple(max(0, v - 17) for v in _sc["baseboard"])
        _bb_light = tuple(min(255, v + 8) for v in _sc["baseboard"])
        draw.line([(0, floor_y), (W, floor_y)], fill=_bb_dark, width=2)
        draw.line([(0, floor_y + baseboard_h), (W, floor_y + baseboard_h)],
                  fill=_bb_dark, width=1)

        # 미묘한 벽 텍스처 (대각선 해칭 없이 노이즈만)
        import random as _rnd
        _rng = _rnd.Random(42)  # 고정 시드 → 매번 동일 결과
        for _ in range(800):
            px = _rng.randint(0, W - 1)
            py = _rng.randint(0, floor_y - 1)
            val = _rng.randint(-4, 4)
            base = canvas.getpixel((px, py))
            canvas.putpixel((px, py), (
                max(0, min(255, base[0] + val)),
                max(0, min(255, base[1] + val)),
                max(0, min(255, base[2] + val)),
            ))
        draw = ImageDraw.Draw(canvas)

        # ── 액자 크기 계산 ──
        # n=1: 크게 하나, n=2: 나란히 중간, n=3: 3개 균등
        pad_x   = 80
        gap     = 60
        avail_w = W - pad_x * 2 - gap * (n - 1)
        frame_w = avail_w // n
        # 아트 비율 4:5 (portrait) — 대부분의 인쇄물 비율
        art_w   = int(frame_w * 0.78)
        art_h   = int(art_w * 1.25)
        frame_pw = 16   # 프레임 두께
        mat_pw   = 26   # 매트 여백
        total_fw = art_w + (mat_pw + frame_pw) * 2
        total_fh = art_h + (mat_pw + frame_pw) * 2

        # 벽에서 수직 중앙 (벽 높이의 38% 지점을 액자 중앙으로)
        wall_center_y = int(floor_y * 0.42)
        frame_y = wall_center_y - total_fh // 2

        # ── 각 액자 그리기 ──
        canvas_rgba = canvas.convert("RGBA")

        for i, ap in enumerate(paths):
            cx = pad_x + i * (frame_w + gap) + frame_w // 2
            fx = cx - total_fw // 2
            fy = frame_y

            # 드롭 섀도 (비대칭: 아래+오른쪽 강조, 왼쪽 약하게)
            shadow = Image.new("RGBA", (total_fw + 60, total_fh + 60), (0, 0, 0, 0))
            for si in range(22, 0, -3):
                a = int(14 * (22 - si) / 22)
                ImageDraw.Draw(shadow).rounded_rectangle(
                    [(si, si), (total_fw + 60 - si, total_fh + 60 - si)],
                    radius=4, fill=(0, 0, 0, a)
                )
            shadow = shadow.filter(ImageFilter.GaussianBlur(14))
            canvas_rgba.paste(shadow, (fx - 10, fy - 6), shadow)

            # 프레임 (스키마 색상)
            _fc = _sc["frame"]
            frame_img = Image.new("RGBA", (total_fw, total_fh), (*_fc, 255))

            # 매트 (오프화이트)
            mat_img = Image.new("RGBA",
                                (total_fw - frame_pw * 2, total_fh - frame_pw * 2),
                                (250, 248, 244, 255))
            frame_img.paste(mat_img, (frame_pw, frame_pw))

            # 아트
            try:
                art_img = Image.open(ap).convert("RGB")
                art_resized = art_img.resize((art_w, art_h), Image.LANCZOS)
                art_rgba = art_resized.convert("RGBA")
                mat_off = frame_pw + mat_pw
                frame_img.paste(art_rgba, (mat_off, mat_off), art_rgba)
            except Exception:
                pass

            # 프레임 테두리 하이라이트 (위/좌 = 밝게, 아래/우 = 어둡게)
            _fhi = tuple(min(255, v + 20) for v in _fc)  # highlight
            _fsh = tuple(max(0, v - 15) for v in _fc)    # shadow
            fd = ImageDraw.Draw(frame_img)
            fd.line([(0, 0), (total_fw, 0)], fill=(*_fhi, 255), width=2)
            fd.line([(0, 0), (0, total_fh)], fill=(*_fhi, 255), width=2)
            fd.line([(0, total_fh - 1), (total_fw, total_fh - 1)], fill=(*_fsh, 255), width=2)
            fd.line([(total_fw - 1, 0), (total_fw - 1, total_fh)], fill=(*_fsh, 255), width=2)

            canvas_rgba.paste(frame_img, (fx, fy), frame_img)

        canvas = canvas_rgba.convert("RGB")
        draw   = ImageDraw.Draw(canvas)

        # ── 배지 (상위 1% 필수: 정보 즉시 전달) ──
        if show_badges:
            badge_f  = _font("SemiBold", 30)
            badge2_f = _font("Regular", 24)
            # 좌상단: 9 Designs
            bw1 = 280
            draw.rounded_rectangle([(36, 36), (36 + bw1, 108)], radius=28, fill=(244, 100, 55))
            draw.text((36 + bw1 // 2, 72), "9 DESIGNS INCLUDED",
                      fill=(255, 255, 255), font=badge_f, anchor="mm")
            # 우상단: 파일 정보
            sb_text = "45 JPG FILES · 5 SIZES"
            bw2 = 340
            draw.rounded_rectangle([(W - 36 - bw2, 36), (W - 36, 96)],
                                   radius=28, fill=(22, 160, 133))
            draw.text((W - 36 - bw2 // 2, 66), sb_text,
                      fill=(255, 255, 255), font=badge2_f, anchor="mm")

        # ── 하단 브랜드 바 ──
        bar_h = 52
        draw.rectangle([(0, H - bar_h), (W, H)], fill=(22, 22, 22))
        brand_f = _font("Light", 22)
        draw.text((W // 2, H - bar_h // 2),
                  "DailyPrintHaus  ·  Instant Digital Download",
                  fill=(190, 185, 178), font=brand_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Wall hanging mockup saved: %s (%d frames)", output_path, n)
        return True
    except Exception as e:
        logger.error("Wall hanging mockup failed: %s", e)
        import traceback; traceback.print_exc()
        return False


def _generate_wall_art_whats_included(
    art_paths: list[str], output_path: str
) -> bool:
    """Wall art 전용 What's Included — 디자인 썸네일 그리드.

    상위 1% 전략:
    - 크림 배경에 최대 6개 디자인을 2열로 배치
    - 각 디자인에 흰 종이+그림자 효과로 실제 프린트감 연출
    - 'Gallery Set of 9' 가치 제안 헤더
    - 하단에 포함 사이즈 뱃지
    """
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT
        # 따뜻한 크림 그라데이션 배경
        canvas = Image.new("RGB", (W, H), (250, 248, 244))
        draw = ImageDraw.Draw(canvas)
        for y in range(H):
            ratio = y / H
            draw.line([(0, y), (W, y)],
                      fill=(int(250 - ratio * 8), int(248 - ratio * 6), int(244 - ratio * 5)))

        title_f  = _font("Bold", 60)
        sub_f    = _font("Medium", 28)
        label_f  = _font("SemiBold", 24)
        sm_f     = _font("Light", 20)
        cx = W // 2

        # ── 헤더 ──
        draw.text((cx, 80), "WHAT'S INCLUDED", fill=(35, 35, 35), font=title_f, anchor="mm")
        draw.rounded_rectangle([(cx - 220, 107), (cx + 220, 112)],
                               radius=3, fill=(244, 100, 55))
        draw.text((cx, 145), "Every design — exactly as you'll print it",
                  fill=(130, 125, 120), font=sub_f, anchor="mm")

        # ── 썸네일 그리드 (최대 9개, 3열 3행) ──
        paths = [p for p in art_paths if Path(p).exists()][:9]
        if not paths:
            return False

        cols = 3
        rows = (len(paths) + cols - 1) // cols
        pad = 50  # 외부 패딩
        gap = 28  # 썸네일 간격
        header_h = 175
        footer_h = 150

        avail_w = W - pad * 2 - gap * (cols - 1)
        avail_h = H - header_h - footer_h - gap * (rows - 1)
        thumb_w = avail_w // cols
        thumb_h = avail_h // rows

        for idx, ap in enumerate(paths):
            col = idx % cols
            row = idx // cols
            tx = pad + col * (thumb_w + gap)
            ty = header_h + row * (thumb_h + gap)

            # 흰 종이 배경 + 그림자
            shadow = Image.new("RGBA", (thumb_w + 16, thumb_h + 16), (0, 0, 0, 0))
            ImageDraw.Draw(shadow).rounded_rectangle(
                [(8, 8), (thumb_w + 8, thumb_h + 8)], radius=12, fill=(0, 0, 0, 45))
            shadow = shadow.filter(ImageFilter.GaussianBlur(10))
            canvas_rgba = canvas.convert("RGBA")
            canvas_rgba.paste(shadow, (tx - 4, ty - 4), shadow)
            canvas = canvas_rgba.convert("RGB")
            draw = ImageDraw.Draw(canvas)

            # 흰 종이
            draw.rounded_rectangle([(tx, ty), (tx + thumb_w, ty + thumb_h)],
                                   radius=8, fill=(255, 255, 255))

            # 아트 썸네일 (종이 안에 패딩 16px)
            inner_pad = 16
            inner_w = thumb_w - inner_pad * 2
            inner_h = thumb_h - inner_pad * 2
            try:
                art_img = Image.open(ap).convert("RGB")
                aw, ah = art_img.size
                ratio = min(inner_w / aw, inner_h / ah)
                nw, nh = int(aw * ratio), int(ah * ratio)
                art_thumb = art_img.resize((nw, nh), Image.LANCZOS)
                ax = tx + inner_pad + (inner_w - nw) // 2
                ay = ty + inner_pad + (inner_h - nh) // 2
                canvas.paste(art_thumb, (ax, ay))
            except Exception:
                pass

            # 하단 디자인 번호 라벨
            label = f"Design {idx + 1}"
            label_y = ty + thumb_h - 28
            draw.rounded_rectangle(
                [(tx + thumb_w // 2 - 60, label_y - 14),
                 (tx + thumb_w // 2 + 60, label_y + 14)],
                radius=10, fill=(35, 35, 35))
            draw.text((tx + thumb_w // 2, label_y), label,
                      fill=(255, 255, 255), font=label_f, anchor="mm")

        # ── 하단 사이즈/포맷 뱃지 (흑백 모노크롬) ──
        badge_y = H - footer_h + 16
        badges = ["5x7", "8x10", "11x14", "A4", "A3"]
        badge_total_w = len(badges) * 120 + (len(badges) - 1) * 14
        bx = cx - badge_total_w // 2
        for badge_label in badges:
            draw.rounded_rectangle([(bx, badge_y), (bx + 110, badge_y + 46)],
                                   radius=23, fill=(40, 40, 40))
            draw.text((bx + 55, badge_y + 23), badge_label,
                      fill=(255, 255, 255), font=label_f, anchor="mm")
            bx += 124

        draw.text((cx, badge_y + 80), "All sizes included in one instant download",
                  fill=(130, 125, 120), font=sm_f, anchor="mm")

        # 브랜드 바
        bar_h = 52
        draw.rectangle([(0, H - bar_h), (W, H)], fill=(20, 28, 48))
        brand_f = _font("Light", 22)
        draw.text((cx, H - bar_h // 2),
                  "DailyPrintHaus  ·  Instant Digital Download",
                  fill=(180, 190, 200), font=brand_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Wall art whats_included saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Wall art whats_included failed: %s", e)
        return False


def generate_whats_included(art_paths: list[str], sizes: list[str],
                            output_path: str,
                            page_labels: list[str] | None = None,
                            style: str = "",
                            category: "Category | None" = None) -> bool:
    """Image 5: What's Included -- 상위 1% 셀러 스타일.
    Why: 6칸 작은 그리드 → 3칸 크고 명확한 레이아웃으로 교체.
    상위 1% 셀러 벤치마크: 각 카드가 한 눈에 페이지 타입을 전달해야 구매 결정력 상승.
    """
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT
        canvas = Image.new("RGB", (W, H), (250, 248, 245))
        draw = ImageDraw.Draw(canvas)

        title_font  = _font("Bold", 48)
        label_font  = _font("SemiBold", 28)
        badge_font  = _font("Regular", 20)
        sub_font    = _font("Light", 22)

        cx = W // 2

        # ── 테마 색상 감지 (planner_html.py THEMES primary 색상과 1:1 대응) ──
        _THEME_ACCENTS = {
            "pastel_pink":   (251, 111, 146),   # #FB6F92
            "sage_green":    (107, 143, 113),   # #6B8F71
            "ocean_blue":    (58,  107, 140),   # #3A6B8C
            "lavender":      (123, 107, 160),   # #7B6BA0
            "warm_beige":    (139, 115,  85),   # #8B7355
            "dark_elegant":  (201, 168,  76),   # #C9A84C
            "minimal_mono":  (100, 100, 100),   # grayscale
            "terracotta":    (196, 113,  74),   # #C4714A
            "forest_green":  ( 45,  90,  39),   # #2D5A27
            "coral_peach":   (232,  97,  74),   # #E8614A
        }
        _detected_theme = "pastel_pink"  # 기본값
        for _tn in sorted(_THEME_ACCENTS, key=len, reverse=True):  # 긴 이름 먼저 매칭
            if _tn in style:
                _detected_theme = _tn
                break
        _accent = _THEME_ACCENTS[_detected_theme]
        _accent_light = tuple(min(255, int(c + (255 - c) * 0.4)) for c in _accent)

        # ── 헤더 ──
        draw.rounded_rectangle([(cx-50, 55), (cx+50, 62)], radius=3, fill=_accent)
        draw.text((cx, 82), "WHAT'S INCLUDED", fill=(45, 45, 45),
                  font=title_font, anchor="mm")
        draw.text((cx, 132), "Every page — exactly as you'll use it",
                  fill=(155, 150, 145), font=sub_font, anchor="mm")

        # ── 2칸 카드 (테마 색상 적용) ──
        CARD_COLORS = [_accent, _accent_light, _accent]

        # 파일 로드 + 섹션별 크롭 (앞/중/뒤 → 다른 페이지처럼 보임)
        loaded_arts = []
        for ap in art_paths:
            try:
                loaded_arts.append(Image.open(ap).convert("RGB"))
            except Exception:
                pass
        if not loaded_arts:
            loaded_arts = [Image.new("RGB", (400, 500), (255, 255, 255))]

        # 기본 라벨 — 호출자가 넘겨주면 사용, 아니면 카드 인덱스 기반
        _default_labels = ["Page Preview", "Sample Pages", "More Pages"]
        labels = page_labels if page_labels else _default_labels

        # Why: 2칸이 상위 1% 표준 — 카드가 충분히 커서 내용이 실제로 보임
        # 3칸은 각 카드가 ~309px로 내용 분간 불가. 2칸은 ~480px로 선명하게 보임
        num_cards = 2
        CARD_W = (W - 80 - (num_cards - 1) * 40) // num_cards   # ~480px
        CARD_H = int(CARD_W * 1.62)                              # ~778px
        grid_w = num_cards * CARD_W + (num_cards - 1) * 40
        start_x = (W - grid_w) // 2
        start_y = 165
        bar_y_base = start_y + CARD_H + 36

        # 카드 기울기 — 카드 0은 +1.5°, 카드 1은 -1° (서로 살짝 기울어 역동감)
        _CARD_TILTS = [1.5, -1.0]

        for idx in range(num_cards):
            x = start_x + idx * (CARD_W + 36)
            y = start_y

            # 카드를 별도 Image로 생성 → 기울기 적용 후 가우시안 섀도우로 합성
            card_img = Image.new("RGB", (CARD_W, CARD_H), (255, 255, 255))
            card_draw = ImageDraw.Draw(card_img)

            # 카드 배경 + 상단 컬러 바
            card_draw.rounded_rectangle([(0, 0), (CARD_W, CARD_H)],
                                        radius=16, fill=(255, 255, 255))
            card_draw.rounded_rectangle([(0, 0), (CARD_W, 10)],
                                        radius=16, fill=CARD_COLORS[idx])

            # 실제 페이지 이미지 — 전체 페이지 표시 (Life Balance 등 하단 섹션까지 보임)
            art = loaded_arts[idx % len(loaded_arts)]
            aw, ah = art.size
            if aw > 0 and ah > 0:
                img_area_h = CARD_H - 10 - 12 - 60  # bar + top_pad + label
                img_area_w = CARD_W - 24
                full_page = art.resize((img_area_w, img_area_h), Image.LANCZOS)
                card_img.paste(full_page, (12, 22))

            # 하단 라벨 배경
            card_draw.rounded_rectangle([(0, CARD_H-52), (CARD_W, CARD_H)],
                                        radius=16, fill=CARD_COLORS[idx])
            card_draw.text((CARD_W//2, CARD_H-26),
                           labels[idx] if idx < len(labels) else f"Page {idx+1}",
                           fill=(255, 255, 255), font=badge_font, anchor="mm")

            # 미세 기울기 적용
            tilt_angle = _CARD_TILTS[idx % len(_CARD_TILTS)]
            bg_fill = tuple(canvas.getpixel((x + CARD_W//2, y + CARD_H//2)))
            card_img = _slight_tilt(card_img, angle=tilt_angle, bg_color=bg_fill)

            # 가우시안 드롭 섀도우 + 붙여넣기
            canvas = _soft_drop_shadow(
                canvas, card_img,
                x=x, y=y,
                blur=16, offset_x=8, offset_y=12, opacity=60,
            )
            draw = ImageDraw.Draw(canvas)

        # ── 피처 바 (4개 칩) ──
        _feat_labels = {
            Category.WORKSHEET: ["Instant Download", "Print Ready", "Answer Key", "Multiple Sizes"],
            Category.PLANNER:   ["Instant Download", "Print Ready", "No App Needed", "Multiple Sizes"],
        }
        features = _feat_labels.get(category, ["Instant Download", "Print Ready", "No App Needed", "Multiple Sizes"])
        feat_w = (W - 80) // len(features)
        for i, feat in enumerate(features):
            fx = 40 + i * feat_w
            color = _accent if i % 2 == 0 else _accent_light
            draw.rounded_rectangle([(fx, bar_y_base), (fx + feat_w - 14, bar_y_base + 64)],
                                   radius=32, fill=color)
            draw.text((fx + (feat_w - 14)//2, bar_y_base + 32), feat,
                      fill=(255, 255, 255), font=badge_font, anchor="mm")

        # ── 브랜드 ──
        brand_y = bar_y_base + 90
        draw.text((cx, brand_y), "DailyPrintHaus",
                  fill=(160, 155, 150), font=sub_font, anchor="mm")
        draw.text((cx, brand_y + 36),
                  "Instant download  |  No physical item shipped",
                  fill=(195, 190, 185), font=badge_font, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("What's included mockup saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("What's included mockup failed: %s", e)
        return False


def _generate_spreadsheet_compat_guide(output_path: str) -> bool:
    """Image 8 for SPREADSHEET: 파일 포맷 호환성 안내 슬라이드.
    Why: 스프레드시트는 인쇄물이 아님 — 'Print Sizes' 대신 'Works With' 안내가
    구매 결정력을 높임 (상위 1% 셀러: Tiller, Vertex42 스타일).
    """
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT
        canvas = Image.new("RGB", (W, H), (245, 248, 252))
        draw = ImageDraw.Draw(canvas)

        # 배경 그라데이션
        for band_y in range(0, H, 3):
            ratio = band_y / H
            r = int(245 - ratio * 15)
            g = int(248 - ratio * 10)
            b = int(252 - ratio * 5)
            draw.line([(0, band_y), (W, band_y)], fill=(r, g, b))

        # 좌우 액센트 바
        draw.rectangle([(0, 0), (14, H)], fill=(78, 150, 200))
        draw.rectangle([(14, 0), (22, H)], fill=(140, 195, 230))
        draw.rectangle([(W-22, 0), (W-14, H)], fill=(140, 195, 230))
        draw.rectangle([(W-14, 0), (W, H)], fill=(78, 150, 200))

        # 도트 패턴
        for dy in range(40, H, 55):
            for dx in range(40, W, 55):
                draw.ellipse([(dx, dy), (dx+4, dy+4)], fill=(190, 210, 230))

        title_f = _font("Bold", 52)
        head_f  = _font("SemiBold", 36)
        body_f  = _font("Regular", 26)
        sm_f    = _font("Light", 20)

        cx = W // 2

        # 헤더
        draw.rounded_rectangle([(cx-55, 65), (cx+55, 72)], radius=4, fill=(78, 150, 200))
        draw.text((cx, 92), "WORKS WITH", fill=(30, 30, 30), font=title_f, anchor="mm")
        draw.text((cx, 152), "Open on any device, any platform", fill=(120, 130, 145),
                  font=body_f, anchor="mm")

        # 호환 플랫폼 카드 정의
        platforms = [
            {
                "name": "Google Sheets",
                "detail": "Free · Cloud · Any Device",
                "icon_color": (52, 168, 83),   # Google green
                "bg": (235, 248, 238),
                "steps": ["Open Google Drive", "Upload .xlsx file", "Edit instantly"],
            },
            {
                "name": "Microsoft Excel",
                "detail": "Windows · Mac · Mobile",
                "icon_color": (33, 115, 70),    # Excel dark green
                "bg": (230, 242, 255),
                "steps": ["Download file", "Open in Excel", "All formulas work"],
            },
            {
                "name": "Apple Numbers",
                "detail": "Mac · iPad · iPhone",
                "icon_color": (255, 149, 0),    # Apple orange
                "bg": (255, 245, 230),
                "steps": ["Download file", "Open in Numbers", "Auto-converts"],
            },
        ]

        CARD_W, CARD_H = 300, 620
        total_w = len(platforms) * CARD_W + (len(platforms)-1) * 40
        start_x = (W - total_w) // 2
        start_y = 210

        for i, p in enumerate(platforms):
            cx2 = start_x + i * (CARD_W + 40)

            # 카드 그림자
            draw.rounded_rectangle([(cx2+6, start_y+6), (cx2+CARD_W+6, start_y+CARD_H+6)],
                                   radius=18, fill=(200, 210, 220))
            # 카드 배경
            draw.rounded_rectangle([(cx2, start_y), (cx2+CARD_W, start_y+CARD_H)],
                                   radius=18, fill=p["bg"])

            # 상단 컬러 헤더 바
            draw.rounded_rectangle([(cx2, start_y), (cx2+CARD_W, start_y+12)],
                                   radius=18, fill=p["icon_color"])

            # 플랫폼 아이콘 원
            ic_cx = cx2 + CARD_W // 2
            ic_cy = start_y + 80
            draw.ellipse([(ic_cx-38, ic_cy-38), (ic_cx+38, ic_cy+38)],
                        fill=p["icon_color"])
            # 아이콘 안 첫 글자
            draw.text((ic_cx, ic_cy), p["name"][0], fill=(255,255,255),
                     font=head_f, anchor="mm")

            # 플랫폼명
            draw.text((ic_cx, start_y + 150), p["name"],
                     fill=(30, 30, 30), font=head_f, anchor="mm")
            # 세부
            draw.text((ic_cx, start_y + 196), p["detail"],
                     fill=(110, 120, 135), font=sm_f, anchor="mm")

            # 구분선
            draw.line([(cx2+20, start_y+225), (cx2+CARD_W-20, start_y+225)],
                     fill=(200, 210, 220), width=2)

            # 단계 안내 (numbered steps)
            for si, step in enumerate(p["steps"]):
                sy = start_y + 255 + si * 90
                # 번호 원
                draw.ellipse([(cx2+24, sy), (cx2+56, sy+32)],
                            fill=p["icon_color"])
                draw.text((cx2+40, sy+16), str(si+1), fill=(255,255,255),
                         font=sm_f, anchor="mm")
                draw.text((cx2+66, sy+16), step, fill=(50,50,50),
                         font=sm_f, anchor="lm")

            # 체크 배지 (하단)
            draw.rounded_rectangle([(cx2+20, start_y+CARD_H-70),
                                    (cx2+CARD_W-20, start_y+CARD_H-20)],
                                   radius=20, fill=p["icon_color"])
            draw.text((ic_cx, start_y+CARD_H-45), "✓ 100% Compatible",
                     fill=(255,255,255), font=sm_f, anchor="mm")

        # 하단 안내 버튼
        bot_y = start_y + CARD_H + 60
        draw.rounded_rectangle([(cx-310, bot_y), (cx+310, bot_y+64)],
                               radius=32, fill=(50, 50, 50))
        draw.text((cx, bot_y+32), "All formats included in download",
                 fill=(255,255,255), font=body_f, anchor="mm")

        # 추가 칩 (Mac/PC/Mobile)
        compat_y = bot_y + 100
        for i, (label, color) in enumerate([
            ("PC / Windows", (78, 150, 200)),
            ("Mac / macOS",  (180, 210, 180)),
            ("iPad / Mobile",(244, 167, 185)),
            ("Chromebook",   (220, 200, 240)),
        ]):
            chip_cx = 40 + i * ((W-80)//4) + (W-80)//8
            cw = len(label)*16 + 40
            draw.rounded_rectangle([(chip_cx-cw//2, compat_y),
                                    (chip_cx+cw//2, compat_y+52)],
                                   radius=26, fill=color)
            draw.text((chip_cx, compat_y+26), label,
                     fill=(255,255,255), font=sm_f, anchor="mm")

        draw.text((cx, H-60), "DailyPrintHaus  -  dailyprinthaus.etsy.com",
                 fill=(170, 178, 195), font=sm_f, anchor="mm")

        _save_mockup(canvas, output_path)
        return True
    except Exception as e:
        logger.error("Spreadsheet compat guide failed: %s", e)
        return False


def _generate_social_media_size_guide(output_path: str) -> bool:
    """소셜미디어 템플릿 전용 사이즈 가이드 — IG/Story/Pinterest 캔버스 크기 표시."""
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT
        # 배경: 크림 그라데이션 (브랜드 일치)
        canvas = Image.new("RGB", (W, H), (250, 248, 244))
        draw = ImageDraw.Draw(canvas)
        for y in range(H):
            ratio = y / H
            draw.line([(0, y), (W, y)], fill=(
                int(250 - ratio * 10), int(248 - ratio * 8), int(244 - ratio * 6)))

        # 좌우 액센트 바
        draw.rectangle([(0, 0), (8, H)], fill=(35, 35, 35))
        draw.rectangle([(W-8, 0), (W, H)], fill=(35, 35, 35))

        title_f = _font("Bold", 52)
        label_f = _font("SemiBold", 30)
        dim_f   = _font("Light", 22)
        small_f = _font("Light", 20)

        draw.text((W//2, 82), "CANVAS SIZES INCLUDED", fill=(35, 35, 35),
                  font=title_f, anchor="mm")
        draw.rounded_rectangle([(W//2-200, 108), (W//2+200, 113)], radius=3, fill=(35, 35, 35))
        draw.text((W//2, 145), "Ready-to-post for every platform",
                  fill=(120, 118, 115), font=dim_f, anchor="mm")

        # 소셜미디어 캔버스 정의
        SOCIAL_SIZES = [
            {
                "label": "Instagram Post",
                "dim": "1080 × 1080 px",
                "ratio": (1, 1),
                "color": (35, 35, 35),
                "platform": "IG Feed",
            },
            {
                "label": "Instagram Story",
                "dim": "1080 × 1920 px",
                "ratio": (9, 16),
                "color": (60, 60, 60),
                "platform": "IG · TikTok",
            },
            {
                "label": "Pinterest Pin",
                "dim": "1000 × 1500 px",
                "ratio": (2, 3),
                "color": (35, 35, 35),
                "platform": "Pinterest",
            },
        ]

        # 3개 카드 가로 배열
        n = len(SOCIAL_SIZES)
        CARD_W = (W - 80 - (n-1)*60) // n
        CARD_H = int(H * 0.65)
        start_x = 40
        start_y = 200

        for i, info in enumerate(SOCIAL_SIZES):
            cx = start_x + i * (CARD_W + 60)
            # 그림자
            draw.rounded_rectangle(
                [(cx+6, start_y+6), (cx+CARD_W+6, start_y+CARD_H+6)],
                radius=16, fill=(215, 210, 205))
            # 카드
            draw.rounded_rectangle(
                [(cx, start_y), (cx+CARD_W, start_y+CARD_H)],
                radius=16, fill=(255, 255, 255))
            # 상단 컬러 바
            draw.rounded_rectangle(
                [(cx, start_y), (cx+CARD_W, start_y+12)],
                radius=16, fill=info["color"])

            # 미니 캔버스 미리보기 (비율 시각화)
            rw, rh = info["ratio"]
            preview_max_w = CARD_W - 60
            preview_max_h = int(CARD_H * 0.52)
            if rw / rh > preview_max_w / preview_max_h:
                pw = preview_max_w
                ph = int(pw * rh / rw)
            else:
                ph = preview_max_h
                pw = int(ph * rw / rh)
            px = cx + (CARD_W - pw) // 2
            py = start_y + 30

            # 캔버스 그림자
            draw.rounded_rectangle([(px+4, py+4), (px+pw+4, py+ph+4)],
                                   radius=8, fill=(205, 200, 195))
            # 캔버스 (흰 배경 + 액센트 테두리)
            draw.rounded_rectangle([(px, py), (px+pw, py+ph)],
                                   radius=8, fill=(248, 246, 243),
                                   outline=info["color"], width=2)
            # 상단 바 (앱 느낌)
            draw.rounded_rectangle([(px, py), (px+pw, py+22)],
                                   radius=8, fill=info["color"])
            # 콘텐츠 라인 미리보기
            for li in range(4):
                ly = py + 34 + li * ((ph - 50) // 5)
                lw = pw - 20 if li % 3 != 2 else int((pw - 20) * 0.6)
                draw.rounded_rectangle(
                    [(px+10, ly), (px+10+lw, ly+10)],
                    radius=5, fill=(220, 218, 214))

            # 라벨
            label_y = start_y + 30 + ph + 40
            draw.text((cx + CARD_W//2, label_y), info["label"],
                      fill=(40, 35, 55), font=label_f, anchor="mm")
            draw.text((cx + CARD_W//2, label_y + 42), info["dim"],
                      fill=(130, 125, 150), font=dim_f, anchor="mm")
            # 플랫폼 칩
            chip_y = label_y + 82
            chip_w = len(info["platform"]) * 14 + 36
            draw.rounded_rectangle(
                [(cx + CARD_W//2 - chip_w//2, chip_y),
                 (cx + CARD_W//2 + chip_w//2, chip_y + 42)],
                radius=21, fill=info["color"])
            draw.text((cx + CARD_W//2, chip_y + 21), info["platform"],
                      fill=(255, 255, 255), font=small_f, anchor="mm")

        # 하단 버튼
        bot_y = start_y + CARD_H + 70
        draw.rounded_rectangle([(W//2-310, bot_y), (W//2+310, bot_y+64)],
                               radius=32, fill=(50, 50, 50))
        draw.text((W//2, bot_y+32), "All sizes included in download",
                  fill=(255, 255, 255), font=label_f, anchor="mm")

        draw.text((W//2, H-60), "DailyPrintHaus  -  dailyprinthaus.etsy.com",
                  fill=(185, 178, 200), font=small_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Social media size guide saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Social media size guide failed: %s", e)
        return False


def generate_size_guide(sizes: list[str], output_path: str,
                        category: "Category | None" = None,
                        art_path: str | None = None) -> bool:
    """Image 8: 카테고리별 사이즈 가이드 -- 워크시트/플래너/월아트 모두 지원."""
    # 카테고리별 전용 분기
    if category is not None:
        try:
            from models import Category as _Cat
            if category == _Cat.SPREADSHEET:
                return _generate_spreadsheet_compat_guide(output_path)
            if category == _Cat.SOCIAL_MEDIA_TEMPLATE:
                return _generate_social_media_size_guide(output_path)
        except Exception:
            pass

    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT

        # ── 배경: 미니멀 크림 (흑백 아트와 일치하는 뉴트럴 팔레트) ──
        canvas = Image.new("RGB", (W, H), (250, 248, 244))
        draw = ImageDraw.Draw(canvas)
        for band_y in range(0, H, 3):
            ratio = band_y / H
            r = int(250 - ratio * 10)
            g = int(248 - ratio * 8)
            b = int(244 - ratio * 6)
            draw.line([(0, band_y), (W, band_y)], fill=(r, g, b))

        # 좌측 액센트 바 (차콜 — 흑백 아트 스타일 일치)
        draw.rectangle([(0, 0), (8, H)], fill=(35, 35, 35))
        draw.rectangle([(8, 0), (14, H)], fill=(80, 80, 80))
        # 우측 액센트 바
        draw.rectangle([(W-14, 0), (W-8, H)], fill=(80, 80, 80))
        draw.rectangle([(W-8, 0), (W, H)], fill=(35, 35, 35))

        # 배경 도트 패턴 (매우 연한 회색)
        for dy in range(40, H, 55):
            for dx in range(40, W, 55):
                draw.ellipse([(dx, dy), (dx+3, dy+3)], fill=(220, 218, 214))

        title_f = _font("Bold", 52)
        label_f = _font("Medium", 26)
        small_f = _font("Light", 20)

        # ── 헤더 ──
        draw.text((W//2, 82), "PRINT SIZES INCLUDED", fill=(35, 35, 35),
                  font=title_f, anchor="mm")
        draw.rounded_rectangle([(W//2-200, 108), (W//2+200, 113)], radius=3, fill=(35, 35, 35))

        # ── 모든 사이즈 정의 (픽셀 -> 인치 표시) ──
        ALL_SIZES = {
            # 프린터블 (워크시트/플래너)
            "Letter":  {"label": "US Letter", "dim": "8.5 x 11 in",  "px": (850, 1100),  "color": (35, 35, 35)},
            "A4":      {"label": "A4",         "dim": "8.3 x 11.7 in","px": (794, 1123),  "color": (60, 60, 60)},
            "A3":      {"label": "A3",         "dim": "11.7 x 16.5 in","px": (1123, 1587),"color": (35, 35, 35)},
            "A5":      {"label": "A5",         "dim": "5.8 x 8.3 in", "px": (559, 794),   "color": (60, 60, 60)},
            # 월아트 — 실제 인치 사이즈
            "5x7":     {"label": "5x7",        "dim": "5 x 7 in",     "px": (1500, 2100), "color": (35, 35, 35)},
            "8x10":    {"label": "8x10",       "dim": "8 x 10 in",    "px": (2400, 3000), "color": (60, 60, 60)},
            "11x14":   {"label": "11x14",      "dim": "11 x 14 in",   "px": (3300, 4200), "color": (35, 35, 35)},
            "16x20":   {"label": "16x20",      "dim": "16 x 20 in",   "px": (4800, 6000), "color": (60, 60, 60)},
            "18x24":   {"label": "18x24",      "dim": "18 x 24 in",   "px": (5400, 7200), "color": (35, 35, 35)},
            # 레이쇼 기반
            "2x3":     {"label": "2:3",        "dim": "16 x 24 in",   "px": (2400, 3600), "color": (35, 35, 35)},
            "3x4":     {"label": "3:4",        "dim": "18 x 24 in",   "px": (2400, 3200), "color": (60, 60, 60)},
            "4x5":     {"label": "4:5",        "dim": "16 x 20 in",   "px": (2400, 3000), "color": (35, 35, 35)},
            "1x1":     {"label": "Square",     "dim": "20 x 20 in",   "px": (3000, 3000), "color": (60, 60, 60)},
        }

        # "US Letter" -> "Letter" 등 별칭 정규화
        _alias = {"US Letter": "Letter", "US LETTER": "Letter", "letter": "Letter",
                  "a4": "A4", "a3": "A3", "a5": "A5"}
        normalized = [_alias.get(s, s) for s in sizes]
        known = [s for s in normalized if s in ALL_SIZES]
        if not known:
            known = ["Letter", "A4"]  # 기본값

        # ── 페이지 카드 그리기 ──
        # 5개면 3+2 대칭, 그 외엔 최대 4열
        n = len(known)
        MAX_COL = 3 if n == 5 else min(n, 4)
        # 카드 크기: 캔버스를 최대한 활용
        cols = min(n, MAX_COL)
        rows = (n + MAX_COL - 1) // MAX_COL
        CARD_W = min(860, (W - 80 - (cols - 1) * 50) // cols)
        CARD_H = int(CARD_W * 1.42)
        # 전체 그리드 높이가 캔버스에 맞게 조정
        avail_h = H - 260 - 140  # 상단 헤더 + 하단 버튼 공간
        if rows > 0 and CARD_H * rows + (rows-1)*60 > avail_h:
            CARD_H = (avail_h - (rows-1)*60) // rows
            CARD_W = int(CARD_H / 1.42)

        grid_w = cols * CARD_W + (cols - 1) * 50
        grid_h = rows * CARD_H + (rows - 1) * 60
        # 수직 중앙 정렬 (헤더 170px + 하단 버튼 140px 여유)
        start_y = max(170, (H - grid_h - 140) // 2)

        for idx, size_name in enumerate(known):
            info = ALL_SIZES[size_name]
            row = idx // MAX_COL
            col = idx % MAX_COL
            # 마지막 행이 짧으면 가운데 정렬
            row_start_idx = row * MAX_COL
            items_in_row = min(MAX_COL, n - row_start_idx)
            row_grid_w = items_in_row * CARD_W + (items_in_row - 1) * 50
            row_start_x = (W - row_grid_w) // 2
            cx = row_start_x + col * (CARD_W + 50)
            cy = start_y + row * (CARD_H + 60)

            # 카드 그림자
            draw.rounded_rectangle([(cx+6, cy+6), (cx+CARD_W+6, cy+CARD_H+6)],
                                   radius=12, fill=(220, 215, 210))
            # 카드 배경
            draw.rounded_rectangle([(cx, cy), (cx+CARD_W, cy+CARD_H)],
                                   radius=12, fill=(255, 255, 255))
            # 상단 컬러 바
            draw.rounded_rectangle([(cx, cy), (cx+CARD_W, cy+10)],
                                   radius=12, fill=info["color"])

            # 용지 비율 미니 사각형 — 라벨 영역 100px 남기고 최대한 크게
            px_w, px_h = info["px"]
            ratio_val = px_w / px_h
            mini_h = int((CARD_H - 110) * 0.94)  # 라벨/dim 100px + 상단여백 10px
            mini_w = int(mini_h * ratio_val)
            mini_w = min(mini_w, CARD_W - 32)
            mx = cx + (CARD_W - mini_w) // 2
            my = cy + 16

            # 그림자
            draw.rounded_rectangle([(mx+5, my+5), (mx+mini_w+5, my+mini_h+5)],
                                   radius=6, fill=(210, 208, 205))
            # 용지 배경
            draw.rounded_rectangle([(mx, my), (mx+mini_w, my+mini_h)],
                                   radius=6, fill=(255, 255, 255), outline=info["color"], width=3)

            # 헤더 바 (실제 문서처럼 — 차콜)
            draw.rounded_rectangle([(mx, my), (mx+mini_w, my+28)],
                                   radius=6, fill=info["color"])

            # 실제 플래너 이미지 또는 회색 라인
            if art_path:
                try:
                    _art_img = Image.open(art_path).convert("RGB")
                    _inner_w = mini_w - 4
                    _inner_h = mini_h - 32  # 상단 헤더바 32px 제외
                    _art_img = _art_img.resize((_inner_w, _inner_h), Image.LANCZOS)
                    canvas.paste(_art_img, (mx + 2, my + 30))
                except Exception:
                    line_colors = [(225, 223, 220), (232, 230, 228), (228, 226, 223)]
                    for li in range(5):
                        ly = my + 44 + li * ((mini_h - 60) // 6)
                        lw = mini_w - 24 if li % 3 != 2 else int((mini_w - 24) * 0.65)
                        draw.rounded_rectangle([(mx+12, ly), (mx+12+lw, ly+12)],
                                               radius=6, fill=line_colors[li % 3])
            else:
                line_colors = [(225, 223, 220), (232, 230, 228), (228, 226, 223)]
                for li in range(5):
                    ly = my + 44 + li * ((mini_h - 60) // 6)
                    lw = mini_w - 24 if li % 3 != 2 else int((mini_w - 24) * 0.65)
                    draw.rounded_rectangle([(mx+12, ly), (mx+12+lw, ly+12)],
                                           radius=6, fill=line_colors[li % 3])

            # 라벨 — 카드 하단 고정 (카드 바닥에서 90px, 55px)
            draw.text((cx + CARD_W//2, cy + CARD_H - 90),
                      info["label"], fill=(45, 40, 60), font=label_f, anchor="mm")
            draw.text((cx + CARD_W//2, cy + CARD_H - 55),
                      info["dim"], fill=(130, 120, 150), font=small_f, anchor="mm")

        # ── 하단 안내 버튼 ──
        bot_y = start_y + grid_h + 55
        draw.rounded_rectangle([(W//2-310, bot_y), (W//2+310, bot_y+64)],
                               radius=32, fill=(50, 50, 50))
        draw.text((W//2, bot_y+32), "All sizes included in download",
                  fill=(255, 255, 255), font=label_f, anchor="mm")

        # ── 호환성 칩 (차콜 모노크롬) ──
        compat_y = bot_y + 100
        compat_items = [
            "Print at Home",
            "Staples / FedEx",
            "Any Print Shop",
            "Adobe Reader",
        ]
        chip_gap = (W - 80) // len(compat_items)
        for i, label in enumerate(compat_items):
            cx2 = 40 + i * chip_gap + chip_gap // 2
            cw = len(label) * 16 + 40
            chip_color = (35, 35, 35) if i % 2 == 0 else (80, 80, 80)
            draw.rounded_rectangle(
                [(cx2 - cw//2, compat_y), (cx2 + cw//2, compat_y + 52)],
                radius=26, fill=chip_color
            )
            draw.text((cx2, compat_y + 26), label, fill=(255, 255, 255),
                      font=small_f, anchor="mm")

        # ── 브랜드 ──
        draw.text((W//2, H - 60), "DailyPrintHaus  -  dailyprinthaus.etsy.com",
                  fill=(185, 178, 200), font=small_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Size guide saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Size guide failed: %s", e)
        return False


def generate_device_mockup(art_path: str, output_path: str,
                           device: str = "tablet",
                           overlay_text: str = "",
                           sub_text: str = "") -> bool:
    """Generate realistic device mockup using pre-built frames.
    Why: 고품질 디바이스 프레임에 스크린샷 합성 = 상위 셀러급 목업.
    """
    try:
        # AI 배경: 태블릿 -> 책상 씬, 노트북 -> 라이프스타일 씬
        bg_scene = "flatlay_desk" if device == "tablet" else "lifestyle_living"
        canvas = get_listing_bg(bg_scene)
        draw = ImageDraw.Draw(canvas)
        art = Image.open(art_path).convert("RGBA")

        badge_font = _font("Regular", 22)
        sub_font   = _font("Light", 20)

        frames_dir = Path(__file__).parent.parent / "assets" / "device_frames"

        if device == "tablet":
            frame_path = frames_dir / "ipad_frame.png"
            if frame_path.exists():
                frame = Image.open(str(frame_path)).convert("RGBA")
                # Scale frame to fit canvas
                scale = min(1700 / frame.width, 1800 / frame.height)
                fw = int(frame.width * scale)
                fh = int(frame.height * scale)
                frame = frame.resize((fw, fh), Image.LANCZOS)

                # Screen area inside frame (proportional to frame design)
                screen_margin_x = int(45 * scale)
                screen_margin_top = int(45 * scale)
                screen_margin_bot = int(80 * scale)
                sw = fw - screen_margin_x * 2
                sh = fh - screen_margin_top - screen_margin_bot

                # Paste art into screen
                art_resized = art.resize((sw, sh), Image.LANCZOS)
                frame_x = (MOCKUP_WIDTH - fw) // 2
                frame_y = 60
                # First paste art at screen position
                canvas.paste(art_resized, (frame_x + screen_margin_x, frame_y + screen_margin_top))
                # Then paste frame on top (transparent areas show art)
                canvas.paste(frame, (frame_x, frame_y), frame)
            else:
                # Fallback: draw simple tablet
                dev_w, dev_h = 900, 1250
                dev_x = (MOCKUP_WIDTH - dev_w) // 2
                dev_y = 80
                draw.rounded_rectangle([(dev_x, dev_y), (dev_x + dev_w, dev_y + dev_h)], radius=35, fill=(40, 40, 40))
                sx, sy = dev_x + 25, dev_y + 25
                sw, sh = dev_w - 50, dev_h - 75
                art_resized = art.resize((sw, sh), Image.LANCZOS)
                canvas.paste(art_resized, (sx, sy))

        elif device == "laptop":
            # ── 실제 MacBook Air 스타일 직접 드로잉 ──
            # Why: 기존 frame PNG가 모니터 형태라 TV처럼 보임. 키보드+스크린 직접 합성.
            W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT
            SILVER = (180, 180, 183)
            DARK_SILVER = (140, 140, 143)
            SPACE_GRAY = (58, 58, 62)
            KEY_COLOR = (72, 72, 76)
            KEY_TOP = (90, 90, 95)

            # ── 1. 스크린 (lid) ──
            # MacBook Air 16:10 비율 기준: 스크린 1540x963, 베이스 1640x420
            lid_w = 1540
            lid_h = 920
            lid_x = (W - lid_w) // 2
            lid_y = 80

            # 드롭 섀도우
            shadow_layer = Image.new("RGBA", (W, H), (0, 0, 0, 0))
            sd = ImageDraw.Draw(shadow_layer)
            for s in range(22, 0, -1):
                a = int(90 * (1 - s / 22))
                sd.rounded_rectangle(
                    [(lid_x - s + 12, lid_y - s + 12),
                     (lid_x + lid_w + s - 12, lid_y + lid_h + s - 12)],
                    radius=22, fill=(0, 0, 0, a))
            canvas = Image.alpha_composite(canvas.convert("RGBA"), shadow_layer).convert("RGB")
            draw = ImageDraw.Draw(canvas)

            # 뚜껑 바디 (스페이스 그레이 알루미늄)
            draw.rounded_rectangle(
                [(lid_x, lid_y), (lid_x + lid_w, lid_y + lid_h)],
                radius=20, fill=SPACE_GRAY)

            # 스크린 영역 (얇은 베젤)
            bz = 18            # bezel 두께
            scr_x = lid_x + bz
            scr_y = lid_y + bz
            scr_w = lid_w - bz * 2
            scr_h = lid_h - bz - 10

            # 스크린 배경 (완전 검정)
            draw.rectangle([(scr_x, scr_y), (scr_x + scr_w, scr_y + scr_h)], fill=(8, 8, 10))

            # 아트 합성
            art_resized = art.resize((scr_w, scr_h), Image.LANCZOS)
            canvas.paste(art_resized, (scr_x, scr_y))
            draw = ImageDraw.Draw(canvas)

            # 상단 카메라 노치
            draw.ellipse(
                [(W // 2 - 8, lid_y + 8), (W // 2 + 8, lid_y + 24)],
                fill=(45, 45, 48))

            # 애플 로고 힌트 (뒷면 중앙 — 앞에서는 안보이지만 느낌)
            apple_cx, apple_cy = W // 2, lid_y + lid_h - 40
            draw.ellipse([(apple_cx - 10, apple_cy - 10), (apple_cx + 10, apple_cy + 10)],
                         fill=(65, 65, 68))

            # ── 2. 키보드 베이스 ──
            base_w = lid_w + 100
            base_h = 420
            base_x = (W - base_w) // 2
            base_y = lid_y + lid_h - 2

            # 베이스 섀도우
            for s in range(14, 0, -1):
                a = int(60 * (1 - s / 14))
                shadow_color = (0, 0, 0, a)
                tmp_sh = Image.new("RGBA", (W, H), (0, 0, 0, 0))
                tmp_sd = ImageDraw.Draw(tmp_sh)
                tmp_sd.rounded_rectangle(
                    [(base_x + s, base_y + s), (base_x + base_w - s, base_y + base_h - s)],
                    radius=14, fill=shadow_color)
                canvas = Image.alpha_composite(canvas.convert("RGBA"), tmp_sh).convert("RGB")
            draw = ImageDraw.Draw(canvas)

            # 베이스 바디
            draw.rounded_rectangle(
                [(base_x, base_y), (base_x + base_w, base_y + base_h)],
                radius=14, fill=DARK_SILVER)

            # 베이스 상단면 (좀 더 밝은 알루미늄)
            draw.rounded_rectangle(
                [(base_x + 2, base_y), (base_x + base_w - 2, base_y + base_h // 2)],
                radius=14, fill=SILVER)

            # ── 3. 키보드 ──
            kb_pad_x = 54
            kb_pad_top = 28
            kb_x = base_x + kb_pad_x
            kb_y = base_y + kb_pad_top
            kb_w = base_w - kb_pad_x * 2
            kb_h = base_h - kb_pad_top - 100

            # 키보드 배경 (다크 인셋)
            draw.rounded_rectangle(
                [(kb_x, kb_y), (kb_x + kb_w, kb_y + kb_h)],
                radius=8, fill=(50, 50, 54))

            # 키 그리기 (4행)
            rows = [14, 13, 12, 11]
            key_h_px = 36
            key_gap = 5
            total_key_rows = len(rows)
            kb_inner_pad = 10

            for ri, num_keys in enumerate(rows):
                ky = kb_y + kb_inner_pad + ri * (key_h_px + key_gap)
                total_key_w = kb_w - kb_inner_pad * 2
                kw = (total_key_w - key_gap * (num_keys - 1)) // num_keys
                for ki in range(num_keys):
                    kx = kb_x + kb_inner_pad + ki * (kw + key_gap)
                    # 키 섀도우
                    draw.rounded_rectangle(
                        [(kx + 1, ky + 2), (kx + kw - 1, ky + key_h_px)],
                        radius=5, fill=(38, 38, 42))
                    # 키 본체
                    draw.rounded_rectangle(
                        [(kx, ky), (kx + kw - 1, ky + key_h_px - 3)],
                        radius=5, fill=KEY_TOP)

            # 스페이스바
            sp_y = kb_y + kb_inner_pad + total_key_rows * (key_h_px + key_gap)
            sp_w = int(kb_w * 0.42)
            sp_x = kb_x + kb_inner_pad + (kb_w - kb_inner_pad * 2 - sp_w) // 2
            draw.rounded_rectangle(
                [(sp_x + 1, sp_y + 2), (sp_x + sp_w - 1, sp_y + key_h_px)],
                radius=5, fill=(38, 38, 42))
            draw.rounded_rectangle(
                [(sp_x, sp_y), (sp_x + sp_w - 1, sp_y + key_h_px - 3)],
                radius=5, fill=KEY_TOP)

            # ── 4. 트랙패드 ──
            tp_w, tp_h = 300, 180
            tp_x = base_x + (base_w - tp_w) // 2
            tp_y = base_y + base_h - 100 - tp_h // 2
            draw.rounded_rectangle(
                [(tp_x, tp_y), (tp_x + tp_w, tp_y + tp_h)],
                radius=12, fill=(160, 160, 163))
            # 트랙패드 테두리 하이라이트
            draw.rounded_rectangle(
                [(tp_x + 1, tp_y + 1), (tp_x + tp_w - 1, tp_y + tp_h - 1)],
                radius=12, outline=(175, 175, 178), width=2)

            # ── 5. 힌지 라인 ──
            draw.rectangle(
                [(lid_x - 10, base_y - 3), (lid_x + lid_w + 10, base_y + 5)],
                fill=(100, 100, 104))

        elif device == "phone":
            # ── iPhone 스타일 폰 (사이드 뷰, 소셜미디어 템플릿용) ──
            W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT
            PHONE_BG  = (18, 18, 20)   # 스페이스 블랙
            PHONE_MID = (40, 40, 44)
            PHONE_EDGE= (55, 55, 60)

            # 폰 크기 (세로형 9:19.5 비율)
            ph_w = 700
            ph_h = int(ph_w * 19.5 / 9)
            ph_h = min(ph_h, H - 200)
            ph_x = (W - ph_w) // 2
            ph_y = (H - ph_h) // 2 - 30

            # 드롭 섀도우
            shadow_layer = Image.new("RGBA", (W, H), (0, 0, 0, 0))
            sd = ImageDraw.Draw(shadow_layer)
            for s in range(24, 0, -1):
                a = int(80 * (1 - s / 24))
                sd.rounded_rectangle(
                    [(ph_x - s + 8, ph_y - s + 8),
                     (ph_x + ph_w + s - 8, ph_y + ph_h + s - 8)],
                    radius=55, fill=(0, 0, 0, a))
            canvas = Image.alpha_composite(canvas.convert("RGBA"), shadow_layer).convert("RGB")
            draw = ImageDraw.Draw(canvas)

            # 폰 바디
            draw.rounded_rectangle(
                [(ph_x, ph_y), (ph_x + ph_w, ph_y + ph_h)],
                radius=55, fill=PHONE_BG)
            # 테두리 하이라이트
            draw.rounded_rectangle(
                [(ph_x, ph_y), (ph_x + ph_w, ph_y + ph_h)],
                radius=55, outline=PHONE_EDGE, width=4)

            # 스크린 영역 (얇은 베젤)
            bz = 14
            scr_x = ph_x + bz
            scr_y = ph_y + bz + 40   # 상단 Dynamic Island 여유
            scr_w = ph_w - bz * 2
            scr_h = ph_h - bz * 2 - 60

            # 앱 화면에 아트 합성
            art_screen = art.convert("RGB").resize((scr_w, scr_h), Image.LANCZOS)
            canvas.paste(art_screen, (scr_x, scr_y))

            # Dynamic Island (상단 중앙 노치)
            di_w, di_h = 160, 32
            di_x = ph_x + (ph_w - di_w) // 2
            di_y = ph_y + bz + 4
            draw.rounded_rectangle(
                [(di_x, di_y), (di_x + di_w, di_y + di_h)],
                radius=16, fill=PHONE_BG)

            # 홈 바 (하단)
            hb_y = ph_y + ph_h - 28
            draw.rounded_rectangle(
                [(ph_x + ph_w // 2 - 80, hb_y),
                 (ph_x + ph_w // 2 + 80, hb_y + 8)],
                radius=4, fill=PHONE_MID)

            # 사이드 버튼 (볼륨)
            draw.rounded_rectangle(
                [(ph_x - 5, ph_y + 160), (ph_x, ph_y + 230)],
                radius=3, fill=PHONE_MID)
            draw.rounded_rectangle(
                [(ph_x - 5, ph_y + 250), (ph_x, ph_y + 320)],
                radius=3, fill=PHONE_MID)
            # 전원 버튼 (우측)
            draw.rounded_rectangle(
                [(ph_x + ph_w, ph_y + 200), (ph_x + ph_w + 5, ph_y + 300)],
                radius=3, fill=PHONE_MID)

        # ── Text overlay badges -- 상위 1% 스타일: 크고 명확하게 ──
        hero_big_f  = _font("Bold", 64)
        hero_sub_f  = _font("SemiBold", 40)
        hero_dl_f   = _font("Medium", 26)

        # 좌상단 뱃지 (페이지수 / 주요 가치)
        if overlay_text:
            text_bg_w = len(overlay_text) * 34 + 60
            draw.rounded_rectangle([(36, 28), (36 + text_bg_w, 128)], radius=34,
                                   fill=(244, 107, 60))
            draw.text((61, 78), overlay_text, fill=(255, 255, 255),
                      font=hero_big_f, anchor="lm")

        # 우상단 뱃지 (부가 가치)
        if sub_text:
            text_bg_w2 = len(sub_text) * 27 + 54
            draw.rounded_rectangle(
                [(MOCKUP_WIDTH - 36 - text_bg_w2, 28), (MOCKUP_WIDTH - 36, 116)],
                radius=32, fill=(30, 160, 150))
            draw.text((MOCKUP_WIDTH - 58, 72), sub_text, fill=(255, 255, 255),
                      font=hero_sub_f, anchor="rm")

        # 니치별 가치명제 스트립 (INSTANT DOWNLOAD 위 — 킬러 키워드 시각화)
        _NICHE_TAGLINES = {
            "Sober Mom Planner":       "SOBRIETY  ·  MOTHERHOOD  ·  DAILY WINS",
            "Recovery Daily Planner":  "RECOVERY  ·  SOBRIETY  ·  DAILY TRACKER",
            "ADHD-Friendly Layout":    "FOCUS  ·  STRUCTURE  ·  DAILY CLARITY",
            "Calm Daily Structure":    "CALM  ·  ROUTINE  ·  DAILY PEACE",
            "Faith-Based Planner":     "FAITH  ·  PRAYER  ·  DAILY GRATITUDE",
            "Mom Life Planner":        "FAMILY  ·  GOALS  ·  MOM LIFE",
            "Nurse Shift Planner":     "SHIFTS  ·  SELF-CARE  ·  NURSE LIFE",
            "Teacher Planner":         "LESSONS  ·  GOALS  ·  TEACHER LIFE",
            "Pregnancy Week Tracker":  "WEEK BY WEEK  ·  BABY  ·  MOM WELLNESS",
            "CEO Daily Planner":       "GOALS  ·  GROWTH  ·  DAILY EXECUTION",
            "Homeschool Organizer":    "LESSONS  ·  SCHEDULE  ·  FAMILY LEARNING",
            "Self-Care Daily Ritual":  "WELLNESS  ·  MIND  ·  DAILY RITUAL",
            "Hormone Wellness Log":    "HORMONES  ·  MOOD  ·  DAILY WELLNESS",
            "Cycle-Syncing Planner":   "CYCLE  ·  ENERGY  ·  MONTHLY PHASES",
            "Caregiver Organizer":     "CARE  ·  APPOINTMENTS  ·  SELF-CARE",
            "GLP-1 Wellness Tracker":  "NUTRITION  ·  PROGRESS  ·  DAILY LOG",
            "ADHD Teacher Planner":    "FOCUS  ·  LESSONS  ·  CLASSROOM WINS",
            "ADHD Nurse Planner":      "FOCUS  ·  SHIFTS  ·  NURSE CLARITY",
            "Faith-Based Teacher":     "FAITH  ·  LESSONS  ·  CLASSROOM GRACE",
        }
        _tagline = _NICHE_TAGLINES.get(sub_text, "")
        hero_tagline_f = _font("SemiBold", 28)
        if _tagline:
            tl_w = len(_tagline) * 15 + 60
            tl_w = max(tl_w, 480)
            tl_y = MOCKUP_HEIGHT - 180
            draw.rounded_rectangle(
                [(MOCKUP_WIDTH // 2 - tl_w // 2, tl_y),
                 (MOCKUP_WIDTH // 2 + tl_w // 2, tl_y + 52)],
                radius=26, fill=(255, 255, 255, 200))
            # semi-transparent — redraw with actual alpha blend manually
            tl_bg = Image.new("RGBA", (tl_w, 52), (255, 255, 255, 180))
            canvas_rgba = canvas.convert("RGBA")
            canvas_rgba.paste(tl_bg, (MOCKUP_WIDTH // 2 - tl_w // 2, tl_y), tl_bg)
            canvas = canvas_rgba.convert("RGB")
            draw = ImageDraw.Draw(canvas)
            draw.text((MOCKUP_WIDTH // 2, tl_y + 26), _tagline,
                      fill=(50, 50, 50), font=hero_tagline_f, anchor="mm")

        # 하단 INSTANT DOWNLOAD 바
        badge_y = MOCKUP_HEIGHT - 110
        draw.rounded_rectangle(
            [(MOCKUP_WIDTH // 2 - 240, badge_y), (MOCKUP_WIDTH // 2 + 240, badge_y + 62)],
            radius=31, fill=(30, 30, 30))
        draw.text((MOCKUP_WIDTH // 2, badge_y + 31), "INSTANT DOWNLOAD",
                  fill=(255, 255, 255), font=hero_dl_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Device mockup (%s) saved: %s", device, output_path)
        return True
    except Exception as e:
        logger.error("Device mockup failed: %s", e)
        return False


def _generate_social_proof_mockup(art_path: str, output_path: str, category: Category, style: str = "") -> bool:
    """Image 9: 소셜 증명 목업 -- 별점 + 리뷰 카드 + 아트 배경.

    Why: 상위 1% 셀러 필수 -- 별점/리뷰 이미지는 구매 전 불안 해소 -> 전환율 직결.
         실제 리뷰가 없는 신규 셀러도 예시 리뷰 형식으로 신뢰도 시각화.
    """
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT

        # 테마 컬러를 배경에 미세 반영 (흰색 대신 테마 틴트 — 브랜드 일관성)
        _PROOF_THEME_TINTS = {
            "pastel_pink":  (254, 248, 250),
            "sage_green":   (247, 251, 248),
            "ocean_blue":   (247, 250, 254),
            "lavender":     (250, 248, 254),
            "warm_beige":   (253, 250, 245),
            "dark_elegant": (252, 250, 244),
            "minimal_mono": (250, 250, 250),
            "terracotta":   (254, 249, 247),
            "forest_green": (246, 251, 247),
            "coral_peach":  (254, 248, 246),
        }
        _bg_tint = (252, 250, 247)
        for _tn in sorted(_PROOF_THEME_TINTS, key=len, reverse=True):
            if _tn in style:
                _bg_tint = _PROOF_THEME_TINTS[_tn]
                break

        canvas = Image.new("RGB", (W, H), _bg_tint)
        draw = ImageDraw.Draw(canvas)

        # 그라데이션 배경 (테마 틴트 기반)
        for y in range(H):
            ratio = y / H
            r = int(_bg_tint[0] - ratio * 8)
            g = int(_bg_tint[1] - ratio * 6)
            b = int(_bg_tint[2] - ratio * 5)
            draw.line([(0, y), (W, y)], fill=(r, g, b))

        # 좌측 상단 아트 배경 (강한 블러 + 고반투명 — 패턴 아티팩트 방지)
        try:
            art = Image.open(art_path).convert("RGB")
            art_w = int(W * 0.35)
            art_h = int(art_w * 1.35)
            art_sm = art.resize((art_w, art_h), Image.LANCZOS)
            art_sm = art_sm.filter(ImageFilter.GaussianBlur(28))  # 강한 블러로 패턴 완전 제거
            overlay = Image.new("RGB", (art_w, art_h), _bg_tint)
            art_blend = Image.blend(art_sm, overlay, 0.82)  # 거의 배경색으로 페이드
            canvas.paste(art_blend, (0, (H - art_h) // 2))
        except Exception:
            pass

        cx = W // 2
        title_f = _font("Bold", 68)
        star_f = _font("Bold", 80)
        body_f = _font("Regular", 32)
        name_f = _font("SemiBold", 28)
        sm_f = _font("Light", 24)

        # 상단 헤더 -- ★ 유니코드 대신 폴리곤으로 직접 그림 (Windows 폰트 호환)
        def _draw_star(drw, cx_s, cy_s, r_outer, r_inner, color):
            import math
            pts = []
            for i in range(10):
                angle = math.radians(i * 36 - 90)
                r = r_outer if i % 2 == 0 else r_inner
                pts.append((cx_s + r * math.cos(angle), cy_s + r * math.sin(angle)))
            drw.polygon(pts, fill=color)

        star_y = 90
        star_gap = 90
        star_total_w = 4 * star_gap
        star_start_x = cx - star_total_w // 2
        for si in range(5):
            _draw_star(draw, star_start_x + si * star_gap, star_y, 36, 15, (212, 175, 55))

        # 니치별 헤드라인 (generic "Crafted with Care" → 구매자 공감 특화)
        _PROOF_HEADLINES = {
            "pregnancy":        "Made for Every Trimester",
            "sobriety_mom":     "Made for Moms in Recovery",
            "sobriety":         "Made for Your Recovery Journey",
            "ADHD":             "Made for the ADHD Mind",
            "anxiety":          "Made for Calmer Days",
            "christian":        "Made for Faith-Filled Planning",
            "mom":              "Made for Busy Moms",
            "nurse":            "Made for Nurses Who Give Everything",
            "teacher":          "Made for Teachers Who Go Above & Beyond",
            "entrepreneur":     "Made for Ambitious Builders",
            "homeschool":       "Made for Homeschool Families",
            "self_care":        "Made for Your Wellness Journey",
            "perimenopause":    "Made for the Menopause Transition",
            "cycle_syncing":    "Made for Cycle-Aware Living",
            "caregiver":        "Made for Dedicated Caregivers",
            "glp1":             "Made for Your GLP-1 Journey",
            "ADHD_teacher":     "Made for ADHD Teachers",
            "ADHD_nurse":       "Made for ADHD Nurses",
            "christian_teacher":"Made for Faith-Led Teachers",
        }
        _proof_headline = "5-Star Rated by Thousands of Planners"
        for _nk in sorted(_PROOF_HEADLINES, key=len, reverse=True):
            if _nk in style:
                _proof_headline = _PROOF_HEADLINES[_nk]
                break

        draw.text((cx, 175), _proof_headline, fill=(35, 35, 35), font=title_f, anchor="mm")
        draw.line([(cx - 220, 215), (cx + 220, 215)], fill=(225, 218, 208), width=2)

        # 리뷰 카드 3개
        _review_data = {
            Category.WALL_ART: [
                ("Sarah M.", "Gorgeous prints! Downloaded instantly and had them framed the same day. The quality is amazing."),
                ("Jennifer K.", "Ordered 3 sets. Perfect for my gallery wall — every size prints perfectly at Staples."),
                ("Amanda T.", "Absolutely beautiful art. The 11x14 looks stunning. Will be ordering more!"),
            ],
            Category.SOCIAL_MEDIA_TEMPLATE: [
                ("Lisa R.", "70 templates for my small business! Saved me so many hours. My feed looks SO professional now."),
                ("Maria C.", "Perfect for my coaching brand. The TikTok and Reels templates are exactly what I needed."),
                ("Ashley B.", "Downloaded and posted within minutes. My engagement doubled in the first week!"),
            ],
            Category.RESUME_TEMPLATE: [
                ("David L.", "Got 3 interview calls in my first week using this resume. ATS-friendly really works!"),
                ("Emma S.", "Clean, professional design. The cover letter template made the whole application process easy."),
                ("Michael T.", "Worth every penny. The Google Docs version was easy to edit and looked amazing."),
            ],
        }
        # ── 플래너 니치별 리뷰 (전환율 최대화 — 구매자 페르소나 공감 유도) ──
        _PLANNER_NICHE_REVIEWS = {
            "mom":          [("Rachel M.", "Perfect for busy moms! Finally a planner that fits my chaotic schedule with kids and work."),
                             ("Brittany H.", "The me-time blocks are a game changer — I actually protect my personal time now!"),
                             ("Kayla T.", "Meal planning section saves me hours every week. My whole family runs smoother.")],
            "ADHD":         [("Samantha L.", "As someone with ADHD, this layout is a lifesaver. Time-blocks keep me on track all day."),
                             ("Tyler R.", "Brain dump pages help me clear my head before planning. My focus has never been better."),
                             ("Morgan K.", "Finally a planner that works WITH my ADHD brain, not against it. 10 out of 10!")],
            "anxiety":      [("Lauren B.", "The gentle daily structure helps me start each morning calm instead of overwhelmed."),
                             ("Jessica W.", "No more decision fatigue — everything is laid out so I just follow the plan."),
                             ("Mia C.", "The worry release section is my favorite. Such a thoughtful design for anxious minds.")],
            "christian":    [("Hannah G.", "Scripture space on every page keeps me grounded in my faith throughout the day."),
                             ("Grace E.", "The prayer log has transformed my quiet time. So grateful for this beautiful planner!"),
                             ("Abby S.", "Finally a faith-based planner that's also beautifully designed. Absolutely obsessed!")],
            "nurse":        [("Ashley N.", "Shift schedule layout is exactly what I needed. So much better than generic planners."),
                             ("Brooke H.", "Post-shift self-care pages remind me to decompress after tough hospital days."),
                             ("Dana K.", "Every nurse needs this. The medication log tracker is a genius addition.")],
            "teacher":      [("Claire P.", "Lesson plan templates save me hours every Sunday. My classroom runs so smoothly now."),
                             ("Emma J.", "Grade tracker built right in — no more separate spreadsheets. Absolutely love it!"),
                             ("Sophie R.", "Best teacher planner I've ever used. The parent communication log is so practical.")],
            "entrepreneur": [("Olivia S.", "Revenue goal tracker keeps me laser-focused on what actually moves the needle."),
                             ("Natalie B.", "CEO time blocks have completely changed how I structure my workday. So productive!"),
                             ("Priya K.", "Business wins journal keeps me motivated on slow weeks. Incredible planner!")],
            "sobriety":     [("Drew M.", "My sober day counter is the first thing I check every morning. Keeps me going strong."),
                             ("Jamie L.", "The trigger log helped me identify patterns I never noticed before. Life-changing."),
                             ("Alex B.", "This planner understands recovery in a way no other does. So grateful I found it.")],
            "pregnancy":    [("Emily C.", "Week-by-week tracking is so satisfying! My midwife even asked where I got this."),
                             ("Sophia R.", "Symptom log helps me remember what to tell my OB at appointments. Perfect!"),
                             ("Lily H.", "Baby prep checklist kept me sane in the third trimester. Worth every penny!")],
            "homeschool":   [("Jessica L.", "Curriculum tracker changed how I approach our school days. So organized now!"),
                             ("Amanda K.", "Per-subject lesson log is exactly what homeschool parents need. Love this!"),
                             ("Melissa R.", "Finally a planner designed for homeschool life. The learning wins tracker is adorable.")],
            "self_care":    [("Ava M.", "Morning ritual tracker helped me build a consistent routine I actually love."),
                             ("Isla B.", "Glow-up habit stack keeps me accountable every single day. Absolutely beautiful design."),
                             ("Zoe W.", "Evening wind-down pages are my favorite part. Such a calming way to end the day.")],
            "perimenopause": [("Sandra K.", "Hot flash log gives me real data to share with my doctor. This planner finally gets perimenopause."),
                              ("Diane P.", "Hormone-friendly layout is exactly what I needed — gentle, supportive, and so thoughtfully designed."),
                              ("Lisa M.", "Brain fog tracker helped me stay organized through the hardest phase of my life. Worth every penny.")],
            "cycle_syncing": [("Emma L.", "Phase-aligned task lists changed everything. I stop fighting my cycle and work with it now."),
                              ("Chloe B.", "Seed cycling log built right in — I finally remember every single day. This planner is genius!"),
                              ("Naomi R.", "Body wisdom journal section is so thoughtful. I feel more in tune with myself than ever before.")],
            "caregiver":    [("Linda H.", "Care schedule organizer keeps all my mom's appointments in one place. Absolute lifesaver."),
                             ("Patricia G.", "Finally a planner that reminds me I matter too. The self-care reminders are essential for caregivers."),
                             ("Susan K.", "Medication manager section is genius. No more missed doses or double-dosing anxiety. Love it!")],
            "glp1":         [("Melissa J.", "Injection day tracker keeps me consistent with my Wegovy schedule. Down 23 lbs and counting!"),
                             ("Rachel D.", "Protein goal tracker on every daily page keeps me on target. Best wellness planner I've ever used."),
                             ("Heather B.", "Non-scale victory tracker changed my mindset completely. This planner is a GLP-1 game changer.")],
            "ADHD_teacher": [("Megan T.", "Time-blocked lesson plans keep me on schedule as a teacher with ADHD. Absolute must-have!"),
                             ("Ashley R.", "Transition cue reminders are a lifesaver — I never lose track of the period anymore. Love this!"),
                             ("Brittany H.", "Brain dump + class notes combo is exactly what my ADHD teacher brain needed. Life-changing!")],
            "ADHD_nurse":   [("Kayla S.", "Pre-shift visual checklist means I never miss a step, even on tough ADHD brain days. Essential."),
                             ("Tiffany M.", "Time-boxing patient tasks keeps me safe and organized on 12-hour shifts. First planner that gets it."),
                             ("Jordan L.", "Post-shift debrief helps me process and decompress. Designed for nurse ADHD brains — perfection.")],
            "christian_teacher": [("Faith M.", "Prayer prompts for my students make every Monday morning feel purposeful and spirit-filled."),
                                  ("Hope R.", "Scripture lesson planner section ties my faith into my teaching beautifully. Absolutely love it!"),
                                  ("Grace B.", "The 'teaching with purpose' daily page reminds me why God called me to this work. So meaningful.")],
            "sobriety_mom": [("Stephanie L.", "Sober day counter + mom win combo keeps me motivated for both roles every single day."),
                             ("Melissa R.", "Trigger check daily section helped me identify patterns around bedtime chaos. Game-changer."),
                             ("Amanda K.", "Recovery affirmations for moms hit me in the heart. This planner truly understands us. Amazing!")],
        }
        # 니치 감지 (긴 것 먼저 — 더블니치가 단일니치보다 먼저 매칭)
        _niche_reviews = None
        for _nk in sorted(_PLANNER_NICHE_REVIEWS, key=len, reverse=True):
            if _nk in style:
                _niche_reviews = _PLANNER_NICHE_REVIEWS[_nk]
                break

        default_reviews = _niche_reviews or [
            ("Sarah K.", "Downloaded, printed and laminated same day. The daily time-blocks finally keep me on track. 10/10!"),
            ("Jordan M.", "153 pages and every single one is useful. Monthly, weekly, and daily all in one — worth every penny."),
            ("Casey R.", "Works perfectly in GoodNotes on my iPad. The hyperlinks are a game-changer. Best planner I've ever bought."),
        ]
        reviews = _review_data.get(category, default_reviews)

        # ── 니치별 feature headline (리뷰 카드 내 한 줄 강조) ──
        _NICHE_FEATURE_HEADLINES = {
            "ADHD":             "ADHD-Friendly Time-Blocking",
            "anxiety":          "Calm Structure for Anxious Minds",
            "christian":        "Faith-Integrated Daily Pages",
            "sobriety":         "Recovery-Focused Daily Tracking",
            "mom":              "Built for Busy Moms",
            "nurse":            "Shift-Ready Nurse Layout",
            "teacher":          "Lesson Plans + Grade Tracker",
            "pregnancy":        "Week-by-Week Pregnancy Log",
            "entrepreneur":     "CEO Time-Blocks + Revenue Tracker",
            "homeschool":       "Per-Subject Curriculum Tracker",
            "self_care":        "Morning Ritual + Evening Wind-Down",
            "perimenopause":    "Hormone & Hot Flash Log",
            "cycle_syncing":    "Phase-Aligned Task Planning",
            "caregiver":        "Care Schedule + Med Manager",
            "glp1":             "GLP-1 Injection & Protein Tracker",
            "ADHD_teacher":     "ADHD Teacher Time-Block Layout",
            "ADHD_nurse":       "ADHD Nurse Pre-Shift Checklist",
            "christian_teacher":"Scripture + Prayer for Class",
            "sobriety_mom":     "Sober Day Counter + Mom Wins",
        }
        _feature_headline = "Daily + Weekly + Monthly · 153 Pages"
        for _nk in sorted(_NICHE_FEATURE_HEADLINES, key=len, reverse=True):
            if _nk in style:
                _feature_headline = _NICHE_FEATURE_HEADLINES[_nk]
                break

        # ── 카드 레이아웃: feature headline 포함 430px ──
        card_h  = 430
        card_gap = 44
        card_w  = int(W * 0.84)
        card_x  = (W - card_w) // 2
        pad     = 48

        # 3카드+갭+배지(110)+브랜드바(52)를 세로 중앙 정렬
        _group_h    = 3 * card_h + 2 * card_gap + 30 + 110
        card_start_y = max(240, (H - _group_h - 52) // 2)

        _body_f   = _font("Regular", 56)   # 크고 임팩트 있게
        _vfy_f    = _font("Light", 28)
        _head_f   = _font("SemiBold", 34)  # feature headline 폰트

        for i, (name, review) in enumerate(reviews[:3]):
            cy = card_start_y + i * (card_h + card_gap)

            # 카드 그림자
            draw.rounded_rectangle(
                [(card_x + 7, cy + 7), (card_x + card_w + 7, cy + card_h + 7)],
                radius=20, fill=(228, 222, 214))
            # 카드 배경
            draw.rounded_rectangle(
                [(card_x, cy), (card_x + card_w, cy + card_h)],
                radius=20, fill=(255, 255, 255))
            # 좌측 골드 바 (전체 높이)
            draw.rounded_rectangle(
                [(card_x, cy), (card_x + 10, cy + card_h)],
                radius=20, fill=(212, 175, 55))

            # ── 1. 상단: 별점 + 이름 ──
            star_cy = cy + 42
            _sx = card_x + pad + 6
            for _si in range(5):
                _draw_star(draw, _sx + _si * 30, star_cy, 13, 6, (212, 175, 55))
            draw.text((card_x + card_w - pad, star_cy), f"— {name}",
                      fill=(150, 140, 130), font=sm_f, anchor="rm")

            # 구분선
            div1_y = cy + 76
            draw.line([(card_x + pad, div1_y), (card_x + card_w - pad, div1_y)],
                      fill=(242, 238, 230), width=2)

            # ── 1-b. Feature headline (니치 있을 때만) ──
            if _feature_headline:
                draw.text((card_x + pad + 16, div1_y + 26), f"✦ {_feature_headline}",
                          fill=(180, 120, 60), font=_head_f, anchor="lm")
                _text_top_offset = div1_y + 60
            else:
                _text_top_offset = div1_y + 16

            # ── 2. 중간: 리뷰 텍스트 (수직 중앙 정렬) ──
            max_w = card_w - pad * 2 - 16
            words = review.split()
            lines_text = []
            cur = ""
            for word in words:
                test = (cur + " " + word).strip()
                bbox = draw.textbbox((0, 0), test, font=_body_f)
                if bbox[2] - bbox[0] <= max_w:
                    cur = test
                else:
                    if cur:
                        lines_text.append(cur)
                    cur = word
            if cur:
                lines_text.append(cur)
            lines_text = lines_text[:2]   # 최대 2줄

            line_h = 76   # 56px 폰트 줄 간격
            # 텍스트 영역: _text_top_offset ~ (cy+card_h-66-16)
            _ta_top = _text_top_offset
            _ta_bot = cy + card_h - 66 - 16
            _tb_h   = (len(lines_text) - 1) * line_h
            _ty     = (_ta_top + _ta_bot) // 2 - _tb_h // 2

            for li, line_txt in enumerate(lines_text):
                draw.text((card_x + pad + 16, _ty + li * line_h), line_txt,
                          fill=(35, 35, 35), font=_body_f, anchor="lm")

            # ── 3. 하단: Verified 배지 ──
            div2_y = cy + card_h - 66
            draw.line([(card_x + pad, div2_y), (card_x + card_w - pad, div2_y)],
                      fill=(242, 238, 230), width=1)
            # 초록 체크 원
            draw.ellipse([(card_x + pad + 16, div2_y + 10),
                          (card_x + pad + 46, div2_y + 40)], fill=(39, 174, 96))
            draw.text((card_x + pad + 16 + 22, div2_y + 25), "✓",
                      fill=(255, 255, 255), font=_font("Bold", 20), anchor="mm")
            draw.text((card_x + pad + 56, div2_y + 25), "Verified Etsy Buyer",
                      fill=(160, 150, 140), font=_vfy_f, anchor="lm")

        # 하단 요약 배지
        badge_y = card_start_y + 3 * (card_h + card_gap) + 30
        draw.rounded_rectangle(
            [(cx - 320, badge_y), (cx + 320, badge_y + 90)],
            radius=24, fill=(22, 160, 133))
        # 하단 배지 별점 (폴리곤)
        _badge_sx = cx - 170
        for _si in range(5):
            _draw_star(draw, _badge_sx + _si * 26, badge_y + 30, 10, 4, (255, 215, 100))
        draw.text((cx + 20, badge_y + 30), "  Average Rating", fill=(255, 255, 255),
                  font=name_f, anchor="lm")
        draw.text((cx, badge_y + 65), "Instant Download  ·  100% Satisfaction Guaranteed",
                  fill=(210, 245, 240), font=sm_f, anchor="mm")

        # 브랜드 하단 바
        bar_h = 52
        draw.rectangle(
            [(0, H - bar_h), (W, H)], fill=(15, 23, 42))
        brand_f = _font("Light", 22)
        draw.text((W // 2, H - bar_h // 2),
                  "DailyPrintHaus  ·  Instant Digital Download",
                  fill=(180, 190, 200), font=brand_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Social proof mockup saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Social proof mockup failed: %s", e)
        return False


def _generate_planner_spread_mockup(converted_paths: list[str], output_path: str,
                                    style: str = "") -> bool:
    """Image 06 (PLANNER 전용): 두 페이지 나란히 — 콘텐츠 다양성 직접 증명.
    Why: 팬/스택 목업에서 페이지가 너무 작아 내용 불인식 → 2단 레이아웃으로 가독성 확보.
    Layout: 크림 배경, 중앙 두 페이지 카드, 상단 헤더, 하단 라벨.
    """
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT
        # 배경: 따뜻한 크림 그래디언트
        canvas = Image.new("RGB", (W, H), (252, 249, 244))
        draw = ImageDraw.Draw(canvas)
        for y in range(H):
            r = int(252 - (y / H) * 14)
            g = int(249 - (y / H) * 11)
            b = int(244 - (y / H) * 9)
            draw.line([(0, y), (W, y)], fill=(r, g, b))

        # 상단 장식 바
        draw.rectangle([(0, 0), (W, 6)], fill=(35, 35, 35))

        hdr_f  = _font("Bold", 54)
        sub_f  = _font("Regular", 28)
        lbl_f  = _font("SemiBold", 30)
        sm_f   = _font("Light", 22)

        # 헤더
        draw.text((W // 2, 72), "WHAT'S INSIDE",
                  fill=(35, 35, 35), font=hdr_f, anchor="mm")
        draw.text((W // 2, 118), "Every page, exactly as you'll use it",
                  fill=(150, 140, 130), font=sub_f, anchor="mm")
        draw.line([(W//2 - 220, 144), (W//2 + 220, 144)], fill=(215, 210, 205), width=2)

        # ── 페이지 선택: index 0(Monthly) + index -1(Daily log) ──
        _p_left  = converted_paths[0]
        _p_right = converted_paths[-1]  # Daily log is most content-rich

        # 페이지 카드 크기: 두 개 나란히, 좌우 여백 36px, 중간 갭 36px (상위 1% — 카드가 캔버스 지배)
        pad   = 36
        gap   = 36
        card_w = (W - pad * 2 - gap) // 2   # ~946px
        card_h = int(card_w * 1.38)           # ~1305px (A4 비율)
        start_y = 160
        # 캔버스 높이 맞춤: 라벨+브랜드 160px 여유
        if start_y + card_h > H - 160:
            card_h = H - 160 - start_y
            card_w = int(card_h / 1.38)
            pad    = (W - card_w * 2 - gap) // 2

        for i, page_path in enumerate([_p_left, _p_right]):
            cx = pad + i * (card_w + gap)
            cy = start_y

            # 그림자
            draw.rounded_rectangle(
                [(cx + 8, cy + 8), (cx + card_w + 8, cy + card_h + 8)],
                radius=14, fill=(205, 200, 195))
            # 흰 카드
            draw.rounded_rectangle(
                [(cx, cy), (cx + card_w, cy + card_h)],
                radius=14, fill=(255, 255, 255))

            # 실제 페이지 이미지
            try:
                _pg = Image.open(page_path).convert("RGB")
                _pg = _pg.resize((card_w - 4, card_h - 4), Image.LANCZOS)
                canvas.paste(_pg, (cx + 2, cy + 2))
                # 카드 테두리 (둥근 사각)
                draw.rounded_rectangle(
                    [(cx, cy), (cx + card_w, cy + card_h)],
                    radius=14, fill=None, outline=(220, 215, 210), width=3)
            except Exception:
                pass

            # 하단 라벨 배지
            _labels = ["Monthly Overview", "Daily Log"]
            _lbl = _labels[i]
            _lw = len(_lbl) * 22 + 56
            _lx = cx + card_w // 2
            _ly = cy + card_h + 16
            draw.rounded_rectangle(
                [(_lx - _lw//2, _ly), (_lx + _lw//2, _ly + 62)],
                radius=31, fill=(35, 35, 35))
            draw.text((_lx, _ly + 31), _lbl,
                      fill=(255, 255, 255), font=lbl_f, anchor="mm")

        # 하단 브랜드
        brand_y = start_y + card_h + 100
        draw.text((W // 2, brand_y + 26), "DailyPrintHaus  ·  Instant Download  ·  No App Needed",
                  fill=(170, 160, 150), font=sm_f, anchor="mm")

        _save_mockup(canvas, output_path)
        logger.info("Planner spread mockup saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Planner spread mockup failed: %s", e)
        return False


def _generate_brand_cta(output_path: str, category: Category,
                        art_path: str | None = None,
                        style: str = "") -> bool:
    """Image 10: Brand CTA -- 구매 결정의 마지막 푸시."""
    try:
        W, H = MOCKUP_WIDTH, MOCKUP_HEIGHT

        # 그래디언트 배경 (따뜻한 크림)
        canvas = Image.new("RGB", (W, H), (252, 250, 247))
        draw = ImageDraw.Draw(canvas)
        for y in range(H):
            ratio = y / H
            r = int(252 - ratio * 12)
            g = int(250 - ratio * 10)
            b = int(247 - ratio * 8)
            draw.line([(0, y), (W, y)], fill=(r, g, b))

        # 좌측 제품 썸네일 (상위 1% 필수 — 구매자가 마지막으로 제품 확인)
        _has_thumb = False
        if art_path:
            try:
                _art = Image.open(art_path).convert("RGB")
                _thumb_w = int(W * 0.38)
                _thumb_h = int(_thumb_w * 1.35)
                _art_rs = _art.resize((_thumb_w, _thumb_h), Image.LANCZOS)
                # 좌측 상단 배치 (세로 중앙)
                _tx = 60
                _ty = (H - _thumb_h) // 2
                # 그림자
                for _si in range(12, 0, -3):
                    _alpha = int(18 * _si / 12)
                    draw.rounded_rectangle(
                        [(_tx + _si, _ty + _si), (_tx + _thumb_w + _si, _ty + _thumb_h + _si)],
                        radius=14, fill=(180, 175, 168))
                # 흰 종이 배경
                draw.rounded_rectangle([(_tx, _ty), (_tx + _thumb_w, _ty + _thumb_h)],
                                       radius=12, fill=(255, 255, 255))
                canvas.paste(_art_rs, (_tx, _ty))
                _has_thumb = True
            except Exception:
                pass

        # 콘텐츠 영역: 썸네일 있으면 우측 60%, 없으면 전체 중앙
        cx = int(W * 0.64) if _has_thumb else W // 2
        _feat_max_w = int(W * 0.52) if _has_thumb else 860

        big_f   = _font("Bold", 58)
        title_f = _font("SemiBold", 44)
        feat_f  = _font("SemiBold", 34)
        sub_f   = _font("Regular", 26)
        sm_f    = _font("Light", 20)

        # ── 상단 장식 선 (차콜 모노크롬) ──
        for i, color in enumerate([(35,35,35),(80,80,80),(140,140,140)]):
            draw.rounded_rectangle(
                [(cx - 90 + i*30, 80), (cx - 60 + i*30, 88)],
                radius=4, fill=color
            )

        # ── 메인 타이틀 ──
        draw.text((cx, 120), "INSTANT DOWNLOAD", fill=(35, 35, 35),
                  font=big_f, anchor="mm")
        draw.text((cx, 185), "Everything you need, right now.", fill=(140, 135, 130),
                  font=sub_f, anchor="mm")

        # ── 구분선 ──
        draw.line([(cx-200, 220), (cx+200, 220)], fill=(225, 220, 215), width=2)

        # ── 카테고리별 특징 (체크마크 대신 컬러 원 + 텍스트) ──
        props = {
            Category.WORKSHEET: [
                ("Print at Home or School",     (244, 167, 185)),
                ("Laminate & Reuse Forever",    (167, 216, 220)),
                ("Answer Key Included",         (180, 210, 180)),
                ("3 Difficulty Levels",         (220, 200, 240)),
                ("Instant PDF Download",        (244, 220, 167)),
            ],
            Category.PLANNER: [
                ("Print or Use on iPad/Tablet", (244, 167, 185)),
                ("Undated \u2014 Use Any Year",      (167, 216, 220)),
                ("Hyperlink Navigation",        (180, 210, 180)),
                ("A4 & US Letter Included",     (220, 200, 240)),
                ("Instant PDF Download",        (244, 220, 167)),
            ],
            Category.SPREADSHEET: [
                ("Google Sheets + Excel",       (244, 167, 185)),
                ("Auto-Calculating Formulas",   (167, 216, 220)),
                ("Dashboard with Charts",       (180, 210, 180)),
                ("Protected Formula Cells",     (220, 200, 240)),
                ("Step-by-Step How To Use",     (244, 220, 167)),
            ],
            Category.SOCIAL_MEDIA_TEMPLATE: [
                ("70 Templates in One Bundle",  (244, 167, 185)),
                ("IG · TikTok · Pinterest · FB",(167, 216, 220)),
                ("High-Res PNG Instant Download",(180, 210, 180)),
                ("Brand Kit + Content Calendar",(220, 200, 240)),
                ("Commercial Use License",      (244, 220, 167)),
            ],
            Category.RESUME_TEMPLATE: [
                ("ATS-Friendly DOCX + PDF",     (244, 167, 185)),
                ("3 Industry Versions Included",(167, 216, 220)),
                ("Editable in Word & Google Docs",(180, 210, 180)),
                ("Cover Letter + References",   (220, 200, 240)),
                ("Instant ZIP Download",        (244, 220, 167)),
            ],
            Category.WALL_ART: [
                ("Gallery Set of 9 Designs",    (35, 35, 35)),
                ("45 JPG Files · 5 Print Sizes",(60, 60, 60)),
                ("300 DPI Print-Ready Quality", (35, 35, 35)),
                ("Print at Home or Any Shop",   (60, 60, 60)),
                ("Instant ZIP Download",        (35, 35, 35)),
            ],
        }

        features = props.get(category, props[Category.WORKSHEET])

        # ── PLANNER 니치별 feature 오버라이드 (generic → 니치 특화로 전환율 상승) ──
        _NICHE_CTA_FEATURES = {
            "ADHD":             [("ADHD Time-Block Layout Daily",   (244, 167, 185)),
                                 ("Brain Dump Space Every Page",     (167, 216, 220)),
                                 ("Dopamine Reward Tracker",         (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "anxiety":          [("Gentle Calm Daily Structure",     (244, 167, 185)),
                                 ("Worry Release Prompts Built-In",  (167, 216, 220)),
                                 ("No Overwhelm Layout",             (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "christian":        [("Scripture Space Every Page",      (244, 167, 185)),
                                 ("Prayer Log Daily & Weekly",       (167, 216, 220)),
                                 ("Faith-Based Gratitude Journal",   (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "sobriety":         [("Sober Day Counter Daily",         (244, 167, 185)),
                                 ("Trigger Log & Coping Plan",       (167, 216, 220)),
                                 ("Recovery Wins Tracker",           (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "sobriety_mom":     [("Sober Day + Mom Wins Daily",      (244, 167, 185)),
                                 ("Family & Recovery Balance",        (167, 216, 220)),
                                 ("Trigger Check Built-In",          (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "mom":              [("Family Schedule All-in-One",      (244, 167, 185)),
                                 ("Me-Time Blocks Protected",        (167, 216, 220)),
                                 ("Meal Planning Included",          (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "nurse":            [("Shift Schedule AM/PM/NOC",        (244, 167, 185)),
                                 ("Medication Log Tracker",          (167, 216, 220)),
                                 ("Post-Shift Self-Care Pages",      (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "teacher":          [("Lesson Plan Templates Built-In",  (244, 167, 185)),
                                 ("Grade Tracker Included",          (167, 216, 220)),
                                 ("Parent Communication Log",        (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "pregnancy":        [("Week-by-Week Tracker",            (244, 167, 185)),
                                 ("Symptom Log Daily",               (167, 216, 220)),
                                 ("Baby Prep Checklist",             (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "entrepreneur":     [("CEO Time Blocks Daily",           (244, 167, 185)),
                                 ("Revenue Goal Tracker",            (167, 216, 220)),
                                 ("Client Notes Space",              (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "homeschool":       [("Per-Subject Lesson Log",          (244, 167, 185)),
                                 ("Curriculum Tracker",              (167, 216, 220)),
                                 ("Learning Wins Daily",             (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "self_care":        [("Morning Ritual Tracker",          (244, 167, 185)),
                                 ("Evening Wind-Down Pages",         (167, 216, 220)),
                                 ("Glow-Up Habit Stack",             (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "perimenopause":    [("Hormone & Hot Flash Log",         (244, 167, 185)),
                                 ("Symptom Pattern Tracker",         (167, 216, 220)),
                                 ("Self-Compassion Prompts",         (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "cycle_syncing":    [("Cycle Phase Tracker",             (244, 167, 185)),
                                 ("Phase-Aligned Task Planning",     (167, 216, 220)),
                                 ("Body Wisdom Journal",             (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "caregiver":        [("Care Schedule Organizer",         (244, 167, 185)),
                                 ("Medication Manager Built-In",     (167, 216, 220)),
                                 ("Respite Self-Care Reminders",     (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "glp1":             [("Injection Day Tracker",           (244, 167, 185)),
                                 ("Protein Goal Daily Log",          (167, 216, 220)),
                                 ("Non-Scale Wins Tracker",          (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "ADHD_teacher":     [("ADHD Time-Block Lessons",         (244, 167, 185)),
                                 ("Transition Cue Reminders",        (167, 216, 220)),
                                 ("Brain Dump + Class Notes",        (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "ADHD_nurse":       [("Pre-Shift Visual Checklist",      (244, 167, 185)),
                                 ("Time-Boxing Patient Tasks",       (167, 216, 220)),
                                 ("Post-Shift Debrief Pages",        (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
            "christian_teacher":[("Prayer Over Students Daily",      (244, 167, 185)),
                                 ("Scripture Lesson Planner",        (167, 216, 220)),
                                 ("Faith-First Classroom Pages",     (180, 210, 180)),
                                 ("A4 & US Letter Included",         (220, 200, 240)),
                                 ("Instant PDF Download",            (244, 220, 167))],
        }
        _niche_cta_applied = False
        if category == Category.PLANNER and style:
            for _nk in sorted(_NICHE_CTA_FEATURES, key=len, reverse=True):
                if style.endswith("_" + _nk) or style == _nk:
                    features = _NICHE_CTA_FEATURES[_nk]
                    _niche_cta_applied = True
                    break

        # ── 피처 리스트 -- 행 높이를 동적으로 늘려 캔버스 꽉 채움 ──
        # Why: 상위 1% 셀러는 피처 행이 크고 굵어서 "충실한" 인상을 줌.
        #      헤더(260) + 피처 블록 + 하단(420) = 2000px 꽉 채워야 함.
        _header_h   = 260
        _bottom_h   = 420         # star80 + gap40 + brand120 + cta90 + deco90
        _avail      = MOCKUP_HEIGHT - _header_h - _bottom_h   # = 1320px
        _n          = max(len(features), 1)
        feat_gap    = _avail // _n                             # 264px per item
        row_h       = int(feat_gap * 0.76)                    # 200px — 행 높이
        feat_start_y = _header_h                              # 헤더 바로 아래 시작
        dot_r       = 28                                       # 아이콘도 더 크게

        # 카테고리별 피처 설명 1줄 (Why: 제목만 있으면 빈 느낌 → 설명 추가로 행 꽉 채움)
        # 카테고리별 피처 설명 — features 리스트 순서와 1:1 매핑
        _feat_descs = {
            Category.WORKSHEET: [
                "Works with any home or school printer",   # Print at Home or School
                "Laminate once, reuse again and again",    # Laminate & Reuse Forever
                "Every page includes full answer key",     # Answer Key Included
                "Easy, Medium & Hard all in one pack",     # 3 Difficulty Levels
                "Ready to print in under 1 minute",        # Instant PDF Download
            ],
            Category.PLANNER: [
                "Use on paper or GoodNotes / Notability",  # Print or Use on iPad/Tablet
                "No year printed — works for any year",    # Undated -- Use Any Year
                "Jump to any section with one tap",        # Hyperlink Navigation
                "Both A4 & US Letter ready to print",      # A4 & US Letter Included
                "Download link arrives instantly",         # Instant PDF Download
            ],
            Category.SPREADSHEET: [
                "Upload .xlsx to Drive, edit instantly",   # Google Sheets + Excel
                "Every total updates as you type",         # Auto-Calculating Formulas
                "Pie & bar charts update live",            # Dashboard with Charts
                "No accidental formula deletion",          # Protected Formula Cells
                "Step-by-step guide included",             # Step-by-Step How To Use
            ],
            Category.WALL_ART: [
                "Crystal clear at any print size",         # High Resolution 300 DPI
                "5:4, 4:3, 2:3 and square included",       # Multiple Sizes Included
                "Print at Staples, FedEx, or at home",     # Print at Home or Shop
                "Download link in your Etsy inbox",        # Instant Access After Buy
                "Personal and commercial use ok",          # Commercial Use Friendly
            ],
            Category.SOCIAL_MEDIA_TEMPLATE: [
                "30 IG + 10 Pinterest + 15 Stories + 10 TikTok + 5 FB",  # 70 Templates
                "All sized to exact platform specs",        # IG · TikTok · Pinterest · FB
                "High-res PNGs ready to post instantly",    # High-Res PNG Instant Download
                "Color palette, fonts + 30-day calendar",   # Brand Kit + Content Calendar
                "Use for client or personal projects",      # Commercial Use License
            ],
            Category.RESUME_TEMPLATE: [
                "Editable DOCX opens in Word or Google Docs",  # ATS-Friendly DOCX + PDF
                "Tech, Business & Healthcare versions",     # 3 Industry Versions Included
                "Clean, modern layout — hireable at a glance", # Editable in Word & Google Docs
                "Professional cover letter ready to send",  # Cover Letter + References
                "All files in one ZIP — download instantly",# Instant ZIP Download
            ],
        }
        # 니치 오버라이드 시 generic 설명 대신 feature 제목에서 의미 자명 → 설명 생략
        descs = [""] * len(features) if _niche_cta_applied else _feat_descs.get(category, [""] * len(features))

        for i, (text, color) in enumerate(features):
            fy = feat_start_y + i * feat_gap
            # 행 배경 — 전체 feat_gap 높이로 꽉 채움 (4px 상하 여백)
            bg_fill = (245, 243, 240) if i % 2 == 0 else (250, 249, 247)
            draw.rounded_rectangle(
                [(cx - 430, fy + 4), (cx + 430, fy + feat_gap - 4)],
                radius=18, fill=bg_fill)

            # 좌측 컬러 바 (강조)
            draw.rounded_rectangle(
                [(cx - 430, fy + 4), (cx - 418, fy + feat_gap - 4)],
                radius=18, fill=color)

            # 컬러 원 + 체크마크 — 행 수직 중앙
            icon_cy = fy + feat_gap // 2
            ex = cx - 400
            draw.ellipse([(ex, icon_cy - dot_r), (ex + dot_r*2, icon_cy + dot_r)], fill=color)
            ick_cx, ick_cy = ex + dot_r, icon_cy
            draw.line([(ick_cx-10, ick_cy), (ick_cx-3, ick_cy+8)], fill=(255,255,255), width=4)
            draw.line([(ick_cx-3, ick_cy+8), (ick_cx+10, ick_cy-8)], fill=(255,255,255), width=4)

            # 제목 텍스트 (아이콘 위 기준)
            text_x = cx - 350
            title_y = icon_cy - 20 if descs[i] else icon_cy
            draw.text((text_x, title_y), text, fill=(35, 35, 35), font=feat_f, anchor="lm")

            # 설명 텍스트
            if i < len(descs) and descs[i]:
                draw.text((text_x, icon_cy + 22), descs[i],
                          fill=(130, 125, 120), font=sm_f, anchor="lm")

        # ── 별점 바 (신뢰 시그널) — 폴리곤 별 + 텍스트 ──
        star_y = feat_start_y + len(features) * feat_gap + 30
        draw.rounded_rectangle([(cx - 320, star_y), (cx + 320, star_y + 80)],
                               radius=20, fill=(255, 249, 235))
        # 폴리곤 별 5개 (유니코드 ★ 렌더링 불안정 방지)
        import math as _math
        def _draw_star_cta(drw, cx_s, cy_s, r_o, r_i, col):
            pts = []
            for ii in range(10):
                ang = _math.radians(ii * 36 - 90)
                rr = r_o if ii % 2 == 0 else r_i
                pts.append((cx_s + rr * _math.cos(ang), cy_s + rr * _math.sin(ang)))
            drw.polygon(pts, fill=col)
        _cta_star_sx = cx - 160
        for _si in range(5):
            _draw_star_cta(draw, _cta_star_sx + _si * 26, star_y + 25, 10, 4, (180, 130, 40))
        draw.text((_cta_star_sx + 5 * 26 + 10, star_y + 25), " Top-Rated Design Quality",
                  fill=(180, 130, 40), font=sm_f, anchor="lm")
        draw.text((cx, star_y + 58), "Instant download . No physical item shipped",
                  fill=(140, 135, 130), font=sm_f, anchor="mm")

        # ── 브랜드 블록 ──
        brand_y = star_y + 120
        draw.line([(cx - 280, brand_y), (cx + 280, brand_y)], fill=(215, 210, 205), width=2)
        draw.text((cx, brand_y + 55), "DailyPrintHaus", fill=(55, 55, 55),
                  font=title_f, anchor="mm")
        draw.text((cx, brand_y + 100), "dailyprinthaus.etsy.com", fill=(160, 155, 150),
                  font=sm_f, anchor="mm")

        # ── CTA 버튼 ──
        btn_y = brand_y + 140
        draw.rounded_rectangle([(cx - 270, btn_y), (cx + 270, btn_y + 90)],
                               radius=45, fill=(35, 35, 35))
        draw.text((cx, btn_y + 45), "SHOP THIS ITEM", fill=(255, 255, 255),
                  font=sub_f, anchor="mm")

        # ── 하단 장식 (차콜 모노크롬) ──
        deco_y = btn_y + 130
        for i, color in enumerate([(35,35,35),(80,80,80),(140,140,140),(80,80,80),(35,35,35)]):
            bx = cx - 160 + i * 80
            draw.rounded_rectangle([(bx - 24, deco_y), (bx + 24, deco_y + 8)],
                                   radius=4, fill=color)

        _save_mockup(canvas, output_path)
        logger.info("Brand CTA saved: %s", output_path)
        return True
    except Exception as e:
        logger.error("Brand CTA failed: %s", e)
        return False


def generate_all_mockups(product: Product) -> list[str]:
    """
    Generate complete set of 10 mockup images for a product.
    Why: Different categories need different mockup strategies.
    - Wall art -> framed on walls
    - Worksheet/Planner -> tablet + flat lay
    - Spreadsheet -> laptop screen
    """
    if not product.file_paths:
        logger.error("No art files for product %s", product.product_id)
        return []

    mockup_dir = Path(product.file_paths[0]).parent / "mockups"
    mockup_dir.mkdir(parents=True, exist_ok=True)

    # ── ZIP 우회: wall_art 등은 file_paths=[zip_path] 이므로
    #    ZIP과 같은 디렉토리에 있는 실제 JPG 파일을 찾아서 사용.
    first_fp = product.file_paths[0]
    _art_candidates: list[str] = []
    if Path(first_fp).suffix.lower() == ".zip":
        product_dir = Path(first_fp).parent
        # 5x7 파일 우선 (가장 원본에 가깝고 화질 좋음) → 아트 번호별로 그룹핑
        _5x7_files = sorted([str(p) for p in product_dir.glob("art*_5x7.jpg")])
        _all_art = sorted(
            [str(p) for p in product_dir.glob("art*.jpg")]
            + [str(p) for p in product_dir.glob("art*.png")]
        )
        # 각 디자인의 5x7 대표본을 candidates로 사용 (목업용 베스트 화질)
        _art_candidates = _5x7_files if _5x7_files else _all_art
        if not _art_candidates:
            _art_candidates = sorted(
                [str(p) for p in product_dir.glob("*.jpg") if p.name not in ("SIZE_GUIDE.png",)]
                + [str(p) for p in product_dir.glob("*.png") if p.name not in ("SIZE_GUIDE.png",)]
            )
        if _art_candidates:
            logger.info("ZIP 감지 → 아트 %d개 사용 (5x7 우선): %s ...",
                        len(_art_candidates), Path(_art_candidates[0]).name)
            art_path = _art_candidates[0]
        else:
            art_path = _get_image_from_file(first_fp)
    else:
        art_path = _get_image_from_file(first_fp)

    if not art_path:
        logger.error("Cannot create preview image for %s", first_fp)
        return []

    mockup_paths = []
    is_digital = product.category in (
        Category.WORKSHEET, Category.PLANNER, Category.SPREADSHEET,
        Category.SOCIAL_MEDIA_TEMPLATE, Category.RESUME_TEMPLATE,
    )

    # Convert all files to images for What's Included
    # ZIP이면 _art_candidates 사용, 아니면 file_paths 변환
    converted_paths = list(_art_candidates) if _art_candidates else []
    if not converted_paths:
        for fp in product.file_paths:
            cp = _get_image_from_file(fp)
            if cp:
                converted_paths.append(cp)

    # ── 플래너/워크시트: HTML preview.html에서 다양한 페이지 스크린샷 ──
    # Why: 상위 1% 셀러는 daily log, habit tracker, meal plan 등 실제 내용 페이지를 보여줌
    #      같은 페이지 반복은 구매 결정력 약화 → 다양한 페이지 노출이 전환율 직결
    if product.category in (Category.PLANNER, Category.WORKSHEET) and not _art_candidates:
        _html_preview = Path(first_fp).parent / "preview.html"
        if _html_preview.exists():
            # 페이지 구조: 0=Cover, 1=TOC, 2-3=YearAtGlance, 4=VisionBoard,
            #              5-16=Monthly, 17-28=MonthlyReview, 29-80=Weekly, 81+=Daily
            # Why: 상위 1% 셀러는 Monthly/Weekly/Daily 실제 콘텐츠 페이지 노출 → 구매 결정력 상승
            _multi_paths: list[str] = []
            _page_targets = [5, 20, 30, 100]  # Monthly / MonthlyReview / Weekly / Daily
            for _pi in _page_targets:
                _pg_path = str(Path(first_fp).parent / f"preview_page{_pi}.png")
                try:
                    _r = _screenshot_html(str(_html_preview), _pg_path, page_index=_pi)
                    if _r and Path(_r).exists():
                        _multi_paths.append(_r)
                except Exception as _e:
                    logger.debug("Multi-page screenshot page %d failed: %s", _pi, _e)
            if len(_multi_paths) >= 2:
                logger.info("멀티페이지 스크린샷 %d장 확보 (Monthly/Weekly/Daily)", len(_multi_paths))
                converted_paths = _multi_paths
                # 히어로/detail용 메인: Daily Log 페이지 (가장 콘텐츠 풍부)
                art_path = _multi_paths[-1]  # page 100 = Daily log

    # Why: 상위 셀러는 히어로 목업에 핵심 가치를 텍스트로 보여줌
    # 키워드에서 페이지수 추출, 없으면 카테고리 기본값
    page_count = ""
    for kw in product.keywords:
        if "page" in kw.lower():
            import re as _re
            m = _re.search(r'(\d+)\s*page', kw.lower())
            if m:
                page_count = f"{m.group(1)} Pages"
                break

    _default_pages = {
        Category.WORKSHEET:             "152 Pages",
        Category.PLANNER:               "149 Pages",
        Category.SPREADSHEET:           "Auto-Calculate",
        Category.WALL_ART:              "9 Designs Included",
        Category.SOCIAL_MEDIA_TEMPLATE: "70 Templates",
        Category.RESUME_TEMPLATE:       "3 Industry Versions",
    }
    _NICHE_HERO_TEXT = {
        "ADHD":             "ADHD-Friendly Layout",
        "anxiety":          "Calm Daily Structure",
        "christian":        "Faith-Based Planner",
        "sobriety":         "Recovery Daily Planner",
        "mom":              "Mom Life Planner",
        "nurse":            "Nurse Shift Planner",
        "teacher":          "Teacher Planner",
        "pregnancy":        "Pregnancy Week Tracker",
        "entrepreneur":     "CEO Daily Planner",
        "homeschool":       "Homeschool Organizer",
        "self_care":        "Self-Care Daily Ritual",
        "perimenopause":    "Hormone Wellness Log",
        "cycle_syncing":    "Cycle-Syncing Planner",
        "caregiver":        "Caregiver Organizer",
        "glp1":             "GLP-1 Wellness Tracker",
        "ADHD_teacher":     "ADHD Teacher Planner",
        "ADHD_nurse":       "ADHD Nurse Planner",
        "christian_teacher":"Faith-Based Teacher",
        "sobriety_mom":     "Sober Mom Planner",
    }
    overlay_map = {
        Category.WORKSHEET:             (page_count or _default_pages[Category.WORKSHEET],  "Answer Key Included"),
        Category.PLANNER:               (page_count or _default_pages[Category.PLANNER],    "Undated \u2014 Any Year"),
        Category.SPREADSHEET:           ("Auto-Calculate", "Google Sheets + Excel"),
        Category.WALL_ART:              ("9 Designs Included", "45 JPG Files · 5 Sizes"),
        Category.SOCIAL_MEDIA_TEMPLATE: ("70 Templates Included", "IG · TikTok · Pinterest"),
        Category.RESUME_TEMPLATE:       ("ATS-Friendly DOCX+PDF", "3 Industry Versions"),
    }
    overlay, sub_overlay = overlay_map.get(product.category, ("", ""))
    if product.category == Category.PLANNER:
        _style = product.style or ""
        for _nk in sorted(_NICHE_HERO_TEXT, key=len, reverse=True):
            if _style.endswith("_" + _nk) or _style == _nk:
                sub_overlay = _NICHE_HERO_TEXT[_nk]
                break

    # ── Slot 1: Hero mockup ──
    path = str(mockup_dir / "01_hero.jpg")
    if product.category == Category.WALL_ART:
        # 갤러리 히어로: 3개 디자인 나란히 — 세트 가치 한 눈에 전달
        generate_wall_art_gallery_hero(
            _art_candidates[:3] or [art_path], path, overlay, sub_overlay)
    elif product.category == Category.WORKSHEET:
        generate_device_mockup(art_path, path, "tablet", overlay, sub_overlay)
    elif product.category == Category.SPREADSHEET:
        generate_device_mockup(art_path, path, "laptop", overlay, sub_overlay)
    elif product.category == Category.SOCIAL_MEDIA_TEMPLATE:
        generate_device_mockup(art_path, path, "phone", overlay, sub_overlay)
    elif product.category == Category.RESUME_TEMPLATE:
        generate_device_mockup(art_path, path, "laptop", overlay, sub_overlay)
    elif is_digital:
        generate_device_mockup(art_path, path, "tablet", overlay, sub_overlay)
    else:
        generate_hero_mockup(art_path, path, overlay, sub_overlay)
    mockup_paths.append(path)

    # ── Slot 2: Wall-hanging room scene (wall_art) / Lifestyle mockup ──
    path = str(mockup_dir / "02_lifestyle_living.jpg")
    if product.category == Category.WALL_ART:
        # 벽에 걸린 씬 — 상위 1% 필수: 구매자가 실제 인테리어 상상 가능
        _generate_wall_art_hanging_mockup(_art_candidates[:3] or [art_path], path, n_frames=3)
    elif product.category in (Category.SPREADSHEET, Category.RESUME_TEMPLATE):
        generate_device_mockup(art_path, path, "laptop", overlay, sub_overlay)
    elif product.category == Category.SOCIAL_MEDIA_TEMPLATE:
        generate_device_mockup(art_path, path, "phone", overlay, sub_overlay)
    elif is_digital:
        generate_flatlay_mockup(art_path, path, "wood", overlay, sub_overlay)
    else:
        generate_lifestyle_mockup(art_path, path, "warm_beige")
    mockup_paths.append(path)

    # ── Slot 3: Detail close-up ──
    path = str(mockup_dir / "03_detail.jpg")
    generate_detail_mockup(art_path, path, category=product.category, style=product.style)
    mockup_paths.append(path)

    # ── Slot 4: Second wall-hanging (2 frames) / Second device ──
    path = str(mockup_dir / "04_lifestyle_bedroom.jpg")
    if product.category == Category.WALL_ART:
        # 2프레임 벽 씬 — art4+art5로 다른 디자인 조합 보여줌
        _candidates_4 = _art_candidates[3:5] or _art_candidates[:2] or [art_path]
        _generate_wall_art_hanging_mockup(_candidates_4, path, n_frames=2)
    elif is_digital:
        if product.category == Category.SPREADSHEET:
            generate_device_mockup(art_path, path, "laptop")
        elif product.category == Category.PLANNER:
            # Monthly Review(index 1) — Weekly(index 2)는 빈 박스라 지루함, Review는 Goals/LifeBalance 내용 풍부
            _slot4_art = converted_paths[1] if len(converted_paths) > 1 else art_path
            generate_device_mockup(_slot4_art, path, "tablet", overlay, sub_overlay)
        else:
            generate_flatlay_mockup(art_path, path, "marble", overlay, sub_overlay)
    else:
        generate_lifestyle_mockup(art_path, path, "soft_cream")
    mockup_paths.append(path)

    # ── Slot 5: What's Included ──
    path = str(mockup_dir / "05_whats_included.jpg")
    if product.category == Category.WALL_ART:
        # 전용 썸네일 그리드 — 최대 9개 디자인 실제 모습으로 표시 (3×3 그리드)
        _generate_wall_art_whats_included(_art_candidates[:9] or [art_path], path)
    else:
        # 플래너 니치별 섹션명 (페르소나 특화 기능 명시 → 구매 결정력 상승)
        _PLANNER_NICHE_LABELS: dict[str, list[str]] = {
            "sobriety_mom":      ["Sober Day Counter", "Family Planner"],
            "sobriety":          ["Sobriety Tracker", "Recovery Journal"],
            "ADHD":              ["Focus Sessions", "Habit Tracker"],
            "anxiety":           ["Calm Routine", "Daily Check-in"],
            "christian":         ["Prayer Journal", "Scripture Planner"],
            "mom":               ["Mom Life Daily", "Family Planner"],
            "nurse":             ["Shift Planner", "Self-Care Log"],
            "teacher":           ["Lesson Planner", "Classroom Goals"],
            "pregnancy":         ["Week Tracker", "Baby Journal"],
            "entrepreneur":      ["CEO Daily Plan", "Business Goals"],
            "homeschool":        ["Lesson Schedule", "Curriculum Map"],
            "self_care":         ["Wellness Ritual", "Self-Care Log"],
            "perimenopause":     ["Hormone Log", "Wellness Tracker"],
            "cycle_syncing":     ["Cycle Phases", "Monthly Energy Log"],
            "caregiver":         ["Care Schedule", "Appointment Log"],
            "glp1":              ["Progress Tracker", "Nutrition Log"],
            "ADHD_teacher":      ["Classroom Focus", "Lesson Planner"],
            "ADHD_nurse":        ["Shift Focus", "Self-Care Log"],
            "christian_teacher": ["Faith Lessons", "Prayer Planner"],
        }
        _planner_style = product.style or ""
        _planner_labels = ["Monthly Overview", "Monthly Review"]
        for _nk in sorted(_PLANNER_NICHE_LABELS, key=len, reverse=True):
            if _planner_style.endswith("_" + _nk) or _planner_style == _nk:
                _planner_labels = _PLANNER_NICHE_LABELS[_nk]
                break

        _page_labels = {
            Category.WORKSHEET:             ["Practice Pages", "Answer Key Included"],
            Category.PLANNER:               _planner_labels,
            Category.SPREADSHEET:           ["Monthly Budget Sheet", "Daily Tracker Tab"],
            Category.SOCIAL_MEDIA_TEMPLATE: ["70 PNG Templates", "IG + TikTok + Pinterest + FB"],
            Category.RESUME_TEMPLATE:       ["DOCX + PDF + Cover Letter", "References + Guide Included"],
        }
        # PLANNER 05: 왼=Daily Log(니치 특화 섹션), 오=Monthly Review(Goals/LifeBalance)
        # Why: Daily Log는 니치별 섹션 가장 풍부 → 니치 라벨(Prayer Journal/Focus Sessions)과 실제 일치
        _wi_paths = ([art_path, converted_paths[1]] if product.category == Category.PLANNER
                     and len(converted_paths) > 1
                     else converted_paths or [art_path])
        generate_whats_included(
            _wi_paths,
            product.sizes,
            path,
            page_labels=_page_labels.get(product.category),
            style=product.style,
            category=product.category,
        )
    mockup_paths.append(path)

    # ── Slot 6: Single large frame on wall (wall_art) / Page spread ──
    path = str(mockup_dir / "06_gallery_wall.jpg")
    if product.category == Category.WALL_ART:
        # 1프레임 크게 — 단독 작품 클로즈업 포커스
        _art_slot6 = _art_candidates[5:6] or _art_candidates[:1] or [art_path]
        _generate_wall_art_hanging_mockup(_art_slot6, path, n_frames=1, show_badges=False)
    elif product.category == Category.PLANNER and len(converted_paths) >= 2:
        # PLANNER 전용: Monthly+Daily 두 페이지 나란히 — 콘텐츠 다양성 직접 증명
        _generate_planner_spread_mockup(converted_paths, path, style=product.style)
    else:
        generate_multi_frame_mockup(converted_paths[:3] or [art_path], path)
    mockup_paths.append(path)

    # ── Slot 7: Three-frame wall scene (art 7-9) / Different lifestyle ──
    path = str(mockup_dir / "07_lifestyle_dark.jpg")
    if product.category == Category.WALL_ART:
        # art7-9로 다크 배경 3연작 벽 씬 — 다른 인테리어 무드 어필
        _candidates_7 = _art_candidates[6:9] or _art_candidates[:3] or [art_path]
        _generate_wall_art_hanging_mockup(_candidates_7, path, n_frames=3, show_badges=False,
                                          room_color="charcoal")
    elif is_digital:
        # 07: Daily Log(art_path) 마블 배경 — 02(나무)와 배경만 다르고 가장 내용 풍부한 페이지 유지
        # Why: Monthly 페이지는 하단 40%가 빈 칸 — Daily Log가 콘텐츠 밀도 3배 높음
        generate_flatlay_mockup(art_path, path, "marble", overlay, sub_overlay)
    else:
        generate_lifestyle_mockup(art_path, path, "charcoal")
    mockup_paths.append(path)

    # ── Slot 8: Size guide / Compatibility ──
    path = str(mockup_dir / "08_size_guide.jpg")
    generate_size_guide(product.sizes, path, category=product.category, art_path=art_path)
    mockup_paths.append(path)

    # ── Slot 9: 소셜 증명 목업 (★★★★★ 신뢰 시그널) ──
    # Why: 상위 1% 셀러는 별점/리뷰 이미지로 구매 전 불안 해소 -> 전환율 직결
    path = str(mockup_dir / "09_social_proof.jpg")
    _generate_social_proof_mockup(art_path, path, product.category, style=product.style)
    mockup_paths.append(path)

    # ── Slot 10: Brand CTA ──
    path = str(mockup_dir / "10_brand_cta.jpg")
    _generate_brand_cta(path, product.category, art_path=art_path, style=product.style or "")
    mockup_paths.append(path)

    # ── Video mockup (MP4) -- Etsy 검색 자동재생, 클릭률 30-50% 향상 ──
    try:
        from generator.video_mockup import generate_video_mockup, CATEGORY_FEATURES
        video_path = str(mockup_dir / "00_video_mockup.mp4")
        cat_features = CATEGORY_FEATURES.get(product.category.value, [])
        extra_pages = converted_paths[1:4] if len(converted_paths) > 1 else []
        ok = generate_video_mockup(
            art_path=art_path,
            output_path=video_path,
            category_features=cat_features,
            page_paths=extra_pages,
        )
        if ok:
            mockup_paths.insert(0, video_path)  # 첫 슬롯에 삽입
            logger.info("Video mockup generated: %s", video_path)
    except Exception as ve:
        logger.warning("Video mockup skipped (non-critical): %s", ve)

    logger.info("Generated %d mockups for product %s", len(mockup_paths), product.product_id)
    return mockup_paths


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    # Test with a dummy product
    print("Mockup generator ready. Use generate_all_mockups(product) to generate.")
